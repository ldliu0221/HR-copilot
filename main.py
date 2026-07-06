import base64
import csv
import hashlib
import hmac
import json
import os
import re
import secrets
import shutil
import sys
import traceback
import urllib.request
from datetime import datetime, timedelta, timezone
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse


APP_NAME = "AI 招聘流程 Copilot"
ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
RESUME_DIRS = [DATA_DIR / "简历", DATA_DIR / "resumes"]
SUPPORTED_RESUME_SUFFIXES = [".pdf", ".docx", ".txt", ".md"]
STATUS_DIR = DATA_DIR / "status"
CONFIG_DIR = DATA_DIR / "config"
EXPORT_DIR = DATA_DIR / "exports"
LEGACY_STATUS_FILE = ROOT / "状态同步" / "状态.xlsx"

CANDIDATES_FILE = STATUS_DIR / "candidates.json"
EVENTS_FILE = STATUS_DIR / "events.jsonl"
STATUS_CSV_FILE = STATUS_DIR / "状态.csv"
JOB_PROFILES_FILE = CONFIG_DIR / "job_profiles.json"
APP_CONFIG_FILE = CONFIG_DIR / "app_config.json"
USERS_FILE = CONFIG_DIR / "users.json"
SESSIONS_FILE = CONFIG_DIR / "sessions.json"
TRAINING_EXAMS_FILE = STATUS_DIR / "training_exams.json"
TRAINING_RESULTS_FILE = STATUS_DIR / "training_results.json"

PIPELINE = ["待沟通", "已沟通", "待群面", "已群面", "待提交群面作业", "已提交群面", "待最终面", "已最终面", "淘汰", "通过"]
AUTO_WORKFLOW_RULES = {
    "resume_pass_score": 70,
    "resume_watch_score": 55,
    "group_pass_score": 70,
    "assignment_pass_score": 70,
    "hr_pass_score": 70,
}
WORKFLOW_RULE_LABELS = {
    "resume_pass_score": "简历推进群面线",
    "resume_watch_score": "简历待观察线",
    "group_pass_score": "群面推进作业线",
    "assignment_pass_score": "作业推进终面线",
    "hr_pass_score": "HR面通过线",
}
STAGE_ALIASES = {
    "待二筛": "待群面",
    "待约面": "待最终面",
    "已约面": "待最终面",
    "已面试": "已最终面",
    "待反馈": "已最终面",
    "已淘汰": "淘汰",
    "已录用": "通过",
}
DEFAULT_RESUME_PARSE_PROMPT = (
    "你是招聘简历分析智能体。请只基于简历原文和岗位标准进行分析，不要编造简历中没有的信息。"
    "输出 JSON，字段包括 analysis_summary、education_analysis、skill_analysis、project_analysis、"
    "experience_analysis、risk_points、follow_up_questions、conclusion、confidence。"
)
DEFAULT_CANDIDATE_REVIEW_PROMPT = (
    "你是招聘初筛评审智能体。请根据岗位标准和候选人简历，评估候选人与岗位的适配度。"
    "输出 JSON，字段包括 suitability_score、suitability_level、fit_summary、matched_evidence、"
    "gap_risks、questions、decision、next_stage、reason。suitability_score 为 0-100。"
)
DEFAULT_JD_GENERATE_PROMPT = (
    "你是资深招聘专家和JD撰写助手。请根据输入生成专业、清晰、可发布的岗位JD，并同步抽取可用于简历评分的岗位标准。"
    "如果输入包含 knowledge/行业知识库/RAG上下文，必须把其中的业务术语、真实项目、行业要求融入职位描述、职责、要求、加分项和面试关注点，避免模板化。"
    "生成内容必须覆盖职位名称、工作地点、部门、工作类型、薪资范围、职位描述、岗位职责、任职要求、公司介绍、福利待遇、工作亮点。"
    "输出 JSON，字段包括 title、city、department、work_type、salary、summary、responsibilities、requirements、company_intro、benefits、highlights、bonus_points、interview_focus、skills、abilities、risk_keywords、pass_score。"
)
DEFAULT_NEXT_ACTION_PROMPT = "你是招聘流程推进智能体。请基于候选人状态输出 action、reason、message、suggested_stage、priority。"
DEFAULT_GROUP_SCORE_PROMPT = "你是群面评分智能体。请输出 group_score、dimension_scores、evaluation、decision、reason、risks。"
DEFAULT_INTERVIEW_PLAN_PROMPT = (
    "你是资深技术面试官和招聘面试设计专家。请根据候选人简历、岗位标准和匹配评分，输出个性化面试方案 JSON。"
    "字段包括 interview_goal、candidate_summary、focus_areas、question_sections、project_deep_dive、risk_checks、scoring_rubric、schedule、decision_signals、interviewer_notes。"
    "question_sections 每项包含 title、purpose、questions。"
)
DEFAULT_DAILY_REPORT_PROMPT = "你是招聘日报智能体。请输出 summary、risks、priorities、actions、bottlenecks。"
DEFAULT_TRAINING_EXAM_PROMPT = (
    "你是企业HR培训考试设计专家。请根据培训对象、考试主题、难度、技能点和说明，输出 JSON。"
    "字段包括 title、target、topic、difficulty、duration、total_score、pass_score、description、sections、questions、scoring_rules。"
    "questions 每项包含 id、type、score、stem、options、answer、analysis。"
)
DEFAULT_EXAM_GRADING_PROMPT = (
    "你是企业培训考试阅卷专家。请根据试卷、标准答案和考生答案输出 JSON，字段包括 score、total_score、status、details、comment、review_required。"
    "details 每项包含 id、score、reason。复杂语义题如果不确定请标记 review_required=true。"
)
DEFAULT_CANDIDATE_RESCORE_PROMPT = "你是招聘初筛助手，请输出 JSON，字段包含 score, strengths, risks, questions, suggestion。"


def now_iso():
    return datetime.now().astimezone().isoformat(timespec="seconds")


def ensure_dirs():
    for path in [DATA_DIR, STATUS_DIR, CONFIG_DIR, EXPORT_DIR, *RESUME_DIRS]:
        path.mkdir(parents=True, exist_ok=True)


def read_json(path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def append_event(event_type, candidate_id=None, detail=None, actor=None):
    EVENTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    event = {
        "time": now_iso(),
        "type": event_type,
        "candidate_id": candidate_id,
        "actor": actor or "",
        "detail": detail or {},
    }
    with EVENTS_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")


def candidate_id_from_file(path):
    raw = str(path.resolve()).encode("utf-8", "ignore")
    return hashlib.sha1(raw).hexdigest()[:12]


def clean_text(text):
    return re.sub(r"\n{3,}", "\n\n", text.replace("\x00", "")).strip()


def extract_pdf_text(path):
    errors = []
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            return clean_text("\n".join(page.extract_text() or "" for page in pdf.pages)), ""
    except Exception as exc:
        errors.append(f"pdfplumber: {exc}")
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        return clean_text("\n".join(page.extract_text() or "" for page in reader.pages)), ""
    except Exception as exc:
        errors.append(f"pypdf: {exc}")
    return "", "; ".join(errors)


def extract_docx_text(path):
    try:
        from docx import Document

        doc = Document(str(path))
        return clean_text("\n".join(p.text for p in doc.paragraphs if p.text.strip())), ""
    except Exception as exc:
        return "", str(exc)


def extract_resume_text(path):
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        text, err = extract_pdf_text(path)
        return text, err
    if suffix == ".docx":
        return extract_docx_text(path)
    if suffix in [".txt", ".md"]:
        try:
            return clean_text(path.read_text(encoding="utf-8", errors="ignore")), ""
        except Exception as exc:
            return "", str(exc)
    return "", "暂不支持该文件类型"


GENERIC_NAME_WORDS = {
    "个人简历", "简历", "基本信息", "教育背景", "教育经历", "专业技能", "项目经历", "实习经历",
    "工作经历", "校园经历", "自我评价", "求职意向", "个人信息", "联系方式",
}


def guess_from_filename(path):
    stem = Path(path).stem.strip()
    clean = re.sub(r"\s+", " ", stem)
    name = ""
    position = ""
    if clean.lower().startswith("51job_"):
        parts = [p.strip() for p in clean.split("_") if p.strip()]
        if len(parts) >= 2:
            name = parts[1]
        if len(parts) >= 3:
            position = re.split(r"[（(]", parts[2])[0].strip()
    if not name:
        bracket = re.search(r"】\s*([\u4e00-\u9fa5·]{2,12})", clean)
        if bracket:
            name = bracket.group(1)
    if not name:
        match = re.search(r"([\u4e00-\u9fa5·]{2,12})(?:[-_ ]|$)", clean)
        if match:
            name = match.group(1)
    if not position:
        bracket_pos = re.search(r"【([^】_]+)", clean)
        if bracket_pos:
            position = bracket_pos.group(1).split()[0]
    name = re.sub(r"[^\u4e00-\u9fa5·]", "", name).strip()
    if name in GENERIC_NAME_WORDS:
        name = ""
    return {"name": name or clean[:20], "position": position or "无"}


def fallback_resume_text(path, error):
    guessed = guess_from_filename(path)
    return clean_text(
        "\n".join([
            f"姓名：{guessed['name']}",
            f"求职意向：{guessed['position']}",
            f"简历文件：{Path(path).name}",
            "解析状态：仅从文件名导入，原文待解析",
            f"解析失败原因：{error}",
        ])
    )


def pick_name(text, fallback):
    guessed = guess_from_filename(fallback)
    if guessed.get("name") and ("51job_" in Path(fallback).stem or Path(fallback).stem.startswith("【")):
        return guessed["name"]
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    label_match = re.search(r"姓\s*名\s*[:：]\s*([\u4e00-\u9fa5]{2,5})(?=\s|年|性|$)", text)
    if label_match:
        name = label_match.group(1)
        if name not in GENERIC_NAME_WORDS:
            return name
    for i, line in enumerate(lines[:12]):
        if line in ["个人简历", "简历", "Resume", "RESUME", "Personal resume"]:
            continue
        cleaned = re.sub(r"\s+", "", line)
        if cleaned in GENERIC_NAME_WORDS:
            continue
        if 2 <= len(cleaned) <= 5 and re.fullmatch(r"[\u4e00-\u9fa5]+", cleaned):
            return cleaned
        if line in ["个人简历", "简历"] and i + 1 < len(lines):
            next_line = re.sub(r"\s+", "", lines[i + 1])
            if 2 <= len(next_line) <= 5 and re.fullmatch(r"[\u4e00-\u9fa5]+", next_line):
                return next_line
    stem = Path(fallback).stem
    stem_name = re.split(r"[-_ （(]", stem)[0]
    if 2 <= len(stem_name) <= 5 and re.fullmatch(r"[\u4e00-\u9fa5]+", stem_name) and stem_name not in GENERIC_NAME_WORDS:
        return stem_name
    return guessed.get("name") or stem_name or stem


def find_first(pattern, text):
    match = re.search(pattern, text, re.I)
    return match.group(0) if match else ""


def find_phone(text):
    compact = re.sub(r"(?<=\d)[\s-]+(?=\d)", "", text)
    return find_first(r"(?<!\d)1[3-9]\d{9}(?!\d)", compact)


def find_email(text):
    match = re.search(r"([A-Za-z0-9._%+\-\s]{2,80}@[A-Za-z0-9.\-\s]{2,80}\.[A-Za-z]{2,})", text)
    if not match:
        return ""
    return re.sub(r"\s+", "", match.group(1))


def find_labeled_value(text, labels, max_len=30):
    label_pattern = "|".join(labels)
    match = re.search(rf"(?:{label_pattern})\s*[:：]\s*([^\n\r]{{1,{max_len}}})", text)
    if not match:
        return ""
    value = re.split(r"\s{2,}|电话|邮箱|政治面貌|语言能力|籍贯|年龄|性别", match.group(1).strip())[0].strip()
    return value[:max_len]


def find_gender(text):
    match = re.search(r"性\s*别\s*[:：]\s*([男女])", text)
    return match.group(1) if match else ""


def parse_resume(path, text):
    phone = find_phone(text)
    email = find_email(text)
    gender = find_gender(text)
    expected_salary = find_labeled_value(text, ["期望薪资", "薪资要求", "期望月薪"])
    arrival_time = find_labeled_value(text, ["到岗时间", "最快到岗", "可到岗时间", "实习周期"])
    intent_match = re.search(r"(?:求职意向|应聘岗位|目标岗位)[:：]?\s*([^\n]+)", text)
    guessed = guess_from_filename(path)
    position = intent_match.group(1).strip()[:40] if intent_match else guessed.get("position", "无")
    edu = []
    for keyword in ["博士", "硕士", "本科", "大专", "计算机", "人工智能", "软件工程"]:
        if keyword in text and keyword not in edu:
            edu.append(keyword)
    skill_hits = extract_keywords(text)
    return {
        "id": candidate_id_from_file(path),
        "name": pick_name(text, path.name),
        "gender": gender or "无",
        "phone": phone or "无",
        "wechat": phone or "无",
        "email": email or "无",
        "position": position,
        "education": " / ".join(edu[:5]) if edu else "无",
        "skills": skill_hits,
        "expected_salary": expected_salary or "无",
        "arrival_time": arrival_time or "无",
        "source": "本地简历",
        "resume_file": str(path.relative_to(ROOT)),
        "resume_text": text,
        "parse_status": "parsed",
        "parse_error": "",
        "stage": "待沟通",
        "group_interview": {"status": "待定", "score": ""},
        "assignment": {"status": "待定", "score": ""},
        "hr_interview": {"status": "待定", "score": ""},
        "final_result": "待定",
        "match": {},
        "notes": "",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }


def extract_keywords(text):
    bank = [
        "Python", "Java", "C++", "JavaScript", "HTML", "CSS", "FastAPI", "Django", "Flask",
        "SpringBoot", "Spring AI", "MyBatis", "MySQL", "Redis", "RabbitMQ", "Prometheus", "Grafana",
        "ELK", "Maven", "WebSocket", "SSE", "SQL", "Linux", "Docker", "Git", "LLM", "Prompt",
        "RAG", "MCP", "LangChain", "Milvus", "向量化", "Rerank", "Agent", "机器学习", "深度学习",
        "NLP", "PyTorch", "数据结构", "算法", "接口", "部署",
    ]
    lower = text.lower()
    hits = []
    for word in bank:
        if word.lower() in lower and word not in hits:
            hits.append(word)
    return hits


def default_job_profiles():
    return [
        {
            "id": "ai-intern",
            "name": "AI算法实习生",
            "description": "面向 AI/LLM/算法方向实习生，关注工程落地、模型调用、数据处理和学习能力。",
            "required_skills": [
                {"name": "Python", "weight": 18},
                {"name": "LLM", "weight": 12},
                {"name": "Prompt", "weight": 8},
                {"name": "RAG", "weight": 10},
                {"name": "FastAPI", "weight": 8},
                {"name": "Docker", "weight": 6},
                {"name": "Git", "weight": 5},
                {"name": "算法", "weight": 8},
            ],
            "abilities": [
                {"name": "项目交付经验", "keywords": ["交付", "上线", "联调", "部署"], "weight": 10},
                {"name": "数据与文档解析", "keywords": ["文档解析", "文本切分", "向量化", "数据"], "weight": 8},
                {"name": "沟通协作", "keywords": ["团队", "协助", "需求", "沟通"], "weight": 5},
            ],
            "risk_keywords": ["频繁跳槽", "无实习", "不接受实习", "无法到岗"],
            "pass_score": 70,
        }
    ]


def load_job_profiles():
    profiles = read_json(JOB_PROFILES_FILE, None)
    if not profiles:
        profiles = default_job_profiles()
        write_json(JOB_PROFILES_FILE, profiles)
    return profiles


def load_app_config():
    config = read_json(APP_CONFIG_FILE, None)
    if not config:
        config = {
            "llm_enabled": False,
            "llm_base_url": "https://api.openai.com/v1/chat/completions",
            "llm_model": "gpt-4o-mini",
            "llm_api_key": "",
            "llm_api_key_env": "OPENAI_API_KEY",
            "reminder_hours": 24,
            "manager_register_code": "MANAGER2026",
            "resume_parse_prompt": DEFAULT_RESUME_PARSE_PROMPT,
            "candidate_review_prompt": DEFAULT_CANDIDATE_REVIEW_PROMPT,
            "jd_generate_prompt": DEFAULT_JD_GENERATE_PROMPT,
            "next_action_prompt": DEFAULT_NEXT_ACTION_PROMPT,
            "group_score_prompt": DEFAULT_GROUP_SCORE_PROMPT,
            "interview_plan_prompt": DEFAULT_INTERVIEW_PLAN_PROMPT,
            "daily_report_prompt": DEFAULT_DAILY_REPORT_PROMPT,
            "training_exam_prompt": DEFAULT_TRAINING_EXAM_PROMPT,
            "exam_grading_prompt": DEFAULT_EXAM_GRADING_PROMPT,
            "candidate_rescore_prompt": DEFAULT_CANDIDATE_RESCORE_PROMPT,
            "ui_theme": "light",
            "ui_primary_color": "#126b61",
            "ui_font_size": 14,
            "ui_font_family": "Microsoft YaHei",
            "workflow_rules": AUTO_WORKFLOW_RULES,
            "mask_contact_for_hr": False,
            "data_retention_days": 365,
        }
        write_json(APP_CONFIG_FILE, config)
    changed = False
    defaults = {
        "llm_enabled": False,
        "llm_base_url": "https://api.openai.com/v1/chat/completions",
        "llm_model": "gpt-4o-mini",
        "llm_api_key": "",
        "llm_api_key_env": "OPENAI_API_KEY",
        "reminder_hours": 24,
        "manager_register_code": "MANAGER2026",
        "resume_parse_prompt": DEFAULT_RESUME_PARSE_PROMPT,
        "candidate_review_prompt": DEFAULT_CANDIDATE_REVIEW_PROMPT,
        "jd_generate_prompt": DEFAULT_JD_GENERATE_PROMPT,
        "next_action_prompt": DEFAULT_NEXT_ACTION_PROMPT,
        "group_score_prompt": DEFAULT_GROUP_SCORE_PROMPT,
        "interview_plan_prompt": DEFAULT_INTERVIEW_PLAN_PROMPT,
        "daily_report_prompt": DEFAULT_DAILY_REPORT_PROMPT,
        "training_exam_prompt": DEFAULT_TRAINING_EXAM_PROMPT,
        "exam_grading_prompt": DEFAULT_EXAM_GRADING_PROMPT,
        "candidate_rescore_prompt": DEFAULT_CANDIDATE_RESCORE_PROMPT,
        "ui_theme": "light",
        "ui_primary_color": "#126b61",
        "ui_font_size": 14,
        "ui_font_family": "Microsoft YaHei",
        "workflow_rules": AUTO_WORKFLOW_RULES,
        "mask_contact_for_hr": False,
        "data_retention_days": 365,
    }
    for key, value in defaults.items():
        if key not in config:
            config[key] = value
            changed = True
    if changed:
        write_json(APP_CONFIG_FILE, config)
    return config


def public_app_config():
    config = load_app_config()
    ensure_model_configs(config)
    ensure_collection_sources(config)
    safe = {k: v for k, v in config.items() if k != "llm_api_key"}
    safe["has_llm_api_key"] = bool(config.get("llm_api_key"))
    safe["model_configs"] = [public_model_config(m) for m in config.get("model_configs", [])]
    safe["collection_sources"] = [public_collection_source(s) for s in config.get("collection_sources", [])]
    return safe


def save_app_config_from_payload(payload):
    config = load_app_config()
    ensure_model_configs(config)
    ensure_collection_sources(config)
    for key in ["llm_enabled", "llm_base_url", "llm_model", "llm_api_key_env", "manager_register_code", "resume_parse_prompt", "candidate_review_prompt", "jd_generate_prompt", "next_action_prompt", "group_score_prompt", "interview_plan_prompt", "daily_report_prompt", "training_exam_prompt", "exam_grading_prompt", "candidate_rescore_prompt", "ui_theme", "ui_primary_color", "ui_font_family"]:
        if key in payload:
            config[key] = payload[key]
    if "ui_font_size" in payload:
        try:
            config["ui_font_size"] = max(12, min(18, int(payload["ui_font_size"])))
        except Exception:
            config["ui_font_size"] = 14
    if "reminder_hours" in payload:
        try:
            config["reminder_hours"] = int(payload["reminder_hours"])
        except Exception:
            config["reminder_hours"] = 24
    if "data_retention_days" in payload:
        try:
            config["data_retention_days"] = max(30, int(payload["data_retention_days"]))
        except Exception:
            config["data_retention_days"] = 365
    if "mask_contact_for_hr" in payload:
        config["mask_contact_for_hr"] = bool(payload.get("mask_contact_for_hr"))
    if isinstance(payload.get("workflow_rules"), dict):
        config["workflow_rules"] = normalize_workflow_rules(payload.get("workflow_rules"))
    if payload.get("llm_api_key"):
        config["llm_api_key"] = payload["llm_api_key"]
    if payload.get("clear_llm_api_key"):
        config["llm_api_key"] = ""
    if isinstance(payload.get("model_configs"), list):
        models = []
        for item in payload["model_configs"]:
            model = {
                "id": item.get("id") or secrets.token_hex(6),
                "name": item.get("name") or item.get("model") or "未命名模型",
                "model": item.get("model") or "",
                "base_url": item.get("base_url") or "",
                "api_key": item.get("api_key") or "",
                "api_key_env": item.get("api_key_env") or "OPENAI_API_KEY",
                "temperature": float(item.get("temperature", 0.2) or 0.2),
                "timeout": int(item.get("timeout", 120) or 120),
                "thinking": item.get("thinking") or "OpenAI",
                "source": item.get("source") or "自定义",
                "status": item.get("status") or "未测试",
                "last_test": item.get("last_test", ""),
                "last_test_ok": item.get("last_test_ok", None),
                "last_test_message": item.get("last_test_message", ""),
                "available_models": item.get("available_models", []),
                "model_found": item.get("model_found", None),
                "models_error": item.get("models_error", ""),
            }
            old = find_model_config(config.get("model_configs", []), model["id"])
            if old and not model["api_key"]:
                model["api_key"] = old.get("api_key", "")
            models.append(model)
        config["model_configs"] = models
    if payload.get("active_model_id"):
        config["active_model_id"] = payload["active_model_id"]
    if isinstance(payload.get("collection_sources"), list):
        sources = []
        old_sources = {s.get("id"): s for s in config.get("collection_sources", [])}
        for item in payload["collection_sources"]:
            sid = str(item.get("id") or secrets.token_hex(4)).strip()
            source = {
                "id": sid,
                "name": item.get("name") or sid,
                "enabled": bool(item.get("enabled", True)),
                "api_url": item.get("api_url") or "",
                "method": (item.get("method") or "POST").upper(),
                "api_key": item.get("api_key") or "",
                "headers": item.get("headers") or "",
                "request_body": item.get("request_body") or "{}",
            }
            old = old_sources.get(sid)
            if old and not source["api_key"]:
                source["api_key"] = old.get("api_key", "")
            sources.append(source)
        config["collection_sources"] = sources
    sync_active_model_to_legacy(config)
    write_json(APP_CONFIG_FILE, config)
    return public_app_config()


def normalize_workflow_rules(rules=None):
    source = {**AUTO_WORKFLOW_RULES, **(rules or {})}
    normalized = {}
    for key, default in AUTO_WORKFLOW_RULES.items():
        try:
            normalized[key] = max(0, min(100, int(float(source.get(key, default)))))
        except Exception:
            normalized[key] = default
    return normalized


def load_workflow_rules():
    config = load_app_config()
    rules = normalize_workflow_rules(config.get("workflow_rules"))
    if rules != config.get("workflow_rules"):
        config["workflow_rules"] = rules
        write_json(APP_CONFIG_FILE, config)
    return rules


def default_model_config(config):
    return {
        "id": "default",
        "name": config.get("llm_model") or "默认模型",
        "model": config.get("llm_model", ""),
        "base_url": config.get("llm_base_url", ""),
        "api_key": config.get("llm_api_key", ""),
        "api_key_env": config.get("llm_api_key_env", "OPENAI_API_KEY"),
        "temperature": 0.2,
        "timeout": 120,
        "thinking": "OpenAI",
        "source": "环境配置",
        "status": "未测试",
        "last_test": "",
    }


def ensure_model_configs(config):
    if not config.get("model_configs"):
        config["model_configs"] = [default_model_config(config)]
        config["active_model_id"] = config["model_configs"][0]["id"]
        write_json(APP_CONFIG_FILE, config)
    if not config.get("active_model_id") and config.get("model_configs"):
        config["active_model_id"] = config["model_configs"][0]["id"]


def public_model_config(model):
    safe = {k: v for k, v in model.items() if k != "api_key"}
    safe["has_api_key"] = bool(model.get("api_key"))
    return safe


def find_model_config(models, model_id):
    return next((m for m in models if m.get("id") == model_id), None)


def active_model_config(config=None):
    config = config or load_app_config()
    ensure_model_configs(config)
    return find_model_config(config.get("model_configs", []), config.get("active_model_id")) or config["model_configs"][0]


def sync_active_model_to_legacy(config):
    model = active_model_config(config)
    config["llm_model"] = model.get("model", config.get("llm_model", ""))
    config["llm_base_url"] = model.get("base_url", config.get("llm_base_url", ""))
    config["llm_api_key_env"] = model.get("api_key_env", config.get("llm_api_key_env", "OPENAI_API_KEY"))
    config["llm_api_key"] = model.get("api_key", config.get("llm_api_key", ""))


def load_users():
    return read_json(USERS_FILE, [])


def save_users(users):
    write_json(USERS_FILE, users)


def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000).hex()
    return salt, digest


def public_user(user):
    return {
        "id": user.get("id"),
        "username": user.get("username"),
        "role": user.get("role"),
        "display_name": user.get("display_name") or user.get("username"),
        "avatar_url": user.get("avatar_url", ""),
        "avatar_color": user.get("avatar_color", "#126b61"),
        "department": user.get("department", ""),
        "title": user.get("title", ""),
        "phone": user.get("phone", ""),
        "email": user.get("email", ""),
        "bio": user.get("bio", ""),
        "collection_sources": [public_collection_source(s) for s in normalize_user_collection_sources(user)],
        "created_at": user.get("created_at"),
        "last_login": user.get("last_login", ""),
    }


def create_user(username, password, role="hr", display_name="", manager_code=""):
    username = normalize_legacy_status(username)
    role = role if role in ["hr", "manager"] else "hr"
    if not username or len(username) < 2:
        raise ValueError("用户名至少 2 个字符")
    if not password or len(password) < 6:
        raise ValueError("密码至少 6 位")
    if role == "manager" and manager_code != load_app_config().get("manager_register_code"):
        raise ValueError("管理者注册码不正确")
    users = load_users()
    if any(u.get("username") == username for u in users):
        raise ValueError("用户名已存在")
    salt, digest = hash_password(password)
    user = {
        "id": secrets.token_hex(8),
        "username": username,
        "display_name": display_name or username,
        "role": role,
        "salt": salt,
        "password_hash": digest,
        "created_at": now_iso(),
        "last_login": "",
    }
    users.append(user)
    save_users(users)
    append_event("user_registered", None, {"username": username, "role": role}, actor=username)
    return public_user(user)


def verify_user(username, password):
    for user in load_users():
        if user.get("username") == username:
            _, digest = hash_password(password, user.get("salt"))
            if hmac.compare_digest(digest, user.get("password_hash", "")):
                user["last_login"] = now_iso()
                users = load_users()
                for item in users:
                    if item.get("id") == user.get("id"):
                        item["last_login"] = user["last_login"]
                save_users(users)
                return user
    return None


def update_user_account(user_id, payload):
    users = load_users()
    idx = next((i for i, item in enumerate(users) if item.get("id") == user_id), None)
    if idx is None:
        raise ValueError("用户不存在")
    user = users[idx]
    username = normalize_legacy_status(payload.get("username", user.get("username", ""))).strip()
    if not username or len(username) < 2:
        raise ValueError("用户名至少 2 个字符")
    if any(item.get("id") != user_id and item.get("username") == username for item in users):
        raise ValueError("用户名已存在")
    user["username"] = username
    for key in ["display_name", "avatar_url", "avatar_color", "department", "title", "phone", "email", "bio"]:
        if key in payload:
            user[key] = str(payload.get(key) or "").strip()
    if isinstance(payload.get("collection_sources"), list):
        old_sources = {s.get("id"): s for s in normalize_user_collection_sources(user)}
        sources = []
        for item in payload["collection_sources"]:
            sid = str(item.get("id") or secrets.token_hex(4)).strip()
            source = {
                "id": sid,
                "name": item.get("name") or sid,
                "enabled": bool(item.get("enabled", True)),
                "api_url": item.get("api_url") or "",
                "method": (item.get("method") or "POST").upper(),
                "api_key": item.get("api_key") or "",
                "headers": item.get("headers") or "",
                "request_body": item.get("request_body") or "{}",
            }
            old = old_sources.get(sid)
            if old and not source["api_key"]:
                source["api_key"] = old.get("api_key", "")
            sources.append(source)
        user["collection_sources"] = normalize_collection_sources(sources)
    old_password = payload.get("old_password", "")
    new_password = payload.get("new_password", "")
    if new_password:
        if len(new_password) < 6:
            raise ValueError("新密码至少 6 位")
        _, old_digest = hash_password(old_password, user.get("salt"))
        if not hmac.compare_digest(old_digest, user.get("password_hash", "")):
            raise ValueError("原密码不正确")
        salt, digest = hash_password(new_password)
        user["salt"] = salt
        user["password_hash"] = digest
    users[idx] = user
    save_users(users)
    append_event("account_updated", None, {"username": user.get("username")}, actor=user.get("username"))
    return public_user(user)


def load_sessions():
    return read_json(SESSIONS_FILE, {})


def save_sessions(sessions):
    write_json(SESSIONS_FILE, sessions)


def create_session(user):
    token = secrets.token_urlsafe(32)
    sessions = load_sessions()
    sessions[token] = {"user_id": user.get("id"), "created_at": now_iso()}
    save_sessions(sessions)
    append_event("user_login", None, {"username": user.get("username")}, actor=user.get("username"))
    return token


def clear_session(token):
    sessions = load_sessions()
    if token in sessions:
        del sessions[token]
        save_sessions(sessions)


def get_user_by_session(token):
    if not token:
        return None
    session = load_sessions().get(token)
    if not session:
        return None
    for user in load_users():
        if user.get("id") == session.get("user_id"):
            return user
    return None


def read_events():
    if not EVENTS_FILE.exists():
        return []
    events = []
    with EVENTS_FILE.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                events.append(json.loads(line))
            except Exception:
                pass
    return events


def is_today(iso_text):
    try:
        return datetime.fromisoformat(iso_text).astimezone().date() == datetime.now().astimezone().date()
    except Exception:
        return False


def manager_overview():
    users = [public_user(u) for u in load_users()]
    events = read_events()
    by_user = {u["username"]: {"today_actions": 0, "today_resumes": 0, "last_action": ""} for u in users}
    resume_ops = {"candidate_created", "candidate_scored", "agent_review", "resume_analyzed", "group_scored", "candidate_updated", "score_all"}
    for event in events:
        actor = event.get("actor") or event.get("detail", {}).get("actor") or ""
        if not actor:
            continue
        item = by_user.setdefault(actor, {"today_actions": 0, "today_resumes": 0, "last_action": ""})
        if is_today(event.get("time", "")):
            item["today_actions"] += 1
            if event.get("type") in resume_ops:
                item["today_resumes"] += 1
        item["last_action"] = max(item.get("last_action", ""), event.get("time", ""))
    for user in users:
        user.update(by_user.get(user["username"], {"today_actions": 0, "today_resumes": 0, "last_action": ""}))
    return {"users": users, "events_today": [e for e in events if is_today(e.get("time", ""))][-80:]}


def normalize_legacy_status(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_stage(stage):
    stage = normalize_legacy_status(stage)
    if not stage:
        return "待沟通"
    return STAGE_ALIASES.get(stage, stage if stage in PIPELINE else "待沟通")


def import_legacy_status_rows():
    if not LEGACY_STATUS_FILE.exists():
        return []
    try:
        import openpyxl

        wb = openpyxl.load_workbook(LEGACY_STATUS_FILE, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [normalize_legacy_status(c.value) for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            item = {headers[i]: normalize_legacy_status(row[i]) for i in range(min(len(headers), len(row)))}
            rows.append(item)
        return rows
    except Exception:
        return []


def merge_legacy_status(candidates):
    by_name = {c.get("name"): c for c in candidates if c.get("name")}
    changed = False
    for row in import_legacy_status_rows():
        name = row.get("姓名")
        if not name:
            continue
        c = by_name.get(name)
        if not c:
            c = {
                "id": hashlib.sha1(name.encode("utf-8")).hexdigest()[:12],
                "name": name,
                "position": "无",
                "source": "状态表",
                "resume_file": "",
                "resume_text": "",
                "skills": [],
                "stage": "待沟通",
                "created_at": now_iso(),
            }
            candidates.append(c)
            by_name[name] = c
        c["gender"] = row.get("性别", c.get("gender", "无")) or "无"
        c["graduation_time"] = row.get("毕业时间", c.get("graduation_time", "无")) or "无"
        c["group_interview"] = {"status": row.get("群面状态", ""), "score": row.get("群面分数", "")}
        c["assignment"] = {"status": row.get("作业状态", ""), "score": row.get("作业分数", "")}
        c["hr_interview"] = {"status": row.get("HR面状态", ""), "score": row.get("HR面分数", "")}
        c["final_result"] = row.get("最终结果", c.get("final_result", "待定"))
        c["updated_at"] = now_iso()
        changed = True
    return changed


def load_candidates():
    candidates = read_json(CANDIDATES_FILE, [])
    for c in candidates:
        c["stage"] = normalize_stage(c.get("stage"))
    if merge_legacy_status(candidates):
        save_candidates(candidates)
    return candidates


def save_candidates(candidates):
    for c in candidates:
        c["stage"] = normalize_stage(c.get("stage"))
    write_json(CANDIDATES_FILE, candidates)
    export_candidates_csv(candidates)
    export_status_csv(candidates)


def load_training_exams():
    exams = read_json(TRAINING_EXAMS_FILE, [])
    return exams if isinstance(exams, list) else []


def save_training_exams(exams):
    write_json(TRAINING_EXAMS_FILE, exams)


def load_training_results():
    results = read_json(TRAINING_RESULTS_FILE, [])
    return results if isinstance(results, list) else []


def save_training_results(results):
    write_json(TRAINING_RESULTS_FILE, results)


def training_summary():
    exams = load_training_exams()
    results = load_training_results()
    avg = 0
    if results:
        avg = round(sum(float(r.get("score", 0) or 0) for r in results) / len(results), 1)
    return {
        "exam_count": len(exams),
        "result_count": len(results),
        "avg_score": avg,
        "pass_rate": round(sum(1 for r in results if float(r.get("score", 0) or 0) >= float(r.get("pass_score", 60) or 60)) * 100 / max(1, len(results)), 1),
    }


def default_training_exam(payload):
    title = payload.get("title") or "新员工业务能力测评"
    target = payload.get("target") or "候选人/新员工"
    topic = payload.get("topic") or "岗位基础能力、业务理解和综合素质"
    difficulty = payload.get("difficulty") or "中等"
    total_score = int(payload.get("total_score") or 100)
    pass_score = int(payload.get("pass_score") or 60)
    skills = [x.strip() for x in re.split(r"[,，、\n]+", payload.get("skills", "")) if x.strip()]
    if not skills:
        skills = ["岗位认知", "专业技能", "问题分析", "沟通协作"]
    questions = []
    qid = 1
    for skill in skills[:4]:
        questions.append({
            "id": f"Q{qid}",
            "type": "单选题",
            "score": 5,
            "stem": f"关于{skill}，以下哪一项最能体现岗位所需能力？",
            "options": ["能结合业务目标说明方法选择", "只关注工具名称", "完全依赖他人安排", "忽略结果验证"],
            "answer": "能结合业务目标说明方法选择",
            "analysis": f"该题考察{skill}是否能落到业务目标、执行方法和结果验证。",
        })
        qid += 1
    questions.extend([
        {
            "id": f"Q{qid}",
            "type": "简答题",
            "score": 20,
            "stem": f"请结合{topic}，说明你会如何拆解一个真实工作任务，并识别关键风险。",
            "answer": "参考要点：目标澄清、信息收集、方案拆解、风险识别、里程碑和复盘。",
            "analysis": "重点看结构化表达、业务理解和风险意识。",
        },
        {
            "id": f"Q{qid + 1}",
            "type": "案例题",
            "score": 30,
            "stem": f"假设你负责推进一个{target}相关项目，时间紧、信息不完整，你会如何安排优先级并与团队协作？",
            "answer": "参考要点：明确目标、确定最小可交付、同步干系人、记录假设、及时复盘。",
            "analysis": "重点看项目推进、沟通协作和压力场景下的判断。",
        },
    ])
    return {
        "id": hashlib.sha1((title + now_iso()).encode("utf-8")).hexdigest()[:12],
        "title": title,
        "target": target,
        "topic": topic,
        "difficulty": difficulty,
        "duration": int(payload.get("duration") or 60),
        "total_score": total_score,
        "pass_score": pass_score,
        "description": payload.get("description") or f"面向{target}的{topic}测评，用于培训后验收和岗位胜任力判断。",
        "sections": [
            {"name": "基础认知", "score": 20, "focus": "岗位知识、基本概念、工具理解"},
            {"name": "能力应用", "score": 30, "focus": "任务拆解、方法选择、结果验证"},
            {"name": "案例综合", "score": 50, "focus": "业务场景、协作推进、风险控制"},
        ],
        "questions": questions,
        "scoring_rules": [
            "选择题按标准答案计分。",
            "简答题重点看结构完整度、关键风险识别和表达清晰度。",
            "案例题重点看目标拆解、优先级判断、沟通协作和复盘意识。",
        ],
        "created_at": now_iso(),
        "method": "本地规则生成",
    }


def generate_training_exam(payload):
    config = load_app_config()
    prompt = config.get("training_exam_prompt") or DEFAULT_TRAINING_EXAM_PROMPT
    llm = call_llm_json(prompt, payload, temperature=0.25, timeout=35, cap_timeout=True)
    exam = llm if not llm.get("error") else default_training_exam(payload)
    exam.setdefault("id", hashlib.sha1(((exam.get("title") or payload.get("title") or "exam") + now_iso()).encode("utf-8")).hexdigest()[:12])
    exam.setdefault("created_at", now_iso())
    exam.setdefault("method", "LLM智能生成" if not llm.get("error") else "本地规则生成")
    exam.setdefault("title", payload.get("title") or "新员工业务能力测评")
    exam.setdefault("total_score", int(payload.get("total_score") or 100))
    exam.setdefault("pass_score", int(payload.get("pass_score") or 60))
    if llm.get("error"):
        exam["llm_error"] = llm.get("error")
    exams = load_training_exams()
    exams = [e for e in exams if e.get("id") != exam["id"]]
    exams.insert(0, exam)
    save_training_exams(exams)
    return exam


def upsert_training_result(payload, actor=""):
    results = load_training_results()
    rid = payload.get("id") or hashlib.sha1(((payload.get("exam_id") or "") + (payload.get("student_name") or "") + now_iso()).encode("utf-8")).hexdigest()[:12]
    score = float(payload.get("score") or 0)
    pass_score = float(payload.get("pass_score") or 60)
    result = {
        "id": rid,
        "exam_id": payload.get("exam_id", ""),
        "exam_title": payload.get("exam_title", ""),
        "student_name": payload.get("student_name", ""),
        "department": payload.get("department", ""),
        "score": score,
        "total_score": float(payload.get("total_score") or 100),
        "pass_score": pass_score,
        "status": payload.get("status") or ("通过" if score >= pass_score else "未通过"),
        "submitted_at": payload.get("submitted_at") or now_iso(),
        "answers": payload.get("answers", ""),
        "comment": payload.get("comment", ""),
        "updated_by": actor,
        "updated_at": now_iso(),
    }
    results = [r for r in results if r.get("id") != rid]
    results.insert(0, result)
    save_training_results(results)
    return result


def local_grade_exam(exam, answer_text):
    answer_text = answer_text or ""
    total = float(exam.get("total_score") or 100)
    score = 0
    details = []
    for q in exam.get("questions", []):
        q_score = float(q.get("score") or 0)
        answer = str(q.get("answer") or "")
        stem = str(q.get("stem") or "")
        keywords = [x for x in re.split(r"[\s,，、；;。.\n]+", answer) if len(x) >= 2][:8]
        hit = sum(1 for k in keywords if k and k in answer_text)
        ratio = min(1, hit / max(1, min(4, len(keywords)))) if keywords else (0.7 if stem[:8] in answer_text else 0.4)
        earned = round(q_score * ratio, 1)
        score += earned
        details.append({"id": q.get("id"), "type": q.get("type"), "score": earned, "max_score": q_score, "reason": "命中参考答案关键要点" if hit else "未明显命中参考答案，建议人工复核"})
    score = max(0, min(total, round(score, 1)))
    return {
        "method": "本地标准答案关键词评分",
        "score": score,
        "total_score": total,
        "pass_score": float(exam.get("pass_score") or 60),
        "status": "通过" if score >= float(exam.get("pass_score") or 60) else "未通过",
        "details": details,
        "comment": "本地评分用于快速初筛；复杂语义题建议 HR 结合答题原文复核。",
        "review_required": any(d["score"] < d["max_score"] * 0.6 for d in details),
    }


def grade_training_answers(payload, actor=""):
    exams = load_training_exams()
    exam = next((e for e in exams if e.get("id") == payload.get("exam_id")), None)
    if not exam:
        raise ValueError("exam not found")
    answer_text = payload.get("answers") or payload.get("comment") or ""
    config = load_app_config()
    llm = call_llm_json(
        config.get("exam_grading_prompt") or DEFAULT_EXAM_GRADING_PROMPT,
        {"exam": exam, "answer_text": answer_text[:12000]},
        temperature=0.1,
        timeout=35,
        cap_timeout=True,
    )
    graded = llm if not llm.get("error") else local_grade_exam(exam, answer_text)
    graded.setdefault("method", "LLM智能阅卷" if not llm.get("error") else "本地标准答案关键词评分")
    graded.setdefault("total_score", exam.get("total_score", 100))
    graded.setdefault("pass_score", exam.get("pass_score", 60))
    graded.setdefault("status", "通过" if float(graded.get("score", 0) or 0) >= float(exam.get("pass_score", 60) or 60) else "未通过")
    if llm.get("error"):
        graded["llm_error"] = llm.get("error")
    result_payload = {
        **payload,
        "exam_title": exam.get("title"),
        "score": graded.get("score", 0),
        "total_score": graded.get("total_score", exam.get("total_score", 100)),
        "pass_score": graded.get("pass_score", exam.get("pass_score", 60)),
        "status": graded.get("status"),
        "answers": answer_text,
        "comment": graded.get("comment") or payload.get("comment", ""),
    }
    result = upsert_training_result(result_payload, actor=actor)
    result["grading"] = graded
    results = load_training_results()
    for item in results:
        if item.get("id") == result.get("id"):
            item["grading"] = graded
            item["answers"] = answer_text
    save_training_results(results)
    return result


def number_or_none(value):
    if value in [None, ""]:
        return None
    try:
        return float(value)
    except Exception:
        return None


def workflow_next_action(stage):
    stage = normalize_stage(stage)
    table = {
        "待沟通": ("完成首轮沟通", "确认实习周期、到岗时间和岗位意向", 1),
        "已沟通": ("安排群面", "发送群面时间并同步面试官", 2),
        "待群面": ("等待群面评价", "群面结束后上传评价或面试记录", 1),
        "已群面": ("判断是否进入作业", "录入群面分数并生成作业安排", 1),
        "待提交群面作业": ("催交群面作业", "发送作业要求和截止时间", 2),
        "已提交群面": ("评审作业并约终面", "录入作业评分，满足阈值后安排终面", 1),
        "待最终面": ("安排最终面", "协调候选人与面试官时间", 2),
        "已最终面": ("确认录用结论", "录入 HR 面评价并输出最终结果", 1),
        "淘汰": ("归档并记录原因", "沉淀淘汰原因用于岗位复盘", 0),
        "通过": ("进入入职跟进", "确认 Offer、入职时间和材料清单", 2),
    }
    action, reason, days = table.get(stage, table["待沟通"])
    due = (datetime.now().astimezone() + timedelta(days=days)).isoformat(timespec="seconds") if days else ""
    return {"action": action, "reason": reason, "due_at": due}


def apply_workflow_automation(candidate, trigger="manual", rules=None):
    rules = normalize_workflow_rules(rules or load_workflow_rules())
    previous = normalize_stage(candidate.get("stage"))
    stage = previous
    reason = ""
    final_result = candidate.get("final_result", "待定") or "待定"
    match_score = number_or_none(candidate.get("match", {}).get("score"))
    group = candidate.get("group_interview") or {}
    assignment = candidate.get("assignment") or {}
    hr = candidate.get("hr_interview") or {}
    group_score = number_or_none(group.get("score"))
    assignment_score = number_or_none(assignment.get("score"))
    hr_score = number_or_none(hr.get("score"))
    group_status = group.get("status", "")
    assignment_status = assignment.get("status", "")
    hr_status = hr.get("status", "")

    if final_result in ["通过", "淘汰"]:
        stage = final_result
        reason = f"最终结果为{final_result}，同步到流程终态"
    elif hr_score is not None or hr_status in ["已完成", "已评分", "通过", "淘汰"]:
        if hr_status == "淘汰" or (hr_score is not None and hr_score < rules["hr_pass_score"]):
            stage, final_result, reason = "淘汰", "淘汰", "HR面未通过，自动归档为淘汰"
        elif hr_status == "通过" or (hr_score is not None and hr_score >= rules["hr_pass_score"]):
            stage, final_result, reason = "通过", "通过", "HR面通过，自动进入通过流程"
        else:
            stage, reason = "已最终面", "HR面已完成，等待最终结论"
    elif assignment_score is not None or assignment_status in ["已提交", "已完成", "已评分", "通过", "淘汰"]:
        if assignment_status == "淘汰" or (assignment_score is not None and assignment_score < rules["assignment_pass_score"]):
            stage, final_result, reason = "淘汰", "淘汰", "群面作业未达标，自动归档为淘汰"
        elif assignment_status in ["通过", "已评分", "已完成", "已提交"] or (assignment_score is not None and assignment_score >= rules["assignment_pass_score"]):
            stage, reason = "待最终面", "作业已通过，自动推进到待最终面"
    elif group_score is not None or group_status in ["已完成", "已评分", "通过", "淘汰"]:
        if group_status == "淘汰" or (group_score is not None and group_score < rules["group_pass_score"]):
            stage, final_result, reason = "淘汰", "淘汰", "群面评价未达标，自动归档为淘汰"
        elif group_status in ["通过", "已评分", "已完成"] or (group_score is not None and group_score >= rules["group_pass_score"]):
            stage, reason = "待提交群面作业", "群面评价通过，自动推进到作业环节"
    elif match_score is not None and previous in ["待沟通", "已沟通"]:
        if match_score >= rules["resume_pass_score"]:
            stage, reason = "待群面", "简历匹配分达到通过线，自动推进到待群面"
        elif match_score < rules["resume_watch_score"]:
            stage, reason = "待沟通", "简历匹配分偏低，保留在待沟通待确认"

    candidate["stage"] = normalize_stage(stage)
    candidate["final_result"] = final_result
    next_action = workflow_next_action(candidate["stage"])
    candidate["workflow"] = {
        **candidate.get("workflow", {}),
        "auto_enabled": True,
        "rules": rules,
        "last_trigger": trigger,
        "last_stage_before": previous,
        "last_stage_after": candidate["stage"],
        "last_reason": reason or "未触发阶段变化，仅刷新下一步安排",
        "next_action": next_action,
        "updated_at": now_iso(),
    }
    return previous != candidate["stage"]


def export_candidates_csv(candidates):
    path = EXPORT_DIR / "candidates.csv"
    fields = [
        "id", "name", "phone", "email", "position", "stage", "source", "resume_file",
        "match_score", "group_status", "group_score", "assignment_status", "assignment_score",
        "hr_status", "hr_score", "final_result", "job_intention", "salary_expectation",
        "availability", "location_preference", "stability_risk", "talent_pool",
        "talent_tags", "next_follow_time", "follow_up_priority", "updated_at",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for c in candidates:
            writer.writerow({
                "id": c.get("id", ""),
                "name": c.get("name", ""),
                "phone": c.get("phone", ""),
                "email": c.get("email", ""),
                "position": c.get("position", ""),
                "stage": c.get("stage", ""),
                "source": c.get("source", ""),
                "resume_file": c.get("resume_file", ""),
                "match_score": c.get("match", {}).get("score", ""),
                "group_status": c.get("group_interview", {}).get("status", ""),
                "group_score": c.get("group_interview", {}).get("score", ""),
                "assignment_status": c.get("assignment", {}).get("status", ""),
                "assignment_score": c.get("assignment", {}).get("score", ""),
                "hr_status": c.get("hr_interview", {}).get("status", ""),
                "hr_score": c.get("hr_interview", {}).get("score", ""),
                "final_result": c.get("final_result", ""),
                "job_intention": c.get("job_intention", ""),
                "salary_expectation": c.get("salary_expectation") or c.get("expected_salary", ""),
                "availability": c.get("availability") or c.get("arrival_time", ""),
                "location_preference": c.get("location_preference", ""),
                "stability_risk": c.get("stability_risk", ""),
                "talent_pool": "是" if c.get("talent_pool") else "否",
                "talent_tags": "、".join(c.get("talent_tags") or []),
                "next_follow_time": c.get("next_follow_time", ""),
                "follow_up_priority": c.get("follow_up_priority", ""),
                "updated_at": c.get("updated_at", ""),
            })


def export_status_csv(candidates):
    fields = ["姓名", "性别", "毕业时间", "群面状态", "群面分数", "作业状态", "作业分数", "HR面状态", "HR面分数", "最终结果", "当前阶段", "匹配分", "更新时间"]
    with STATUS_CSV_FILE.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for c in candidates:
            writer.writerow({
                "姓名": c.get("name", ""),
                "性别": c.get("gender", ""),
                "毕业时间": c.get("graduation_time", ""),
                "群面状态": c.get("group_interview", {}).get("status", ""),
                "群面分数": c.get("group_interview", {}).get("score", ""),
                "作业状态": c.get("assignment", {}).get("status", ""),
                "作业分数": c.get("assignment", {}).get("score", ""),
                "HR面状态": c.get("hr_interview", {}).get("status", ""),
                "HR面分数": c.get("hr_interview", {}).get("score", ""),
                "最终结果": c.get("final_result", ""),
                "当前阶段": c.get("stage", ""),
                "匹配分": c.get("match", {}).get("score", ""),
                "更新时间": c.get("updated_at", ""),
            })


def current_resume_files():
    files = []
    for folder in RESUME_DIRS:
        files.extend([
            path for path in folder.rglob("*")
            if path.is_file() and path.suffix.lower() in SUPPORTED_RESUME_SUFFIXES
        ])
    return files


PLATFORM_SOURCES = {
    "boss": "BOSS直聘",
    "liepin": "猎聘",
    "zhilian": "智联招聘",
    "51job": "前程无忧",
    "local": "本地上传",
}


def default_collection_sources():
    return [
        {"id": "boss", "name": "BOSS直聘", "enabled": True, "api_url": "", "method": "POST", "api_key": "", "headers": "", "request_body": "{}"},
        {"id": "liepin", "name": "猎聘", "enabled": True, "api_url": "", "method": "POST", "api_key": "", "headers": "", "request_body": "{}"},
        {"id": "zhilian", "name": "智联招聘", "enabled": True, "api_url": "", "method": "POST", "api_key": "", "headers": "", "request_body": "{}"},
        {"id": "51job", "name": "前程无忧", "enabled": True, "api_url": "", "method": "POST", "api_key": "", "headers": "", "request_body": "{}"},
        {"id": "local", "name": "本地上传", "enabled": True, "api_url": "", "method": "POST", "api_key": "", "headers": "", "request_body": "{}"},
    ]


def normalize_collection_sources(sources=None):
    defaults = {item["id"]: item for item in default_collection_sources()}
    if not isinstance(sources, list) or not sources:
        sources = default_collection_sources()
    merged = []
    seen = set()
    for item in sources:
        sid = str(item.get("id") or secrets.token_hex(4)).strip()
        base = defaults.get(sid, {})
        merged.append({
            **base,
            **item,
            "id": sid,
            "name": item.get("name") or base.get("name") or sid,
            "enabled": bool(item.get("enabled", True)),
            "method": (item.get("method") or "POST").upper(),
            "headers": item.get("headers") or "",
            "request_body": item.get("request_body") or "{}",
        })
        seen.add(sid)
    for sid, item in defaults.items():
        if sid not in seen:
            merged.append(item)
    return merged


def normalize_user_collection_sources(user):
    return normalize_collection_sources(user.get("collection_sources"))


def ensure_collection_sources(config):
    existing = config.get("collection_sources")
    if not isinstance(existing, list) or not existing:
        config["collection_sources"] = default_collection_sources()
        write_json(APP_CONFIG_FILE, config)
        return
    config["collection_sources"] = normalize_collection_sources(existing)


def public_collection_source(item):
    safe = {k: v for k, v in item.items() if k != "api_key"}
    safe["has_api_key"] = bool(item.get("api_key"))
    return safe


def collection_source_map(user=None):
    if user:
        return {item.get("id"): item for item in normalize_user_collection_sources(user)}
    config = load_app_config()
    ensure_collection_sources(config)
    return {item.get("id"): item for item in config.get("collection_sources", [])}


def platform_key(source, user=None):
    raw = (source or "local").strip().lower()
    configured = collection_source_map(user)
    if raw in configured:
        return raw
    if "boss" in raw or "boss直聘" in raw:
        return "boss"
    if "liepin" in raw or "猎聘" in raw:
        return "liepin"
    if "zhilian" in raw or "智联" in raw:
        return "zhilian"
    if "51job" in raw or "前程" in raw:
        return "51job"
    return "local"


def source_from_resume_path(path, user=None):
    text = str(path).lower()
    configured = collection_source_map(user)
    labels = {key: value.get("name", key) for key, value in configured.items()}
    labels.update(PLATFORM_SOURCES)
    for key, label in labels.items():
        if key in text:
            return label
    return "本地简历"


def safe_upload_name(name):
    name = Path(name or "resume.txt").name
    stem = re.sub(r"[^\w\u4e00-\u9fa5·.-]+", "_", Path(name).stem).strip("._") or "resume"
    suffix = Path(name).suffix.lower()
    if suffix not in SUPPORTED_RESUME_SUFFIXES:
        suffix = ".txt"
    return stem[:80] + suffix


def save_collected_resumes(payload, actor=None, user=None):
    ensure_dirs()
    key = platform_key(payload.get("source"), user)
    configured = collection_source_map(user)
    source = configured.get(key, {}).get("name") or PLATFORM_SOURCES.get(key, "本地上传")
    target_dir = RESUME_DIRS[0] / key
    target_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    skipped = []
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    for i, item in enumerate(payload.get("files") or []):
        name = safe_upload_name(item.get("name") or f"resume-{i + 1}.txt")
        suffix = Path(name).suffix.lower()
        if suffix not in SUPPORTED_RESUME_SUFFIXES:
            skipped.append({"name": name, "reason": "unsupported_type"})
            continue
        raw = item.get("content") or ""
        if "," in raw and raw.split(",", 1)[0].startswith("data:"):
            raw = raw.split(",", 1)[1]
        try:
            data = base64.b64decode(raw)
        except Exception as exc:
            skipped.append({"name": name, "reason": str(exc)})
            continue
        path = target_dir / f"{stamp}_{i + 1}_{name}"
        path.write_bytes(data)
        saved.append(str(path.relative_to(ROOT)))
    for i, item in enumerate(payload.get("texts") or []):
        text = (item.get("text") or "").strip()
        if not text:
            continue
        name = safe_upload_name(item.get("name") or f"{source}-候选人-{i + 1}.txt")
        path = target_dir / f"{stamp}_text_{i + 1}_{Path(name).stem}.txt"
        header = f"来源：{source}\n采集人：{actor or ''}\n采集时间：{now_iso()}\n\n"
        path.write_text(header + text, encoding="utf-8")
        saved.append(str(path.relative_to(ROOT)))
    result = scan_resumes(actor=actor, remove_missing=False, user=user)
    append_event("resume_collected", None, {"source": source, "saved": saved, "skipped": skipped}, actor=actor)
    return {"ok": True, "source": source, "saved": saved, "saved_count": len(saved), "skipped": skipped, "scan": result}


def collect_resumes_from_api(source_id, actor=None, user=None):
    sources = normalize_user_collection_sources(user or {})
    source = next((s for s in sources if s.get("id") == source_id), None)
    if not source:
        return {"ok": False, "error": "collection source not found"}
    if not source.get("enabled"):
        return {"ok": False, "error": "collection source disabled"}
    if not source.get("api_url"):
        return {"ok": False, "error": "api_url_empty", "message": "该来源还没有配置 API，可先使用本地文件或粘贴文本导入。"}
    headers = {"Content-Type": "application/json"}
    if source.get("api_key"):
        headers["Authorization"] = f"Bearer {source.get('api_key')}"
    try:
        custom_headers = json.loads(source.get("headers") or "{}")
        if isinstance(custom_headers, dict):
            headers.update({str(k): str(v) for k, v in custom_headers.items()})
    except Exception:
        pass
    try:
        request_body = json.loads(source.get("request_body") or "{}")
    except Exception:
        request_body = {}
    request_body.setdefault("source", source.get("id"))
    request_body.setdefault("requested_by", actor or "")
    data = json.dumps(request_body, ensure_ascii=False).encode("utf-8")
    method = (source.get("method") or "POST").upper()
    req = urllib.request.Request(source.get("api_url"), data=data if method != "GET" else None, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except Exception as exc:
        return {"ok": False, "source": source.get("name"), "error": str(exc)}
    payload = {"source": source.get("id"), "files": result.get("files") or [], "texts": result.get("texts") or []}
    for i, candidate in enumerate(result.get("candidates") or []):
        if isinstance(candidate, dict):
            text = candidate.get("resume_text") or candidate.get("text") or json.dumps(candidate, ensure_ascii=False, indent=2)
            name = candidate.get("name") or f"{source.get('name')}-候选人-{i + 1}"
            payload["texts"].append({"name": name + ".txt", "text": text})
    if not payload["files"] and not payload["texts"]:
        return {"ok": True, "source": source.get("name"), "saved_count": 0, "message": "API 已返回，但未识别到 files/texts/candidates 字段。", "raw": result}
    collected = save_collected_resumes(payload, actor=actor, user=user)
    collected["api_source"] = source.get("name")
    return collected


def scan_resumes(actor=None, remove_missing=False, user=None):
    ensure_dirs()
    candidates = load_candidates()
    by_id = {c["id"]: c for c in candidates}
    scanned = 0
    failed = []
    seen_ids = set()
    for path in current_resume_files():
            scanned += 1
            cid = candidate_id_from_file(path)
            seen_ids.add(cid)
            text, err = extract_resume_text(path)
            if not text:
                failed.append({"file": str(path), "error": err, "imported_from_filename": True})
                text = fallback_resume_text(path, err)
            parsed = parse_resume(path, text)
            parsed["source"] = source_from_resume_path(path, user)
            if err:
                parsed["parse_status"] = "filename_only"
                parsed["parse_error"] = err
                parsed["notes"] = (parsed.get("notes") or "") + "简历原文解析失败，已先按文件名导入，待安装 PDF/DOCX 解析依赖后重新解析。"
            if cid in by_id:
                existing = by_id[cid]
                if err and existing.get("parse_status") == "parsed" and not (existing.get("resume_text", "").startswith("姓名：") and "仅从文件名导入" in existing.get("resume_text", "")):
                    existing["parse_error"] = err
                    existing["updated_at"] = now_iso()
                    continue
                preserved = {
                    "stage": existing.get("stage"),
                    "notes": existing.get("notes"),
                    "group_interview": existing.get("group_interview"),
                    "assignment": existing.get("assignment"),
                    "hr_interview": existing.get("hr_interview"),
                    "final_result": existing.get("final_result"),
                    "match": existing.get("match"),
                    "created_at": existing.get("created_at"),
                }
                existing.update(parsed)
                for key, value in preserved.items():
                    if value not in [None, "", {}, []]:
                        existing[key] = value
                if err and existing.get("resume_text", "").startswith("姓名："):
                    existing["parse_status"] = "filename_only"
                    existing["parse_error"] = err
                existing["updated_at"] = now_iso()
            else:
                candidates.append(parsed)
                by_id[cid] = parsed
                append_event("candidate_created", cid, {"file": parsed["resume_file"]}, actor=actor)
    removed = []
    if remove_missing:
        kept = []
        for c in candidates:
            cid = c.get("id")
            if c.get("source") == "本地简历" and c.get("resume_file") and cid not in seen_ids:
                removed.append({"id": cid, "name": c.get("name", ""), "resume_file": c.get("resume_file", "")})
            else:
                kept.append(c)
        candidates = kept
    save_candidates(candidates)
    return {"scanned": scanned, "total": len(candidates), "failed": failed, "removed": removed, "removed_count": len(removed)}


def score_candidate(candidate, profile):
    text = (candidate.get("resume_text") or "") + "\n" + " ".join(candidate.get("skills") or [])
    lower = text.lower()
    max_score = 0
    score = 0
    matched = []
    missing = []
    for skill in profile.get("required_skills", []):
        name = skill.get("name", "")
        weight = float(skill.get("weight", 0) or 0)
        max_score += weight
        if name and name.lower() in lower:
            score += weight
            matched.append(name)
        else:
            missing.append(name)
    ability_hits = []
    for ability in profile.get("abilities", []):
        weight = float(ability.get("weight", 0) or 0)
        max_score += weight
        keywords = ability.get("keywords") or [ability.get("name", "")]
        hits = [k for k in keywords if k and k.lower() in lower]
        if hits:
            score += weight
            ability_hits.append(ability.get("name", ""))
    risk_hits = [k for k in profile.get("risk_keywords", []) if k and k.lower() in lower]
    normalized = round((score / max_score) * 100) if max_score else 0
    skill_total = sum(float(s.get("weight", 0) or 0) for s in profile.get("required_skills", []))
    skill_score = sum(float(s.get("weight", 0) or 0) for s in profile.get("required_skills", []) if s.get("name", "").lower() in lower)
    ability_total = sum(float(a.get("weight", 0) or 0) for a in profile.get("abilities", []))
    ability_score = 0
    for item in profile.get("abilities", []):
        keywords = item.get("keywords") or [item.get("name", "")]
        if any(k and k.lower() in lower for k in keywords):
            ability_score += float(item.get("weight", 0) or 0)
    risk_penalty = min(20, len(risk_hits) * 6)
    dimensions = [
        {"name": "技能匹配", "score": round(skill_score / skill_total * 100) if skill_total else 0, "evidence": matched[:6]},
        {"name": "能力信号", "score": round(ability_score / ability_total * 100) if ability_total else 0, "evidence": ability_hits[:6]},
        {"name": "风险控制", "score": max(0, 100 - risk_penalty), "evidence": risk_hits[:6]},
        {"name": "岗位相关", "score": min(100, normalized + (8 if matched and ability_hits else 0)), "evidence": [profile.get("name", "")]},
    ]
    questions = []
    if missing:
        questions.append("请候选人补充说明：" + "、".join(missing[:4]) + " 的实际项目经验。")
    questions.append("确认可实习周期、每周到岗天数、工作地点接受度和最快到岗时间。")
    if "LLM" in matched or "RAG" in matched:
        questions.append("追问最近一个 LLM/RAG 项目中的数据流、评测指标和上线问题。")
    return {
        "score": min(100, normalized),
        "profile_id": profile.get("id"),
        "profile_name": profile.get("name"),
        "matched_skills": matched,
        "missing_skills": missing,
        "abilities": ability_hits,
        "dimensions": dimensions,
        "risks": risk_hits or (["关键技能覆盖不足"] if normalized < profile.get("pass_score", 70) else []),
        "strengths": build_strengths(matched, ability_hits),
        "questions": questions,
        "suggestion": "建议推进群面" if normalized >= profile.get("pass_score", 70) else "建议补充追问后再决定",
        "scored_at": now_iso(),
        "method": "规则匹配，可在配置中接入大模型复评",
    }


def build_strengths(matched, abilities):
    items = []
    if matched:
        items.append("技能匹配：" + "、".join(matched[:6]))
    if abilities:
        items.append("能力信号：" + "、".join(abilities[:4]))
    if not items:
        items.append("简历中暂未识别到明显岗位匹配信号")
    return items


def try_llm_rescore(candidate, profile):
    config = load_app_config()
    if not config.get("llm_enabled"):
        return None
    model_cfg = active_model_config(config)
    api_key = get_llm_api_key(config, model_cfg)
    payload = {
        "model": model_cfg.get("model") or config.get("llm_model", "gpt-4o-mini"),
        "messages": [
            {"role": "system", "content": config.get("candidate_rescore_prompt") or DEFAULT_CANDIDATE_RESCORE_PROMPT},
            {"role": "user", "content": json.dumps({"job": profile, "resume": candidate.get("resume_text", "")[:8000]}, ensure_ascii=False)},
        ],
        "temperature": float(model_cfg.get("temperature", 0.2) or 0.2),
    }
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = urllib.request.Request(normalize_llm_endpoint(model_cfg.get("base_url") or config.get("llm_base_url")), data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=int(model_cfg.get("timeout", 120) or 120)) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        content = data["choices"][0]["message"]["content"]
        match = re.search(r"\{.*\}", content, re.S)
        return json.loads(match.group(0) if match else content)
    except Exception as exc:
        return {"error": str(exc)}


def call_llm_json(system_prompt, user_payload, temperature=0.2, timeout=25, cap_timeout=False):
    config = load_app_config()
    if not config.get("llm_enabled"):
        return {"error": "llm_disabled"}
    model_cfg = active_model_config(config)
    api_key = get_llm_api_key(config, model_cfg)
    payload = {
        "model": model_cfg.get("model") or config.get("llm_model", "gpt-4o-mini"),
        "messages": [
            {"role": "system", "content": system_prompt + " 只输出 JSON，不要输出 Markdown。"},
            {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
        ],
        "temperature": float(model_cfg.get("temperature", temperature) or temperature),
    }
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = urllib.request.Request(normalize_llm_endpoint(model_cfg.get("base_url") or config.get("llm_base_url")), data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
    try:
        request_timeout = int(model_cfg.get("timeout", timeout) or timeout)
        if cap_timeout:
            request_timeout = min(request_timeout, timeout)
        with urllib.request.urlopen(req, timeout=request_timeout) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        content = data["choices"][0]["message"]["content"]
        match = re.search(r"\{.*\}", content, re.S)
        return json.loads(match.group(0) if match else content)
    except Exception as exc:
        return {"error": str(exc)}


def normalize_list(value):
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    if value:
        return [x.strip() for x in re.split(r"[；;\n]+", str(value)) if x.strip()]
    return []


def jd_profile_from_result(result):
    skills = normalize_list(result.get("skills") or result.get("skill_tags"))
    abilities = normalize_list(result.get("abilities") or result.get("competencies"))
    if not skills:
        skills = ["沟通协作", "学习能力", "执行力"]
    if not abilities:
        abilities = ["业务理解", "项目推进", "跨部门协作"]
    skill_weight = max(5, round(70 / max(1, len(skills))))
    ability_weight = max(5, round(30 / max(1, len(abilities))))
    return {
        "name": result.get("title", "新岗位"),
        "description": result.get("summary", ""),
        "pass_score": int(result.get("pass_score", 70) or 70),
        "required_skills": [{"name": s, "weight": skill_weight} for s in skills[:10]],
        "abilities": [{"name": a, "keywords": [a], "weight": ability_weight} for a in abilities[:8]],
        "risk_keywords": normalize_list(result.get("risk_keywords"))[:10],
    }


def jd_text_from_result(result):
    sections = [
        ("岗位职责", normalize_list(result.get("responsibilities"))),
        ("任职要求", normalize_list(result.get("requirements"))),
        ("公司介绍", normalize_list(result.get("company_intro") or result.get("company"))),
        ("福利待遇", normalize_list(result.get("benefits"))),
        ("工作亮点", normalize_list(result.get("highlights") or result.get("work_highlights"))),
        ("加分项", normalize_list(result.get("bonus_points"))),
        ("面试关注点", normalize_list(result.get("interview_focus"))),
    ]
    lines = [
        f"职位名称：{result.get('title', '新岗位')}",
        f"工作地点：{result.get('city', '不限')}",
        f"部门：{result.get('department', '待定')}",
        f"工作类型：{result.get('work_type', '全职')}",
        f"薪资范围：{result.get('salary', '面议')}",
        "",
        "职位描述：",
        result.get("summary", "").strip(),
    ]
    for title, items in sections:
        if items:
            lines.extend(["", title + "：", *[f"{i + 1}. {item}" for i, item in enumerate(items)]])
    return "\n".join([line for line in lines if line is not None]).strip()


def local_generate_jd(payload):
    title = (payload.get("title") or "新岗位").strip()
    city = (payload.get("city") or "不限").strip()
    department = (payload.get("department") or "待定").strip()
    salary = (payload.get("salary") or "面议").strip()
    education = payload.get("education") or "本科"
    work_type = payload.get("work_type") or "全职"
    experience = (payload.get("experience") or "具备相关岗位经验").strip()
    headcount = payload.get("headcount") or 1
    skills = normalize_list(payload.get("skills")) or ["沟通协作", "学习能力", "执行力"]
    highlights = normalize_list(payload.get("highlights"))
    company = payload.get("company") or "我们是一家重视长期发展与人才培养的公司，持续为员工提供清晰的成长路径和开放协作的工作环境。"
    knowledge = (payload.get("knowledge") or "").strip()
    background = payload.get("background") or f"围绕{title}相关业务场景，完成需求分析、方案设计、开发交付与持续优化。"
    if knowledge:
        background = f"{background} 参考行业知识与业务上下文：{knowledge[:500]}"
    result = {
        "method": "本地JD模板兜底" + (" + 知识库上下文" if knowledge else ""),
        "title": title,
        "city": city,
        "department": department,
        "salary": salary,
        "work_type": work_type,
        "summary": f"我们正在寻找一位经验丰富的{title}加入{department}，负责{background}。候选人需要具备扎实的编程基础和项目经验，能够独立完成开发任务并与团队高效协作。",
        "responsibilities": [
            f"参与{title}产品或系统的设计、开发和优化，确保高性能、可扩展性和稳定性。",
            "使用" + "、".join(skills[:5]) + "等技术栈进行功能开发、接口联调和问题排查。",
            "参与需求分析、技术方案设计及代码评审，推动项目高质量交付。",
            "解决开发过程中的技术难题，优化系统架构，提升开发效率。",
            "与产品、测试等团队紧密合作，确保项目按时高质量完成。",
        ],
        "requirements": [
            f"{education}及以上学历，计算机或相关专业优先。",
            experience,
            "掌握" + "、".join(skills[:6]) + "等相关技能。",
            "具备清晰表达、问题分析和跨团队协作能力。",
        ],
        "company_intro": [company],
        "benefits": ["五险一金", "带薪年假", "弹性工作", "技术培训"] if not highlights else highlights,
        "highlights": highlights or ["核心项目实践机会", "完善的技术成长路径", "开放协作的团队氛围"],
        "bonus_points": ["有复杂系统开发、性能优化或团队协作项目经验优先。"] + (["能结合行业知识库中的业务术语、项目场景和目标用户做方案设计。"] if knowledge else []),
        "interview_focus": ["岗位动机", "项目经历真实性", "问题拆解能力", "稳定性与到岗时间"] + (["是否理解行业知识库/RAG上下文中的真实业务场景"] if knowledge else []),
        "skills": skills,
        "abilities": ["项目推进", "沟通协作", "问题分析", "结果复盘"],
        "risk_keywords": ["频繁跳槽", "到岗时间不确定", "项目描述不清晰"],
        "pass_score": 70,
    }
    result["jd_text"] = jd_text_from_result(result)
    result["profile"] = jd_profile_from_result(result)
    return result


def generate_job_description(payload):
    config = load_app_config()
    llm = call_llm_json(
        config.get("jd_generate_prompt") or DEFAULT_JD_GENERATE_PROMPT,
        payload,
        temperature=0.35,
        timeout=35,
        cap_timeout=True,
    )
    if llm.get("error"):
        return local_generate_jd(payload)
    llm["method"] = "LLM JD生成"
    llm.setdefault("title", payload.get("title") or "新岗位")
    llm.setdefault("summary", "")
    llm["jd_text"] = llm.get("jd_text") or jd_text_from_result(llm)
    llm["profile"] = jd_profile_from_result(llm)
    return llm


def get_llm_api_key(config, model_cfg=None):
    model_cfg = model_cfg or active_model_config(config)
    return model_cfg.get("api_key") or config.get("llm_api_key") or os.environ.get(model_cfg.get("api_key_env") or config.get("llm_api_key_env", "OPENAI_API_KEY"), "")


def normalize_llm_endpoint(base_url):
    url = (base_url or "").strip().rstrip("/")
    if not url:
        return "https://api.openai.com/v1/chat/completions"
    if url.endswith("/chat/completions"):
        return url
    if url.endswith("/v1"):
        return url + "/chat/completions"
    return url


def normalize_models_endpoint(base_url):
    url = (base_url or "").strip().rstrip("/")
    if not url:
        return "https://api.openai.com/v1/models"
    if url.endswith("/chat/completions"):
        url = url[: -len("/chat/completions")]
    if url.endswith("/v1"):
        return url + "/models"
    return url + "/models"


def test_model_config(model):
    api_key = model.get("api_key") or os.environ.get(model.get("api_key_env") or "OPENAI_API_KEY", "")
    payload = {
        "model": model.get("model", ""),
        "messages": [
            {"role": "system", "content": "????????????? JSON?"},
            {"role": "user", "content": "{\"task\":\"ping\"}"},
        ],
        "temperature": float(model.get("temperature", 0.2) or 0.2),
    }
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = urllib.request.Request(
        normalize_llm_endpoint(model.get("base_url")),
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=min(int(model.get("timeout", 30) or 30), 30)) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        available_models = []
        models_error = ""
        try:
            models_req = urllib.request.Request(normalize_models_endpoint(model.get("base_url")), headers=headers, method="GET")
            with urllib.request.urlopen(models_req, timeout=10) as resp:
                models_data = json.loads(resp.read().decode("utf-8"))
            available_models = [
                item.get("id") or item.get("model") or item.get("name")
                for item in models_data.get("data", [])
                if isinstance(item, dict)
            ]
        except Exception as exc:
            models_error = str(exc)
        target = model.get("model", "")
        model_found = (not available_models) or target in available_models
        message = "API ???" + ("?????" if model_found else "?????????")
        return {
            "ok": True,
            "status": "??",
            "message": message,
            "raw_message": content[:180],
            "available_models": available_models[:20],
            "models_error": models_error,
            "model_found": model_found,
            "time": now_iso(),
        }
    except Exception as exc:
        return {"ok": False, "status": "??", "message": str(exc), "time": now_iso()}


class RecruitingAgent:
    def __init__(self, profiles=None):
        self.profiles = profiles or load_job_profiles()

    def profile_for(self, candidate, profile_id=None):
        if profile_id:
            for profile in self.profiles:
                if profile.get("id") == profile_id:
                    return profile
        position = candidate.get("position", "")
        for profile in self.profiles:
            if profile.get("name") and profile.get("name") in position:
                return profile
        return self.profiles[0]

    def analyze_resume(self, candidate, profile_id=None):
        profile = self.profile_for(candidate, profile_id)
        config = load_app_config()
        llm = call_llm_json(
            config.get("resume_parse_prompt") or DEFAULT_RESUME_PARSE_PROMPT,
            {"job_profile": profile, "resume_text": candidate.get("resume_text", "")[:9000]},
        )
        if not llm.get("error"):
            llm["method"] = "LLM结构化解析"
            return llm
        text = candidate.get("resume_text", "")
        return {
            "method": "本地规则解析兜底",
            "analysis_summary": build_resume_summary(candidate),
            "education_analysis": f"学历信息：{candidate.get('education', '无')}",
            "skill_analysis": "技能信号：" + ("、".join((candidate.get("skills") or [])[:10]) or "无"),
            "project_analysis": "；".join(extract_section_lines(text, ["项目经历", "项目经验", "个人项目"], 4)) or "无",
            "experience_analysis": "；".join(extract_section_lines(text, ["实习经历", "工作经历"], 4)) or "无",
            "risk_points": candidate.get("match", {}).get("risks", []),
            "follow_up_questions": candidate.get("match", {}).get("questions", []),
            "conclusion": "基于本地规则生成分析；启用并连通大模型后可获得更完整的语义分析。",
            "confidence": "中",
            "llm_error": llm.get("error"),
        }

    def review_candidate(self, candidate, profile_id=None):
        profile = self.profile_for(candidate, profile_id)
        rule_score = score_candidate(candidate, profile)
        config = load_app_config()
        llm = call_llm_json(
            config.get("candidate_review_prompt") or DEFAULT_CANDIDATE_REVIEW_PROMPT,
            {"job_profile": profile, "rule_score": rule_score, "candidate": compact_candidate(candidate)},
        )
        if not llm.get("error"):
            if "score" not in llm and "suitability_score" in llm:
                llm["score"] = llm.get("suitability_score")
            llm.setdefault("score", rule_score.get("score", 0))
            llm.setdefault("suitability_score", llm.get("score", rule_score.get("score", 0)))
            llm.setdefault("suitability_level", suitability_level(llm.get("suitability_score", 0)))
            llm.setdefault("strengths", rule_score.get("strengths", []))
            llm.setdefault("risks", rule_score.get("risks", []))
            llm.setdefault("questions", rule_score.get("questions", []))
            llm.setdefault("decision", "建议推进群面" if float(llm.get("score", 0)) >= profile.get("pass_score", 70) else "建议补充追问")
            llm["method"] = "LLM评审"
            return llm
        return {
            "method": "本地规则评审兜底",
            "score": rule_score.get("score", 0),
            "suitability_score": rule_score.get("score", 0),
            "suitability_level": suitability_level(rule_score.get("score", 0)),
            "fit_summary": f"与岗位「{profile.get('name', '无')}」的适配度为 {rule_score.get('score', 0)} 分。",
            "matched_evidence": rule_score.get("strengths", []),
            "gap_risks": rule_score.get("risks", []),
            "decision": rule_score.get("suggestion"),
            "strengths": rule_score.get("strengths", []),
            "risks": rule_score.get("risks", []),
            "questions": rule_score.get("questions", []),
            "next_stage": "待群面" if rule_score.get("score", 0) >= profile.get("pass_score", 70) else "待沟通",
            "reason": "基于技能权重、能力关键词和风险关键词生成。",
            "llm_error": llm.get("error"),
        }

    def recommend_next_action(self, candidate):
        config = load_app_config()
        stage = normalize_stage(candidate.get("stage"))
        table = {
            "待沟通": ("发送首次沟通话术", "确认实习周期、到岗时间、地点接受度和岗位意向。", "已沟通"),
            "已沟通": ("判断是否推进群面", "查看匹配分和追问结果，满足通过线则推进群面。", "待群面"),
            "待群面": ("发送群面通知", "同步群面时间、形式、注意事项，并记录候选人确认状态。", "已群面"),
            "已群面": ("录入群面评价", "补充群面表现、维度分和是否进入作业环节。", "待提交群面作业"),
            "待提交群面作业": ("催交群面作业", "发送作业提交提醒，并记录截止时间。", "已提交群面"),
            "已提交群面": ("评审作业并约最终面", "结合简历、群面和作业表现决定是否进入最终面。", "待最终面"),
            "待最终面": ("发送最终面邀约", "协调候选人与面试官时间，并发送最终面提醒。", "已最终面"),
            "已最终面": ("输出录用建议", "汇总所有环节证据，建议通过、淘汰或补充追问。", "通过"),
            "淘汰": ("归档候选人", "记录淘汰原因，沉淀到岗位复盘中。", "淘汰"),
            "通过": ("进入入职跟进", "确认 Offer、入职时间和材料清单。", "通过"),
        }
        action, reason, suggested_stage = table.get(stage, table["待沟通"])
        llm = call_llm_json(
            config.get("next_action_prompt") or DEFAULT_NEXT_ACTION_PROMPT,
            {"candidate": compact_candidate(candidate), "current_stage": stage, "default_action": action},
        )
        if not llm.get("error"):
            llm.setdefault("suggested_stage", suggested_stage)
            llm["method"] = "LLM下一步推荐"
            return llm
        return {
            "method": "本地流程规则兜底",
            "action": action,
            "reason": reason,
            "message": build_templates(candidate).get("群面确认", next(iter(build_templates(candidate).values()))),
            "suggested_stage": suggested_stage,
            "priority": "高" if stage.startswith("待") else "中",
            "llm_error": llm.get("error"),
        }

    def score_group_interview(self, candidate, text):
        config = load_app_config()
        llm = call_llm_json(
            config.get("group_score_prompt") or DEFAULT_GROUP_SCORE_PROMPT,
            {"candidate": compact_candidate(candidate), "group_record_or_homework": text[:6000]},
        )
        if not llm.get("error"):
            llm["method"] = "LLM群面评分"
            return llm
        score = local_group_score(text)
        return {
            "method": "本地关键词评分兜底",
            "group_score": score,
            "dimension_scores": {
                "表达能力": min(20, 12 + text.count("沟通") + text.count("表达")),
                "逻辑能力": min(20, 12 + text.count("逻辑") + text.count("分析")),
                "专业理解": min(20, 12 + text.count("项目") + text.count("技术")),
                "协作意识": min(20, 12 + text.count("团队") + text.count("协作")),
                "岗位匹配度": min(20, 12 + text.count("AI") + text.count("Python")),
            },
            "evaluation": "根据输入记录识别到候选人具备一定表达、项目和岗位相关信号。",
            "decision": "建议进入最终面" if score >= 70 else "建议暂缓推进",
            "reason": "未启用大模型，使用关键词和维度规则生成兜底评分。",
            "risks": [] if score >= 70 else ["群面证据不足或岗位信号偏弱"],
        }

    def interview_plan(self, candidate, profile_id=None):
        profile = self.profile_for(candidate, profile_id)
        rule_score = candidate.get("match") or score_candidate(candidate, profile)
        config = load_app_config()
        llm = call_llm_json(
            config.get("interview_plan_prompt") or DEFAULT_INTERVIEW_PLAN_PROMPT,
            {"job_profile": profile, "rule_score": rule_score, "candidate": compact_candidate(candidate)},
            temperature=0.25,
            timeout=35,
            cap_timeout=True,
        )
        if not llm.get("error"):
            llm["method"] = "LLM智能面试方案"
            return llm
        return local_interview_plan(candidate, profile, rule_score, llm.get("error"))

    def daily_report(self, candidates):
        summary = build_summary(candidates)
        config = load_app_config()
        llm = call_llm_json(
            config.get("daily_report_prompt") or DEFAULT_DAILY_REPORT_PROMPT,
            {"summary": summary, "candidates": [compact_candidate(c) for c in candidates[:40]]},
        )
        if not llm.get("error"):
            llm["method"] = "LLM招聘日报"
            return llm
        pending = [c for c in candidates if normalize_stage(c.get("stage")).startswith("待")]
        return {
            "method": "本地日报兜底",
            "summary": f"当前共有 {len(candidates)} 位候选人，推进中 {summary.get('active', 0)} 位，平均匹配分 {summary.get('avg_score', 0)}。",
            "risks": ["待处理候选人较多，建议优先处理超时阶段。"] if pending else [],
            "priorities": [f"{c.get('name')}：{normalize_stage(c.get('stage'))}" for c in pending[:5]],
            "actions": ["优先跟进待沟通/待群面候选人", "对已提交群面作业候选人尽快安排最终面", "复盘淘汰原因和岗位标准"],
            "bottlenecks": [k for k, v in summary.get("stages", {}).items() if v > 0 and k.startswith("待")],
        }


def extract_section_lines(text, titles, limit=6):
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    found = []
    capture = False
    for line in lines:
        if any(title in line for title in titles):
            capture = True
            continue
        if capture and re.search(r"教育背景|专业技能|个人荣誉|项目经历|实习经历|工作经历", line) and not any(t in line for t in titles):
            break
        if capture:
            found.append(line[:120])
        if len(found) >= limit:
            break
    return found


def build_resume_summary(candidate):
    skills = "、".join((candidate.get("skills") or [])[:8])
    return f"{candidate.get('name', '候选人')}投递{candidate.get('position', '无')}，学历信号：{candidate.get('education', '无')}，技能信号：{skills or '无'}。"


def local_interview_plan(candidate, profile, rule_score, llm_error=""):
    skills = candidate.get("skills") or rule_score.get("matched_skills") or []
    missing = rule_score.get("missing_skills") or []
    risks = rule_score.get("risks") or []
    questions = rule_score.get("questions") or []
    project_lines = extract_section_lines(candidate.get("resume_text") or "", ["项目经历", "项目经验", "个人项目"], 5)
    exp_lines = extract_section_lines(candidate.get("resume_text") or "", ["实习经历", "工作经历"], 4)
    focus = ["简历真实性与项目贡献边界", "岗位核心技能匹配度", "问题拆解、沟通表达和复盘能力"]
    if missing:
        focus.append("补齐缺口：" + "、".join(missing[:4]))
    if risks:
        focus.append("风险验证：" + "、".join(risks[:3]))
    return {
        "method": "本地规则面试方案兜底",
        "llm_error": llm_error,
        "interview_goal": f"验证候选人对{profile.get('name', candidate.get('position', '目标岗位'))}的真实匹配度，确认项目深度、技能掌握和推进风险。",
        "candidate_summary": build_resume_summary(candidate),
        "focus_areas": focus,
        "question_sections": [
            {"title": "一、开场与动机确认（5分钟）", "purpose": "确认岗位意愿、到岗安排和候选人表达状态。", "questions": ["请用 2 分钟介绍自己，并说明为什么关注这个岗位。", "你对这个岗位的核心工作内容如何理解？", "最快到岗时间、可实习/工作周期、工作地点接受度分别是什么？"]},
            {"title": "二、技术/能力基础验证（15分钟）", "purpose": "围绕简历技能和岗位标准验证基础能力。", "questions": [f"请说明你最熟悉的技能：{('、'.join(skills[:5]) or '简历中的核心技能')}，分别用在什么场景？", "遇到一个陌生业务需求时，你通常如何拆解、排期和交付？", *questions[:3]]},
            {"title": "三、项目深挖（20分钟）", "purpose": "确认项目真实性、个人贡献、复杂问题和结果产出。", "questions": ["请选择一个最能代表你能力的项目，说明背景、目标、你的职责和最终结果。", "项目中最难的问题是什么？你尝试过哪些方案，为什么最终选择当前方案？", "如果让你重新做一次，你会如何优化架构、流程或协作方式？", *project_lines[:3]]},
            {"title": "四、风险与稳定性确认（10分钟）", "purpose": "验证简历薄弱项、沟通风险和入职确定性。", "questions": ["简历里哪些内容你认为面试官最应该继续追问？", "如果入职后发现工作节奏/技术栈和预期不同，你会怎么处理？", "请补充说明：" + ("、".join(missing[:4]) if missing else "简历中未展开但与岗位相关的能力证明。")]},
        ],
        "project_deep_dive": project_lines or exp_lines or ["简历未识别到清晰项目段落，建议要求候选人现场讲述最近一次完整交付经历。"],
        "risk_checks": risks or ["简历信息完整度、项目贡献真实性、到岗稳定性"],
        "scoring_rubric": [
            {"dimension": "岗位技能匹配", "weight": 30, "pass_signal": "能解释关键技能的业务场景和实现细节"},
            {"dimension": "项目深度与真实性", "weight": 30, "pass_signal": "能讲清个人贡献、关键问题和结果指标"},
            {"dimension": "问题拆解与学习能力", "weight": 20, "pass_signal": "面对追问能结构化分析并给出取舍"},
            {"dimension": "沟通协作与稳定性", "weight": 20, "pass_signal": "表达清晰，动机稳定，到岗条件明确"},
        ],
        "schedule": [{"part": "开场与岗位动机", "minutes": 5}, {"part": "基础技能验证", "minutes": 15}, {"part": "项目深挖", "minutes": 20}, {"part": "风险确认与候选人提问", "minutes": 10}],
        "decision_signals": {"strong_hire": "项目讲述扎实，关键技能可落地，风险项解释充分。", "hold": "基础能力可接受，但项目贡献或稳定性需要二次确认。", "reject": "关键经历无法自洽，岗位核心技能明显缺口，或到岗条件不匹配。"},
        "interviewer_notes": "面试官记录时建议标注：证据、追问、风险、是否推进、下一步动作。",
    }


def compact_candidate(candidate):
    return {
        "id": candidate.get("id"),
        "name": candidate.get("name"),
        "position": candidate.get("position"),
        "stage": normalize_stage(candidate.get("stage")),
        "education": candidate.get("education"),
        "skills": candidate.get("skills", []),
        "match": candidate.get("match", {}),
        "group_interview": candidate.get("group_interview", {}),
        "assignment": candidate.get("assignment", {}),
        "hr_interview": candidate.get("hr_interview", {}),
        "notes": candidate.get("notes", ""),
        "resume_text": (candidate.get("resume_text") or "")[:5000],
    }


def local_group_score(text):
    base = 55
    for keyword in ["逻辑", "表达", "协作", "项目", "技术", "AI", "Python", "主动", "负责", "分析"]:
        if keyword.lower() in text.lower():
            base += 4
    return max(0, min(100, base))


def suitability_level(score):
    try:
        score = float(score)
    except Exception:
        score = 0
    if score >= 85:
        return "高度适配"
    if score >= 70:
        return "较适配"
    if score >= 55:
        return "一般适配"
    return "低适配"


def build_summary(candidates):
    total = len(candidates)
    stages = {stage: 0 for stage in PIPELINE}
    for c in candidates:
        stage = normalize_stage(c.get("stage"))
        stages[stage] = stages.get(stage, 0) + 1
    active = [c for c in candidates if normalize_stage(c.get("stage")) not in ["淘汰", "通过"]]
    scored = [c for c in candidates if c.get("match", {}).get("score") not in [None, ""]]
    avg_score = round(sum(float(c.get("match", {}).get("score", 0)) for c in scored) / len(scored), 1) if scored else 0
    pass_count = len([c for c in scored if float(c.get("match", {}).get("score", 0)) >= 70])
    metrics = build_metrics(candidates, stages, scored, pass_count)
    duplicates = detect_duplicate_candidates(candidates)
    reminders = []
    for c in candidates:
        if normalize_stage(c.get("stage")) in ["待沟通", "待群面", "待提交群面作业", "待最终面"]:
            reminders.append({
                "id": c.get("id"),
                "name": c.get("name"),
                "stage": normalize_stage(c.get("stage")),
                "message": f"{c.get('name')} 当前处于「{normalize_stage(c.get('stage'))}」，建议今天完成跟进。",
            })
    return {
        "total": total,
        "active": len(active),
        "avg_score": avg_score,
        "pass_rate": round(pass_count / len(scored) * 100, 1) if scored else 0,
        "stages": stages,
        "metrics": metrics,
        "reminders": reminders[:12],
        "duplicates": duplicates[:20],
        "talent_pool_count": len([c for c in candidates if c.get("talent_pool")]),
        "exports": {
            "csv": str((EXPORT_DIR / "candidates.csv").relative_to(ROOT)),
            "status_csv": str(STATUS_CSV_FILE.relative_to(ROOT)),
        },
    }


def pct(num, den):
    return round(num / den * 100, 1) if den else 0


def build_metrics(candidates, stages, scored, pass_count):
    total = len(candidates)
    group_done = stages.get("已群面", 0) + stages.get("待提交群面作业", 0) + stages.get("已提交群面", 0) + stages.get("待最终面", 0) + stages.get("已最终面", 0) + stages.get("通过", 0)
    final_invited = stages.get("待最终面", 0) + stages.get("已最终面", 0) + stages.get("通过", 0)
    interviewed = stages.get("已最终面", 0) + stages.get("通过", 0)
    offered = stages.get("通过", 0)
    rejected = stages.get("淘汰", 0)
    by_position = {}
    for c in candidates:
        position = c.get("position") or "无"
        item = by_position.setdefault(position, {
            "position": position,
            "total": 0,
            "active": 0,
            "avg_score": 0,
            "scored": 0,
            "pass": 0,
            "offer": 0,
            "rejected": 0,
            "talent_pool": 0,
        })
        item["total"] += 1
        stage = normalize_stage(c.get("stage"))
        if stage not in ["淘汰", "通过"]:
            item["active"] += 1
        if stage == "通过":
            item["offer"] += 1
        if stage == "淘汰":
            item["rejected"] += 1
        if c.get("talent_pool"):
            item["talent_pool"] += 1
        score = c.get("match", {}).get("score")
        if score not in [None, ""]:
            item["avg_score"] += float(score)
            item["scored"] += 1
            if float(score) >= 70:
                item["pass"] += 1
    for item in by_position.values():
        item["avg_score"] = round(item["avg_score"] / item["scored"], 1) if item["scored"] else 0
        item["pass_rate"] = pct(item["pass"], item["scored"])
        item["offer_rate"] = pct(item["offer"], item["total"])
    return {
        "resume_pass_rate": pct(pass_count, len(scored)),
        "screen_pass_rate": pct(group_done, total),
        "interview_invite_rate": pct(final_invited, total),
        "interview_show_rate": pct(interviewed, final_invited),
        "offer_conversion_rate": pct(offered, total),
        "loss_rate": pct(rejected, total),
        "positions": sorted(by_position.values(), key=lambda x: (-x["total"], x["position"])),
    }


def duplicate_key(value):
    value = str(value or "").strip().lower()
    return value if value and value not in ["无", "none", "null"] else ""


def detect_duplicate_candidates(candidates):
    buckets = {}
    for c in candidates:
        keys = [
            ("phone", duplicate_key(c.get("phone"))),
            ("email", duplicate_key(c.get("email"))),
            ("name_position", duplicate_key(c.get("name")) + "|" + duplicate_key(c.get("position"))),
        ]
        for kind, key in keys:
            if not key or key == "|":
                continue
            buckets.setdefault((kind, key), []).append(c)
    duplicates = []
    seen = set()
    for (kind, key), items in buckets.items():
        if len(items) < 2:
            continue
        ids = tuple(sorted(i.get("id", "") for i in items))
        if ids in seen:
            continue
        seen.add(ids)
        duplicates.append({
            "kind": kind,
            "key": key,
            "count": len(items),
            "candidates": [{"id": i.get("id"), "name": i.get("name"), "position": i.get("position"), "source": i.get("source")} for i in items],
        })
    return duplicates


def mask_text(value, keep=3):
    value = str(value or "")
    if not value:
        return ""
    if "@" in value:
        name, domain = value.split("@", 1)
        return (name[:2] + "***@" + domain) if name else "***@" + domain
    if len(value) <= keep * 2:
        return value[:1] + "***"
    return value[:keep] + "****" + value[-keep:]


def candidates_for_user(candidates, user):
    config = load_app_config()
    if user.get("role") == "manager" or not config.get("mask_contact_for_hr"):
        return candidates
    safe = []
    for c in candidates:
        item = dict(c)
        item["phone"] = mask_text(item.get("phone"))
        item["email"] = mask_text(item.get("email"), keep=2)
        safe.append(item)
    return safe


def build_templates(candidate):
    name = candidate.get("name") or "同学"
    position = candidate.get("position") or "该岗位"
    questions = candidate.get("match", {}).get("questions") or ["确认可实习周期、工作地点接受度和最快到岗时间。"]
    score = candidate.get("match", {}).get("score", "未评分")
    return {
        "群面确认": f"{name}你好，我是招聘团队 HR。看到你投递/沟通的是{position}，简历匹配度为{score}分。想再和你确认一下实习周期、每周到岗天数、工作地点接受度和最快到岗时间，方便我们推进群面安排。",
        "追问问题": "\n".join([f"{i + 1}. {q}" for i, q in enumerate(questions[:4])]),
        "最终面邀约": f"{name}你好，你的{position}方向群面/作业已完成。想和你约一次最终面，请你回复 2-3 个方便的时间段，我们会尽快协调面试官。",
        "最终面提醒": f"{name}你好，提醒一下今天有{position}岗位最终面安排。请提前 5 分钟准备好简历、项目介绍和网络环境，如时间有变化请及时告诉我。",
        "反馈催办": f"面试官老师好，{name}的{position}最终面已完成，麻烦方便时同步一下评价、是否通过以及建议关注点，感谢。",
    }


HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>AI 招聘流程 Copilot</title>
  <style>
    :root{--bg:#f6f7fb;--panel:#fff;--ink:#172033;--muted:#667085;--line:#d9dee8;--brand:#126b61;--brand2:#2f6fed;--warn:#c97913;--bad:#c43d4b;--ok:#228b5b;--shadow:0 12px 32px rgba(26,34,52,.09)}
    body.theme-dark{--bg:#0f141d;--panel:#171e29;--ink:#edf3fb;--muted:#a8b4c4;--line:#2c3747;--shadow:0 16px 36px rgba(0,0,0,.28)}
    *{box-sizing:border-box} body{margin:0;background:var(--bg);color:var(--ink);font-family:"Microsoft YaHei",Arial,sans-serif;letter-spacing:0}
    .app{min-height:100vh;display:grid;grid-template-columns:248px 1fr}
    aside{background:#172033;color:#f7fbff;padding:22px 18px;position:sticky;top:0;height:100vh}
    .brand{font-size:22px;font-weight:800;line-height:1.2;margin-bottom:22px}.brand small{display:block;font-size:12px;color:#adc0d8;margin-top:8px;font-weight:500}
    nav button{width:100%;border:0;background:transparent;color:#dce7f5;text-align:left;padding:11px 12px;border-radius:8px;margin:3px 0;cursor:pointer;font-size:14px}
    nav button.active,nav button:hover{background:#243149;color:#fff}
    main{padding:24px 28px 40px}.topbar{display:flex;align-items:center;justify-content:space-between;gap:16px;margin-bottom:18px}.topbar h1{font-size:24px;margin:0}.sub{color:var(--muted);font-size:13px;margin-top:4px}
    .actions{display:flex;gap:10px;flex-wrap:wrap}.btn{border:1px solid var(--line);background:var(--panel);border-radius:8px;padding:10px 14px;cursor:pointer;color:var(--ink);font-weight:700;white-space:nowrap}.btn.primary{background:var(--brand);color:#fff;border-color:var(--brand)}.btn.blue{background:var(--brand2);color:#fff;border-color:var(--brand2)}.btn:disabled{opacity:.55;cursor:not-allowed}
    .grid{display:grid;gap:16px}.kpis{grid-template-columns:repeat(4,minmax(150px,1fr));margin-bottom:16px}.card{background:var(--panel);border:1px solid var(--line);border-radius:8px;box-shadow:var(--shadow)}
    .kpi{padding:18px}.kpi .label{font-size:13px;color:var(--muted)}.kpi .value{font-size:30px;font-weight:800;margin-top:8px}
    .layout{display:grid;grid-template-columns:1.5fr .9fr;gap:16px}.section{padding:18px}.section h2{font-size:17px;margin:0 0 14px}.muted{color:var(--muted)}
    .stage-row{display:grid;grid-template-columns:86px 1fr 42px;align-items:center;gap:10px;margin:10px 0}.bar{height:10px;background:#e8edf5;border-radius:999px;overflow:hidden}.bar span{display:block;height:100%;background:linear-gradient(90deg,var(--brand),var(--brand2));border-radius:999px}
    table{width:100%;border-collapse:collapse}th,td{text-align:left;border-bottom:1px solid var(--line);padding:11px 8px;font-size:13px;vertical-align:middle}th{color:var(--muted);font-weight:800;background:rgba(127,139,158,.09)}tr:hover td{background:rgba(127,139,158,.07)}
    .pill{display:inline-flex;align-items:center;gap:4px;border-radius:999px;padding:4px 9px;background:#eef4ff;color:#2455a6;font-weight:700;font-size:12px}.pill.ok{background:#eaf7ef;color:var(--ok)}.pill.warn{background:#fff5e7;color:var(--warn)}.pill.bad{background:#fff0f1;color:var(--bad)}
    .toolbar{display:flex;gap:10px;align-items:center;margin-bottom:12px;flex-wrap:wrap}.toolbar input{min-width:200px}.toolbar select{min-width:180px}input,select,textarea{border:1px solid var(--line);border-radius:8px;padding:10px 11px;background:var(--panel);color:var(--ink);font:inherit}input{height:40px}select{height:44px;line-height:22px;padding:8px 34px 8px 11px}textarea{width:100%;min-height:90px;resize:vertical}
    .detail{display:grid;gap:12px}.detail .row{display:grid;grid-template-columns:82px 1fr;gap:10px;font-size:13px}.chips{display:flex;flex-wrap:wrap;gap:6px}.chip{background:#f0f3f8;border:1px solid #e1e6ef;border-radius:999px;padding:5px 8px;font-size:12px}
    .kanban{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.lane{background:#fff;border:1px solid var(--line);border-radius:8px;min-height:180px;padding:12px}.lane h3{font-size:14px;margin:0 0 10px}.mini{border:1px solid var(--line);border-radius:8px;padding:10px;margin-bottom:8px;background:#fbfcff}.mini b{display:block;font-size:13px}.mini span{font-size:12px;color:var(--muted)}
    .profile-editor{display:grid;grid-template-columns:320px 1fr;gap:16px}.list button{display:block;width:100%;padding:12px;text-align:left;background:#fff;border:1px solid var(--line);border-radius:8px;margin-bottom:8px;cursor:pointer}.list button.active{border-color:var(--brand);box-shadow:0 0 0 2px rgba(18,107,97,.12)}
    .profile-panel{padding:0;overflow:hidden}.profile-head{background:linear-gradient(135deg,#f0f8f6,#f4f7ff);border-bottom:1px solid var(--line);padding:18px 20px;display:flex;align-items:flex-start;justify-content:space-between;gap:16px}.profile-head h2{margin:0;font-size:18px}.profile-head p{margin:6px 0 0;color:var(--muted);font-size:13px}.profile-stats{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}.stat{min-width:82px;border:1px solid #d8e4ee;background:rgba(255,255,255,.72);border-radius:8px;padding:8px 10px;text-align:center}.stat b{display:block;font-size:18px;color:var(--brand)}.stat span{display:block;font-size:12px;color:var(--muted);white-space:nowrap}.jd-builder{display:grid;grid-template-columns:1fr 1fr;gap:12px}.jd-form{display:grid;gap:10px}.jd-preview{min-height:280px;max-height:520px;overflow:auto;white-space:pre-wrap;line-height:1.65;color:#263348}.jd-actions{display:flex;gap:8px;flex-wrap:wrap;align-items:center}.input-row{display:grid;grid-template-columns:1fr 1fr;gap:10px}
    .profile-form{padding:18px 20px 20px;display:grid;gap:14px}.form-grid{display:grid;grid-template-columns:1fr 180px;gap:12px}.field{display:grid;gap:7px}.field label{font-size:13px;font-weight:800;color:#263348}.hint{font-size:12px;color:var(--muted);font-weight:500}.field input,.field textarea{width:100%}.textarea-short{min-height:86px}.textarea-tall{min-height:132px}.score-blocks{display:grid;grid-template-columns:1fr 1fr;gap:12px}.config-box{border:1px solid var(--line);border-radius:8px;background:#fbfcff;padding:12px}.config-box h3{margin:0 0 8px;font-size:14px}.config-box textarea{border-color:#cfd8e6;background:#fff;font-family:Consolas,"Microsoft YaHei",monospace;line-height:1.45}.risk-row{display:grid;grid-template-columns:1fr 210px;gap:12px;align-items:end}.save-bar{display:flex;align-items:center;justify-content:space-between;gap:12px;border-top:1px solid var(--line);padding-top:14px}.save-note{font-size:12px;color:var(--muted)}
    .metrics{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:14px}.metric{border:1px solid var(--line);border-radius:8px;padding:12px;background:#fbfcff;min-width:0}.metric b{display:block;font-size:20px;margin-top:6px}.metric span{display:block;white-space:normal}.template{border:1px solid var(--line);border-radius:8px;padding:10px;background:#fbfcff}.template h4{margin:0 0 8px;font-size:13px}.template pre{white-space:pre-wrap;margin:0;color:#344054;font-family:inherit;font-size:13px;line-height:1.55}.copy{float:right;border:0;background:#e8f0ff;color:#2455a6;border-radius:6px;padding:4px 7px;cursor:pointer}
    .agent-card{border:1px solid #cfe0da;background:linear-gradient(180deg,#fbfffd,#f7faff);border-radius:8px;padding:12px}.agent-card h3{margin:0 0 8px;font-size:14px}.agent-card pre{white-space:pre-wrap;margin:0;color:#2b3648;font-family:inherit;font-size:13px;line-height:1.55}.agent-actions{display:flex;gap:8px;flex-wrap:wrap}.btn.ghost{background:#f3f7fb;color:#263348}.btn.small{padding:7px 10px;font-size:12px}.ai-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px}.daily-list{display:grid;gap:8px}.daily-list div{border:1px solid var(--line);border-radius:8px;background:#fbfcff;padding:10px;font-size:13px}
    .collapsible{border:1px solid var(--line);border-radius:8px;background:#fff;overflow:hidden}.collapse-head{width:100%;border:0;background:#f8fafc;color:var(--ink);display:flex;align-items:center;justify-content:space-between;gap:10px;padding:10px 12px;cursor:pointer;font:inherit;font-weight:800;text-align:left}.collapse-arrow{width:18px;height:18px;display:inline-grid;place-items:center;transition:.16s transform;color:var(--muted)}.collapsible.collapsed .collapse-arrow{transform:rotate(-90deg)}.collapse-body{padding:10px}.collapsible.collapsed .collapse-body{display:none}.collapse-body .template{border:0;background:transparent;padding:0}
    .review-form{display:grid;gap:12px}.review-grid{display:grid;grid-template-columns:1fr 96px;gap:10px}.review-line{display:grid;grid-template-columns:92px 1fr;gap:10px;align-items:center}.review-line b{font-size:13px}.review-form input[type=number]{width:100%;min-width:0}.review-summary{display:grid;grid-template-columns:repeat(3,1fr);gap:8px}.review-summary .metric{padding:9px 10px}.review-summary b{font-size:18px}
    .auth-screen{position:fixed;inset:0;background:linear-gradient(135deg,#172033,#126b61);display:none;align-items:center;justify-content:center;padding:24px;z-index:20}.auth-box{width:min(440px,100%);background:#fff;border-radius:8px;box-shadow:0 24px 80px rgba(0,0,0,.28);padding:24px}.auth-box h2{margin:0 0 8px}.auth-tabs{display:flex;background:#f2f5f9;border-radius:8px;padding:4px;margin:16px 0}.auth-tabs button{flex:1;border:0;background:transparent;border-radius:6px;padding:9px;cursor:pointer;font-weight:800}.auth-tabs button.active{background:#fff;color:var(--brand);box-shadow:0 2px 8px rgba(22,33,52,.08)}.auth-form{display:grid;gap:10px}.role-row{display:grid;grid-template-columns:minmax(116px,.42fr) minmax(0,.58fr);gap:8px}.role-row select,.role-row input{min-width:0;width:100%;overflow:hidden;text-overflow:ellipsis}.userbar{position:relative;color:#dce7f5;font-size:13px;margin:14px 0;border:1px solid rgba(255,255,255,.12);border-radius:8px;overflow:hidden}.userbar b{color:#fff}.account-trigger{width:100%;border:0;background:transparent;color:#dce7f5;display:flex;align-items:center;gap:8px;padding:10px;cursor:pointer;text-align:left}.avatar{width:34px;height:34px;border-radius:999px;display:inline-grid;place-items:center;flex:0 0 34px;background:var(--brand);color:#fff;font-weight:900;overflow:hidden}.avatar img{width:100%;height:100%;object-fit:cover}.avatar.large{width:58px;height:58px;flex-basis:58px;font-size:20px}.account-info{display:flex;align-items:center;gap:8px;min-width:0;flex:1}.account-text{display:grid;gap:2px;min-width:0}.account-text b{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.account-text span{font-size:12px;color:#adc0d8}.account-arrow{margin-left:auto;transition:.16s transform}.userbar.open .account-arrow{transform:rotate(180deg)}.account-menu{display:none;border-top:1px solid rgba(255,255,255,.1);padding:8px;gap:6px;background:rgba(255,255,255,.03)}.userbar.open .account-menu{display:grid}.account-menu button{border:0;border-radius:7px;background:rgba(255,255,255,.08);color:#eef5ff;text-align:left;padding:8px 10px;cursor:pointer;font-weight:700}.account-menu button:hover{background:rgba(255,255,255,.14)}.settings-head{display:flex;align-items:center;gap:14px}.admin-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:14px}.admin-table td,.admin-table th{font-size:13px}
    .model-stats{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.model-tools{display:grid;grid-template-columns:1fr 160px 160px auto;gap:10px;margin:12px 0}.model-list{display:grid;gap:10px}.model-card{border:1px solid var(--line);border-radius:8px;background:#fbfcff;overflow:hidden}.model-card.active{border-color:var(--brand2);box-shadow:0 0 0 2px rgba(47,111,237,.12)}.model-title{display:flex;align-items:center;justify-content:space-between;padding:12px 14px;border-bottom:1px solid var(--line)}.model-title h3{margin:0;font-size:15px}.model-meta{display:grid;grid-template-columns:80px 1fr 80px 1fr 80px 1fr;gap:8px 10px;padding:12px 14px;font-size:13px}.model-label{color:var(--brand2);font-weight:800}.model-actions{display:flex;gap:8px;flex-wrap:wrap;padding:0 14px 12px}.badge{display:inline-flex;border-radius:999px;padding:4px 9px;font-size:12px;font-weight:800;background:#eef2f7;color:#475467}.badge.ok{background:#eaf7ef;color:var(--ok)}.badge.active{background:#eef4ff;color:var(--brand2)}.model-form{display:grid;gap:10px}.model-form-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px}.model-test-result{margin:0 14px 12px;border:1px solid #b7e0c5;border-radius:8px;background:#edf8f0;color:#1f7a42;padding:10px 12px;font-size:13px;line-height:1.6}.model-test-result.bad{border-color:#f1b8c0;background:#fff1f3;color:#a12c3a}.modal-mask{position:fixed;inset:0;background:rgba(15,23,42,.45);display:none;align-items:center;justify-content:center;padding:24px;z-index:30}.modal-mask.open{display:flex}.modal-panel{width:min(920px,96vw);max-height:88vh;overflow:auto;background:var(--panel);border:1px solid var(--line);border-radius:8px;box-shadow:0 28px 80px rgba(0,0,0,.28)}.modal-head{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:16px 18px;border-bottom:1px solid var(--line)}.modal-head h3{margin:0;font-size:17px}.modal-body{padding:18px}.icon-btn{border:1px solid var(--line);background:var(--panel);color:var(--ink);border-radius:8px;width:34px;height:34px;cursor:pointer;font-weight:900}
    .toast{position:fixed;right:20px;bottom:20px;background:#172033;color:#fff;padding:12px 14px;border-radius:8px;box-shadow:var(--shadow);display:none;max-width:360px}.view{display:none}.view.active{display:block}
    @media(max-width:980px){.app{grid-template-columns:1fr}aside{height:auto;position:relative}.layout,.profile-editor,.form-grid,.score-blocks,.risk-row,.ai-grid,.admin-grid,.model-stats,.model-tools,.model-form-grid,.model-meta,.review-grid,.review-line,.review-summary,.jd-builder,.input-row{grid-template-columns:1fr}.profile-head{display:block}.profile-stats{justify-content:flex-start;margin-top:12px}.kpis{grid-template-columns:repeat(2,1fr)}.kanban{grid-template-columns:repeat(2,1fr)}main{padding:18px}}
  </style>
</head>
<body>
<div class="auth-screen" id="authScreen">
  <div class="auth-box">
    <h2>AI 招聘流程 Copilot</h2>
    <div class="sub">请登录后使用。普通 HR 可操作招聘流程，管理者可查看账号和今日操作统计。</div>
    <div class="auth-tabs"><button id="loginTab" class="active">登录</button><button id="registerTab">注册</button></div>
    <div class="auth-form">
      <input id="authUsername" placeholder="用户名" />
      <input id="authPassword" type="password" placeholder="密码，至少 6 位" />
      <input id="authDisplay" placeholder="显示名称（注册时可填）" style="display:none" />
      <div class="role-row" id="roleRow" style="display:none">
        <select id="authRole"><option value="hr">HR</option><option value="manager">管理者</option></select>
        <input id="managerCode" placeholder="管理者注册码" />
      </div>
      <button class="btn primary" id="authSubmit">登录</button>
      <div class="muted" id="authHint">管理者注册需输入注册码，默认可在配置文件中修改。</div>
    </div>
  </div>
</div>
<div class="app">
  <aside>
    <div class="brand">AI 招聘流程 Copilot<small>信息沉淀 · 流程跟进 · 数据分析</small></div>
    <div class="userbar" id="userbar" style="display:none"></div>
    <nav>
      <button class="active" data-view="dashboard">数据看板</button>
      <button data-view="collection">简历采集</button>
      <button data-view="interview">智能面试</button>
      <button data-view="training">智能培训</button>
      <button data-view="candidates">候选人库</button>
      <button data-view="talentPool">人才池</button>
      <button data-view="pipeline">流程跟进</button>
      <button data-view="profiles">岗位标准</button>
      <button data-view="admin" id="adminNav" style="display:none">管理者</button>
      <button data-view="modelSettings" id="modelSettingsNav" style="display:none">模型设置</button>
      <button data-view="promptSettings" id="promptSettingsNav" style="display:none">提示词设置</button>
      <button data-view="settings" id="settingsNav" style="display:none">系统设置</button>
    </nav>
  </aside>
  <main>
    <div class="topbar">
      <div><h1 id="title">数据看板</h1><div class="sub">本地低成本部署，所有数据写入 data 目录</div></div>
      <div class="actions">
        <button class="btn primary" id="scanBtn">扫描简历</button>
        <button class="btn blue" id="scoreAllBtn">批量评分</button>
        <button class="btn" id="exportBtn">导出 CSV</button>
      </div>
    </div>

    <section id="dashboard" class="view active">
      <div class="grid kpis">
        <div class="card kpi"><div class="label">候选人总数</div><div class="value" id="kTotal">0</div></div>
        <div class="card kpi"><div class="label">推进中</div><div class="value" id="kActive">0</div></div>
        <div class="card kpi"><div class="label">平均匹配分</div><div class="value" id="kScore">0</div></div>
        <div class="card kpi"><div class="label">推荐推进率</div><div class="value" id="kPass">0%</div></div>
      </div>
      <div class="layout">
        <div class="card section"><h2>招聘漏斗</h2><div id="stageBars"></div><div id="metrics" class="metrics"></div></div>
        <div class="card section"><h2>智能提醒</h2><div id="reminders" class="detail"></div></div>
      </div>
      <div class="layout" style="margin-top:16px">
        <div class="card section"><h2>岗位招聘看板</h2><div id="positionBoard" class="detail"></div></div>
        <div class="card section"><h2>风险与人才池</h2><div id="riskBoard" class="detail"></div></div>
      </div>
      <div class="card section" style="margin-top:16px"><h2>AI 招聘日报</h2><div id="dailyReport" class="daily-list muted">正在生成日报...</div></div>
    </section>

    <section id="collection" class="view">
      <div class="card section" style="margin-bottom:16px">
        <div class="profile-head" style="margin:-18px -18px 14px">
          <div><h2>简历采集</h2><p>每位 HR 使用自己的后台软件/API 配置，选择来源后一键采集；系统会做字段标准化、去重和文本清洗。</p></div>
          <div class="actions">
            <button class="btn" type="button" onclick="showCollectGuide()">配置教程</button>
            <button class="btn primary" type="button" onclick="collectSelectedSource()">采集</button>
          </div>
        </div>
        <div class="toolbar">
          <select id="collectSource"></select>
          <input id="resumeFiles" type="file" multiple accept=".pdf,.docx,.txt,.md" style="display:none" onchange="uploadResumeFiles(this.files)">
          <span class="muted" id="collectHint">来源由个人配置决定；未配置 API 时可用本地上传或粘贴文本兜底。</span>
        </div>
        <div id="collectGuide" class="template" style="display:none;margin-bottom:12px">
          <h4>个人账号/API 配置教程</h4>
          <p>1. 在本页“个人招聘软件/API”中维护自己的 BOSS、猎聘等后台软件。</p>
          <p>2. 为每个来源填写 API 地址、请求方式、API Key、请求头和请求体。</p>
          <p>3. API 返回建议包含 <b>files</b>、<b>texts</b> 或 <b>candidates</b> 字段，系统会统一清洗为姓名、岗位、联系方式、简历文本和来源。</p>
          <p>4. 每位 HR 的配置只保存在自己的账号里，互不共用；遇到反爬或格式不统一时，优先使用官方/后台接口。</p>
        </div>
        <div id="collectionSettings"></div>
        <div class="field">
          <label>平台候选人文本</label>
          <textarea id="platformResumeText" class="textarea-short" placeholder="从 BOSS/猎聘候选人详情页复制姓名、岗位、联系方式、经历等内容，粘贴后点击导入。"></textarea>
        </div>
        <div class="actions" style="margin-top:10px">
          <button class="btn primary" type="button" onclick="importPlatformText()">导入粘贴内容</button>
          <button class="btn" type="button" onclick="$('#platformResumeText').value=''">清空</button>
          <span class="muted" id="collectResult"></span>
        </div>
      </div>
    </section>

    <section id="interview" class="view">
      <div class="profile-editor">
        <div class="card section">
          <div class="toolbar">
            <input id="interviewSearch" placeholder="搜索候选人、岗位、技能" oninput="renderInterview()">
            <button class="btn small" type="button" onclick="renderInterview()">刷新</button>
          </div>
          <div id="interviewCandidateList" class="list"></div>
        </div>
        <div class="card section">
          <div class="profile-head" style="margin:-18px -18px 14px">
            <div><h2 id="interviewTitle">智能面试</h2><p>根据候选人简历、岗位标准和评分风险，生成定制化面试方案。</p></div>
            <div class="actions"><button class="btn primary" type="button" onclick="generateInterviewPlan()">生成面试方案</button></div>
          </div>
          <div class="toolbar">
            <button class="btn small" id="resumeTabBtn" type="button" onclick="setInterviewTab('resume')">简历预览</button>
            <button class="btn small blue" id="planTabBtn" type="button" onclick="setInterviewTab('plan')">面试方案</button>
          </div>
          <div id="interviewPanel" class="detail muted">请选择左侧候选人。</div>
        </div>
      </div>
    </section>

    <section id="training" class="view">
      <div class="grid kpis">
        <div class="card kpi"><div class="label">试卷数量</div><div class="value" id="tExamCount">0</div></div>
        <div class="card kpi"><div class="label">考试记录</div><div class="value" id="tResultCount">0</div></div>
        <div class="card kpi"><div class="label">平均得分</div><div class="value" id="tAvgScore">0</div></div>
        <div class="card kpi"><div class="label">通过率</div><div class="value" id="tPassRate">0%</div></div>
      </div>
      <div class="profile-editor">
        <div class="card section">
          <div class="profile-head" style="margin:-18px -18px 14px">
            <div><h2>试卷生成</h2><p>根据培训主题、岗位技能和难度生成试卷，生成后进入考试管理。</p></div>
          </div>

            <div class="field"><label>考试名称</label><input id="examTitle" placeholder="例如：调研试卷考核"></div>
            <div class="field"><label>培训对象</label><input id="examTarget" placeholder="例如：AI算法实习生 / 新员工 / HR"></div>
            <div class="field"><label>考试主题</label><input id="examTopic" placeholder="例如：大模型应用开发、业务知识、招聘流程"></div>
            <div class="input-row">
              <div class="field"><label>难度</label><select id="examDifficulty"><option>中等</option><option>基础</option><option>进阶</option><option>高阶</option></select></div>
              <div class="field"><label>考试时长</label><input id="examDuration" type="number" value="60"></div>
            </div>
            <div class="input-row">
              <div class="field"><label>总分</label><input id="examTotalScore" type="number" value="100"></div>
              <div class="field"><label>通过分</label><input id="examPassScore" type="number" value="60"></div>
            </div>
            <div class="field"><label>技能点</label><textarea id="examSkills" class="textarea-short" placeholder="每行或用顿号分隔：Python、沟通表达、业务理解、项目拆解"></textarea></div>
            <div class="field"><label>补充说明</label><textarea id="examDescription" class="textarea-short" placeholder="说明考试场景、题型偏好、重点考察内容"></textarea></div>
            <div class="actions"><button class="btn blue" type="button" onclick="generateTrainingExam()">生成试卷</button><button class="btn" type="button" onclick="clearExamForm()">清空</button></div>
          </div>
          <h2 style="margin-top:18px">试卷列表</h2>
          <div id="trainingExamList" class="list"></div>
        </div>
        <div class="card section">
          <div class="profile-head" style="margin:-18px -18px 14px">
            <div><h2 id="trainingTitle">考试管理</h2><p>基于标准答案模板评分，复杂语义题标记人工复核，降低 AI 误判风险。</p></div>
            <div class="actions"><button class="btn small" type="button" onclick="renderTraining()">刷新数据</button></div>
          </div>
          <div class="toolbar">
            <button class="btn small blue" id="examPreviewTab" type="button" onclick="setTrainingTab('preview')">试卷预览</button>
            <button class="btn small" id="examResultTab" type="button" onclick="setTrainingTab('results')">考试管理</button>
          </div>
          <div id="trainingPanel" class="detail muted">请选择左侧试卷，或先生成一份试卷。</div>
        </div>
      </div>
    </section>

    <section id="candidates" class="view">
      <div class="layout">
        <div class="card section">
          <div class="toolbar">
            <input id="search" placeholder="搜索姓名、岗位、技能" />
            <select id="stageFilter"><option value="">全部阶段</option></select>
          </div>
          <table><thead><tr><th>姓名</th><th>岗位</th><th>阶段</th><th>匹配</th><th>来源</th><th>更新时间</th></tr></thead><tbody id="candidateRows"></tbody></table>
        </div>
        <div class="card section"><h2>候选人详情</h2><div id="candidateDetail" class="detail muted">选择左侧候选人查看解析、评分和追问建议。</div></div>
      </div>
    </section>

    <section id="talentPool" class="view">
      <div class="card section">
        <div class="toolbar">
          <input id="talentSearch" placeholder="搜索姓名、岗位、标签" oninput="renderTalentPool()">
          <button class="btn small" type="button" onclick="renderTalentPool()">刷新</button>
        </div>
        <div id="talentPoolList" class="detail"></div>
      </div>
    </section>

    <section id="pipeline" class="view">
      <div class="card section"><h2>流程看板</h2><div id="kanban" class="kanban"></div></div>
    </section>

    <section id="profiles" class="view">
      <div class="profile-editor">
        <div class="card section"><h2>岗位列表</h2><div id="profileList" class="list"></div><button class="btn primary" id="newProfileBtn">新增岗位标准</button></div>
        <div><div id="profileForm"></div></div>
      </div>
    </section>

    <section id="admin" class="view">
      <div class="card section"><h2>账号与今日操作</h2><div id="adminPanel" class="muted">仅管理者可见。</div></div>
    </section>

    <section id="modelSettings" class="view">
      <div class="card section"><h2>模型设置</h2><div id="modelSettingsPanel" class="muted">仅管理者可配置。</div></div>
    </section>

    <section id="promptSettings" class="view">
      <div class="card section"><h2>提示词设置</h2><div id="promptSettingsPanel" class="muted">仅管理者可配置。</div></div>
    </section>

    <section id="settings" class="view">
      <div class="card section"><h2>系统设置</h2><div id="settingsPanel" class="muted">仅管理者可配置。</div></div>
    </section>
  </main>
</div>
<div class="toast" id="toast"></div>
<script>
const PIPELINE = ["待沟通","已沟通","待群面","已群面","待提交群面作业","已提交群面","待最终面","已最终面","淘汰","通过"];
let candidates=[], summary={}, profiles=[], collectionSources=[], trainingData={exams:[],results:[],summary:{}}, dailyReport=null, currentUser=null, adminOverview=null, appConfig=null, selectedCandidate=null, selectedInterviewId=null, interviewTab="plan", selectedTrainingExamId=null, trainingTab="preview", selectedProfile=null, editingModelId=null, editingCollectId=null, jdDraft=null, authMode="login";
const $=s=>document.querySelector(s);
function toast(msg){const t=$("#toast");t.textContent=msg;t.style.display="block";setTimeout(()=>t.style.display="none",3200)}
async function api(url, opts={}){const r=await fetch(url,{credentials:"same-origin",headers:{"Content-Type":"application/json"},...opts});if(!r.ok){const text=await r.text();const err=new Error(text||r.statusText);err.status=r.status;throw err}return r.headers.get("content-type")?.includes("json")?r.json():r.text()}
async function init(){try{const me=await api("/api/me");currentUser=me.user;if(!currentUser){showAuth(true);return}showAuth(false);renderUserbar();await refresh()}catch(e){showAuth(true);toast("登录状态已失效，请重新登录")}}
function showAuth(show){$("#authScreen").style.display=show?"flex":"none"}
async function refresh(){
  dailyReport=null;
  try{
    [candidates,summary,profiles]=await Promise.all([api("/api/candidates"),api("/api/summary"),api("/api/job-profiles")]);
    const optional=await Promise.allSettled([api("/api/collection-sources"),api("/api/training")]);
    collectionSources=optional[0].status==="fulfilled"?optional[0].value:(currentUser?.collection_sources||[]);
    trainingData=optional[1].status==="fulfilled"?optional[1].value:{exams:[],results:[],summary:{}};
    if(optional.some(x=>x.status==="rejected")) toast("部分扩展数据加载失败，候选人数据已优先显示");
  }catch(e){
    if(e.status===401){showAuth(true);toast("登录状态已失效，请重新登录");return}
    toast("基础数据加载失败："+e.message);return
  }
  renderSummary();renderCollection();renderInterview();renderTraining();renderCandidates();renderTalentPool();renderKanban();renderProfiles();renderDaily();
  loadDailyReport();
  if(currentUser?.role==="manager"){
    const settled=await Promise.allSettled([api("/api/admin/overview"),api("/api/app-config")]);
    if(settled[0].status==="fulfilled")adminOverview=settled[0].value;
    if(settled[1].status==="fulfilled")appConfig=settled[1].value;
    applyAppearance();renderAdmin();renderModelSettings();renderPromptSettings();renderSettings();renderCandidates();
  }else{
    applyAppearance();renderSettings();
  }
}
function renderAll(){applyAppearance();renderSummary();renderCollection();renderInterview();renderTraining();renderCandidates();renderTalentPool();renderKanban();renderProfiles();renderAdmin();renderModelSettings();renderPromptSettings();renderSettings()}
function avatarHtml(user=currentUser,cls=""){const name=(user?.display_name||user?.username||"U").slice(0,1).toUpperCase();return `<span class="avatar ${cls}" style="background:${user?.avatar_color||"#126b61"}">${user?.avatar_url?`<img src="${esc(user.avatar_url)}">`:esc(name)}</span>`}
function renderUserbar(){if(!currentUser)return;const u=$("#userbar");u.style.display="block";u.classList.remove("open");u.innerHTML=`<button class="account-trigger" onclick="toggleAccountMenu(event)"><span class="account-info">${avatarHtml(currentUser)}<span class="account-text"><b>${esc(currentUser.display_name||currentUser.username)}</b><span>${currentUser.role==="manager"?"管理者":"HR"} ${currentUser.title?"/ "+esc(currentUser.title):""}</span></span></span><span class="account-arrow">⌄</span></button><div class="account-menu"><button onclick="document.querySelector('[data-view=settings]').click()">个人设置</button><button onclick="switchUser()">切换用户</button><button onclick="logout()">退出登录</button></div>`;const isMgr=currentUser.role==="manager";["adminNav","modelSettingsNav","promptSettingsNav"].forEach(id=>$("#"+id).style.display=isMgr?"block":"none");$("#settingsNav").style.display="block"}
function toggleAccountMenu(e){e?.stopPropagation();$("#userbar").classList.toggle("open")}
function applyAppearance(){const c=appConfig||{};document.body.classList.toggle("theme-dark",c.ui_theme==="dark");document.documentElement.style.setProperty("--brand",c.ui_primary_color||"#126b61");document.body.style.fontSize=(c.ui_font_size||14)+"px";document.body.style.fontFamily=`"${c.ui_font_family||"Microsoft YaHei"}",Arial,sans-serif`}
function renderSummary(){
  $("#kTotal").textContent=summary.total||0; $("#kActive").textContent=summary.active||0; $("#kScore").textContent=summary.avg_score||0; $("#kPass").textContent=(summary.pass_rate||0)+"%";
  const max=Math.max(1,...Object.values(summary.stages||{}));
  $("#stageBars").innerHTML=PIPELINE.map(s=>`<div class="stage-row"><b>${s}</b><div class="bar"><span style="width:${((summary.stages?.[s]||0)/max)*100}%"></span></div><span>${summary.stages?.[s]||0}</span></div>`).join("");
  const m=summary.metrics||{};
  $("#metrics").innerHTML=[
    ["简历通过率",m.resume_pass_rate],["群面推进率",m.screen_pass_rate],["终面邀约率",m.interview_invite_rate],
    ["终面完成率",m.interview_show_rate],["通过转化率",m.offer_conversion_rate],["淘汰率",m.loss_rate]
  ].map(x=>`<div class="metric"><span class="muted">${x[0]}</span><b>${x[1]||0}%</b></div>`).join("");
  $("#reminders").innerHTML=(summary.reminders||[]).length?summary.reminders.map(r=>`<div class="mini"><b>${r.name}</b><span>${r.message}</span></div>`).join(""):`<div class="muted">暂无需要催办的候选人。</div>`;
  renderPositionBoard();
  renderDaily();
}
function renderPositionBoard(){
  const p=$("#positionBoard"), r=$("#riskBoard"); if(!p||!r)return;
  const positions=summary.metrics?.positions||[];
  p.innerHTML=positions.length?`<table><thead><tr><th>岗位</th><th>总数</th><th>推进中</th><th>均分</th><th>通过率</th><th>Offer率</th><th>人才池</th></tr></thead><tbody>${positions.slice(0,8).map(x=>`<tr><td><b>${esc(x.position||"无")}</b></td><td>${x.total||0}</td><td>${x.active||0}</td><td>${x.avg_score||0}</td><td>${x.pass_rate||0}%</td><td>${x.offer_rate||0}%</td><td>${x.talent_pool||0}</td></tr>`).join("")}</tbody></table>`:`<div class="muted">暂无岗位数据。</div>`;
  const dups=summary.duplicates||[];
  const topLoss=(summary.metrics?.loss_rate||0)>=30?`<div class="mini"><b>淘汰率偏高</b><span>当前淘汰率 ${summary.metrics.loss_rate}% ，建议复盘岗位标准和筛选口径。</span></div>`:"";
  const dupHtml=dups.length?`<div class="mini"><b>疑似重复候选人 ${dups.length} 组</b><span>${dups.slice(0,3).map(d=>`${d.candidates.map(c=>c.name).join(" / ")} (${d.kind})`).join("<br>")}</span></div>`:`<div class="mini"><b>重复识别</b><span>暂无明显重复候选人。</span></div>`;
  const pool=`<div class="mini"><b>人才池</b><span>已沉淀 ${summary.talent_pool_count||0} 位可复联候选人。</span></div>`;
  r.innerHTML=topLoss+dupHtml+pool;
}
function renderDaily(){if(dailyReport===null){$("#dailyReport").innerHTML=`<div class="muted">基础数据已加载，AI 招聘日报正在后台生成...</div>`;return}const r=dailyReport||{};$("#dailyReport").innerHTML=`<div><b>今日概览</b><br>${r.summary||"暂无日报"}</div><div><b>优先跟进</b><br>${(r.priorities||[]).join("<br>")||"暂无"}</div><div><b>建议动作</b><br>${(r.actions||[]).join("<br>")||"暂无"}</div><div><b>风险/卡点</b><br>${[...(r.risks||[]),...(r.bottlenecks||[]).map(x=>"卡点："+x)].join("<br>")||"暂无"}</div><div class="muted">生成方式：${r.method||"未知"}</div>`}
async function loadDailyReport(){try{dailyReport=await api("/api/daily-report")}catch(e){dailyReport={summary:"日报生成失败："+e.message,priorities:[],actions:[],risks:[],bottlenecks:[],method:"异常"}}renderDaily()}
function renderCollection(){const select=$("#collectSource");if(!select)return;const selected=select.value;const enabled=(collectionSources||[]).filter(s=>s.enabled!==false);select.innerHTML=enabled.map(s=>`<option value="${esc(s.id)}">${esc(s.name)}${s.api_url?" · API":" · 手动"}</option>`).join("")||`<option value="local">本地上传</option>`;if(selected&&enabled.some(s=>s.id===selected))select.value=selected;renderCollectionSettings()}
function renderCollectionSettings(){const box=$("#collectionSettings");if(!box)return;box.innerHTML=`<div class="config-box" style="margin-bottom:12px"><h3>个人招聘软件/API</h3><div class="hint">针对招聘网站反爬和数据格式不统一，建议配置标准化后台接口；系统会统一清洗 files、texts 或 candidates 字段，并同步到候选人库。</div><div id="collectSourceForms" class="model-list" style="margin-top:10px">${(collectionSources||[]).map((s,i)=>collectSourceForm(s,i)).join("")}</div><div class="actions" style="margin-top:10px"><button class="btn small ghost" type="button" onclick="addCollectionSource()">+ 新增后台软件</button><button class="btn small primary" type="button" onclick="saveCollectionSources()">保存个人采集配置</button></div></div>`}
function collectSourceForm(s,i){return `<div class="model-card"><div class="model-title"><h3>${esc(s.name||s.id)}</h3><label class="hint"><input type="checkbox" id="csEnabled${i}" ${s.enabled!==false?"checked":""}> 启用</label></div><div class="model-form" style="padding:12px 14px"><div class="model-form-grid"><div class="field"><label>软件名称</label><input id="csName${i}" value="${esc(s.name||"")}"></div><div class="field"><label>来源ID</label><input id="csId${i}" value="${esc(s.id||"")}" placeholder="boss"></div></div><div class="field"><label>API 地址</label><input id="csApi${i}" value="${esc(s.api_url||"")}" placeholder="https://example.com/api/resumes"></div><div class="model-form-grid"><div class="field"><label>请求方式</label><select id="csMethod${i}"><option ${s.method!=="GET"?"selected":""}>POST</option><option ${s.method==="GET"?"selected":""}>GET</option></select></div><div class="field"><label>API Key</label><input id="csKey${i}" type="password" placeholder="${s.has_api_key?"已保存，留空不修改":"可选"}"></div></div><div class="field"><label>请求头 JSON</label><textarea id="csHeaders${i}" class="textarea-short" placeholder='{"X-Tenant":"xxx"}'>${esc(s.headers||"")}</textarea></div><div class="field"><label>请求体 JSON</label><textarea id="csBody${i}" class="textarea-short" placeholder='{"page":1,"page_size":20}'>${esc(s.request_body||"{}")}</textarea></div><div class="actions"><button class="btn small" type="button" onclick="removeCollectionSource(${i})">删除</button></div></div></div>`}
function addCollectionSource(){collectionSources=[...(collectionSources||[]),{id:"custom_"+Date.now().toString(36),name:"新后台软件",enabled:true,api_url:"",method:"POST",headers:"",request_body:"{}"}];renderCollection()}
function removeCollectionSource(i){collectionSources=(collectionSources||[]).filter((_,idx)=>idx!==i);renderCollection()}
function collectCollectionSourcesFromForm(){return (collectionSources||[]).map((s,i)=>({id:$("#csId"+i).value.trim()||s.id,name:$("#csName"+i).value.trim()||s.name,enabled:$("#csEnabled"+i).checked,api_url:$("#csApi"+i).value.trim(),method:$("#csMethod"+i).value,api_key:$("#csKey"+i).value.trim(),headers:$("#csHeaders"+i).value.trim(),request_body:$("#csBody"+i).value.trim()||"{}"}))}
async function saveCollectionSources(){const payload={username:currentUser.username,display_name:currentUser.display_name,department:currentUser.department,title:currentUser.title,phone:currentUser.phone,email:currentUser.email,avatar_url:currentUser.avatar_url,avatar_color:currentUser.avatar_color,bio:currentUser.bio,collection_sources:collectCollectionSourcesFromForm()};const r=await api("/api/account",{method:"POST",body:JSON.stringify(payload)});currentUser=r.user;collectionSources=currentUser.collection_sources||[];renderCollection();toast("个人采集配置已保存")}
function renderAdmin(){if(!currentUser||currentUser.role!=="manager")return;const users=adminOverview?.users||[];const totalActions=users.reduce((s,u)=>s+(u.today_actions||0),0);const totalResumes=users.reduce((s,u)=>s+(u.today_resumes||0),0);$("#adminPanel").innerHTML=`<div class="admin-grid"><div class="metric"><span class="muted">账号数</span><b>${users.length}</b></div><div class="metric"><span class="muted">今日操作</span><b>${totalActions}</b></div><div class="metric"><span class="muted">今日简历相关</span><b>${totalResumes}</b></div><div class="metric"><span class="muted">管理者</span><b>${users.filter(u=>u.role==="manager").length}</b></div></div><table class="admin-table"><thead><tr><th>账号</th><th>角色</th><th>显示名</th><th>今日操作</th><th>今日处理简历</th><th>最近操作</th><th>创建时间</th></tr></thead><tbody>${users.map(u=>`<tr><td><b>${u.username}</b></td><td>${u.role==="manager"?"管理者":"普通HR"}</td><td>${u.display_name||""}</td><td>${u.today_actions||0}</td><td>${u.today_resumes||0}</td><td>${(u.last_action||"").slice(0,19).replace("T"," ")}</td><td>${(u.created_at||"").slice(0,10)}</td></tr>`).join("")}</tbody></table>`}
function renderSettings(){if(!currentUser||currentUser.role!=="manager")return;const c=appConfig||{};const models=c.model_configs||[];const activeId=c.active_model_id;const online=models.filter(m=>m.status==="在线").length;$("#settingsPanel").innerHTML=`<div class="profile-form"><div class="config-box"><h3>模型设置</h3><div class="model-stats"><div class="metric"><span class="muted">总模型数</span><b>${models.length}</b></div><div class="metric"><span class="muted">当前启用</span><b>${models.find(m=>m.id===activeId)?.name||"无"}</b></div><div class="metric"><span class="muted">在线</span><b>${online}</b></div><div class="metric"><span class="muted">自定义</span><b>${models.filter(m=>m.source==="自定义").length}</b></div></div><div class="model-tools"><input id="modelSearch" placeholder="搜索模型名称、模型标识、API 地址..." oninput="renderModelList()"><select id="modelSourceFilter" onchange="renderModelList()"><option value="">全部来源</option><option>环境配置</option><option>自定义</option></select><select id="modelStatusFilter" onchange="renderModelList()"><option value="">全部状态</option><option>在线</option><option>失败</option><option>未测试</option></select><button class="btn primary" onclick="newModel()">+ 添加模型</button></div><div id="modelList" class="model-list"></div></div><div class="config-box"><h3 id="modelFormTitle">新增模型</h3><div class="model-form"><div class="model-form-grid"><div class="field"><label>显示名称</label><input id="modelName" placeholder="例如：qwen3.6-27b(A100)"></div><div class="field"><label>模型标识</label><input id="modelIdName" placeholder="例如：qwen36-27b"></div></div><div class="field"><label>API 地址</label><input id="modelBaseUrl" placeholder="可填 http://host:port/v1"></div><div class="model-form-grid"><div class="field"><label>API Key</label><input id="modelApiKey" type="password" placeholder="本地模型可留空，编辑时留空表示不修改"></div><div class="field"><label>环境变量名</label><input id="modelKeyEnv" value="OPENAI_API_KEY"></div></div><div class="model-form-grid"><div class="field"><label>Temperature</label><input id="modelTemp" type="number" step="0.1" value="0.2"></div><div class="field"><label>Timeout 秒</label><input id="modelTimeout" type="number" value="120"></div></div><div class="model-form-grid"><div class="field"><label>Thinking / Provider</label><input id="modelThinking" placeholder="OpenAI / DashScope / 不关闭"></div><div class="field"><label>来源</label><select id="modelSource"><option>自定义</option><option>环境配置</option></select></div></div><div class="actions"><button class="btn primary" onclick="saveModel()">保存模型</button><button class="btn" onclick="newModel()">清空表单</button></div></div></div><div class="config-box"><h3>提示词管理</h3><div class="field"><label>AI解析简历提示词</label><textarea id="cfgParsePrompt" class="textarea-tall">${c.resume_parse_prompt||""}</textarea></div><div class="field"><label>AI综合评审提示词</label><textarea id="cfgReviewPrompt" class="textarea-tall">${c.candidate_review_prompt||""}</textarea></div><div class="hint">提示词要求模型输出 JSON。字段可自定义，但建议保留分析摘要、风险点、适配度和决策建议。</div></div><div class="config-box"><h3>系统规则</h3><div class="form-grid"><div class="field"><label>启用大模型</label><select id="cfgEnabled"><option value="false">关闭，使用本地兜底</option><option value="true" ${c.llm_enabled?"selected":""}>开启，使用当前模型</option></select></div><div class="field"><label>提醒小时数</label><input id="cfgReminder" type="number" min="1" value="${c.reminder_hours||24}"></div></div><div class="field"><label>管理者注册码</label><input id="cfgManagerCode" value="${c.manager_register_code||""}"></div></div><div class="save-bar"><div class="save-note">模型配置、提示词和系统规则都会保存到 data/config/app_config.json。</div><button class="btn primary" onclick="saveSettings()">保存系统设置</button></div></div>`;renderModelList();if(editingModelId){fillModelForm(models.find(m=>m.id===editingModelId))}}
function renderCandidates(){
  const sf=$("#stageFilter"); if(sf.children.length===1) PIPELINE.forEach(s=>sf.insertAdjacentHTML("beforeend",`<option value="${s}">${s}</option>`));
  const q=$("#search").value.trim().toLowerCase(), st=sf.value;
  const rows=candidates.filter(c=>(!st||c.stage===st)&&(!q||JSON.stringify(c).toLowerCase().includes(q)));
  $("#candidateRows").innerHTML=rows.map(c=>`<tr onclick="selectCandidate('${c.id}')"><td><b>${c.name||""}</b><div class="muted">${c.phone||c.email||""}</div></td><td>${c.position||""}</td><td>${pill(c.stage)}</td><td>${scorePill(c.match?.score)}</td><td>${c.source||""}</td><td>${(c.updated_at||"").slice(0,16).replace("T"," ")}</td></tr>`).join("");
  if(selectedCandidate) renderCandidateDetail(selectedCandidate);
}
function renderTalentPool(){
  const box=$("#talentPoolList"); if(!box)return;
  const q=($("#talentSearch")?.value||"").trim().toLowerCase();
  const rows=candidates.filter(c=>c.talent_pool&&(!q||JSON.stringify(c).toLowerCase().includes(q)));
  box.innerHTML=rows.length?`<table><thead><tr><th>姓名</th><th>岗位</th><th>标签</th><th>最近阶段</th><th>下次复联</th><th>备注</th></tr></thead><tbody>${rows.map(c=>`<tr onclick="document.querySelector('[data-view=candidates]').click();selectCandidate('${c.id}')"><td><b>${esc(c.name||"")}</b><div class="muted">${esc(c.phone||c.email||"")}</div></td><td>${esc(c.position||"")}</td><td>${arr(c.talent_tags).map(x=>`<span class="chip">${esc(x)}</span>`).join(" ")}</td><td>${pill(c.stage)}</td><td>${esc(c.next_follow_time||"")}</td><td>${esc((c.notes||"").slice(0,40))}</td></tr>`).join("")}</tbody></table>`:`<div class="muted">暂无人才池候选人。可在候选人详情里勾选“加入人才池”。</div>`;
}
function renderInterview(){
  const list=$("#interviewCandidateList"); if(!list)return;
  const q=($("#interviewSearch")?.value||"").trim().toLowerCase();
  const rows=candidates.filter(c=>!q||JSON.stringify(c).toLowerCase().includes(q));
  if(!selectedInterviewId&&rows[0]) selectedInterviewId=rows[0].id;
  list.innerHTML=rows.map(c=>`<button class="${selectedInterviewId===c.id?"active":""}" onclick="selectInterviewCandidate('${c.id}')"><b>${esc(c.name||"未命名")}</b><div class="muted">${esc(c.position||"无")} · ${c.match?.score??"未评分"}分</div><div class="hint">${esc((c.education||"").slice(0,18))} ${esc((c.skills||[]).slice(0,3).join("、"))}</div><div class="actions" style="margin-top:8px"><span class="pill ${c.interview_plan?'ok':'warn'}">${c.interview_plan?'已生成':'待生成'}</span></div></button>`).join("")||"<div class='muted'>暂无候选人。</div>";
  renderInterviewPanel();
}
function selectInterviewCandidate(id){selectedInterviewId=id;renderInterview()}
function setInterviewTab(tab){interviewTab=tab;renderInterviewPanel()}
function currentInterviewCandidate(){return candidates.find(c=>c.id===selectedInterviewId)}
function renderInterviewPanel(){
  const c=currentInterviewCandidate(), panel=$("#interviewPanel"); if(!panel)return;
  $("#resumeTabBtn")?.classList.toggle("blue",interviewTab==="resume");
  $("#planTabBtn")?.classList.toggle("blue",interviewTab==="plan");
  if(!c){panel.innerHTML="请选择左侧候选人。";return}
  $("#interviewTitle").textContent=`${c.name||"候选人"} · 智能面试`;
  panel.classList.remove("muted");
  panel.innerHTML=interviewTab==="resume"?renderResumePreview(c):renderInterviewPlan(c.interview_plan);
}
function renderResumePreview(c){const skills=(c.skills||[]).map(s=>`<span class="chip">${esc(s)}</span>`).join("");return `<div class="template"><h4>${esc(c.name||"候选人")} - ${esc(c.position||"无")}</h4><p><b>阶段：</b>${esc(c.stage||"")}　<b>匹配：</b>${c.match?.score??"未评分"}　<b>来源：</b>${esc(c.source||"")}</p><p><b>联系方式：</b>${esc(c.phone||"")} ${esc(c.email||"")}</p><p><b>教育：</b>${esc(c.education||"无")}</p><p><b>技能：</b></p><div class="chips">${skills||"<span class='muted'>未识别</span>"}</div><h4 style="margin-top:12px">简历原文</h4><pre>${esc((c.resume_text||"暂无简历文本").slice(0,9000))}</pre></div>`}
function renderInterviewPlan(plan){if(!plan)return `<div class="muted">尚未生成面试方案。点击右上角“生成面试方案”，系统会根据该候选人的简历、岗位标准、匹配分、风险点生成定制问题。</div>`;const sections=(plan.question_sections||[]).map(sec=>`<div class="template" style="margin-top:10px"><h4>${esc(sec.title||"问题模块")}</h4><p class="muted">${esc(sec.purpose||"")}</p><ol>${(sec.questions||[]).map(q=>`<li>${esc(q)}</li>`).join("")}</ol></div>`).join("");const rubric=(plan.scoring_rubric||[]).map(r=>`<tr><td>${esc(r.dimension||"")}</td><td>${esc(r.weight??"")}</td><td>${esc(r.pass_signal||"")}</td></tr>`).join("");const schedule=(plan.schedule||[]).map(x=>`<span class="chip">${esc(x.part||"环节")} ${esc(x.minutes??"")}min</span>`).join("");return `<div class="detail"><div class="template"><h4>${esc(plan.method||"智能面试方案")}</h4><p><b>面试目标：</b>${esc(plan.interview_goal||"")}</p><p><b>候选人摘要：</b>${esc(plan.candidate_summary||"")}</p><p><b>重点关注：</b>${arr(plan.focus_areas).map(esc).join("；")||"无"}</p><div class="chips">${schedule}</div></div>${sections}<div class="template"><h4>项目深挖与风险验证</h4><p><b>项目深挖：</b>${arr(plan.project_deep_dive).map(esc).join("；")||"无"}</p><p><b>风险验证：</b>${arr(plan.risk_checks).map(esc).join("；")||"无"}</p></div><div class="template"><h4>评分 Rubric</h4><table><thead><tr><th>维度</th><th>权重</th><th>通过信号</th></tr></thead><tbody>${rubric}</tbody></table></div><div class="template"><h4>决策信号</h4><pre>${esc(pretty(plan.decision_signals||{}))}</pre><p><b>面试官记录：</b>${esc(plan.interviewer_notes||"")}</p></div></div>`}
function renderTraining(){
  const list=$("#trainingExamList"); if(!list)return;
  const exams=trainingData.exams||[], results=trainingData.results||[], s=trainingData.summary||{};
  $("#tExamCount").textContent=s.exam_count||exams.length||0; $("#tResultCount").textContent=s.result_count||results.length||0; $("#tAvgScore").textContent=s.avg_score||0; $("#tPassRate").textContent=(s.pass_rate||0)+"%";
  if(!selectedTrainingExamId&&exams[0]) selectedTrainingExamId=exams[0].id;
  list.innerHTML=exams.map(e=>`<button class="${selectedTrainingExamId===e.id?"active":""}" onclick="selectTrainingExam('${esc(e.id)}')"><b>${esc(e.title||"未命名试卷")}</b><div class="muted">${esc(e.topic||"通用培训")} · ${esc(e.difficulty||"中等")} · ${esc(e.total_score||100)}分</div><div class="hint">${esc((e.created_at||"").slice(0,16).replace("T"," "))} · ${esc(e.method||"")}</div></button>`).join("")||"<div class='muted'>暂无试卷，请先生成。</div>";
  renderTrainingPanel();
}
function selectTrainingExam(id){selectedTrainingExamId=id;renderTraining()}
function setTrainingTab(tab){trainingTab=tab;renderTrainingPanel()}
function currentTrainingExam(){return (trainingData.exams||[]).find(e=>e.id===selectedTrainingExamId)}
function renderTrainingPanel(){
  const panel=$("#trainingPanel"); if(!panel)return;
  $("#examPreviewTab")?.classList.toggle("blue",trainingTab==="preview");
  $("#examResultTab")?.classList.toggle("blue",trainingTab==="results");
  const exam=currentTrainingExam();
  if(!exam){panel.innerHTML="请选择左侧试卷，或先生成一份试卷。";return}
  $("#trainingTitle").textContent=`${exam.title||"考试"} · 考试管理`;
  panel.classList.remove("muted");
  panel.innerHTML=trainingTab==="preview"?renderExamPreview(exam):renderExamResults(exam);
}
function renderExamPreview(exam){
  const sections=(exam.sections||[]).map(x=>`<span class="chip">${esc(x.name||"模块")} ${esc(x.score??"")}分</span>`).join("");
  const qs=(exam.questions||[]).map((q,i)=>`<div class="template"><h4>${i+1}. ${esc(q.type||"题目")} · ${esc(q.score??"")}分</h4><p>${esc(q.stem||"")}</p>${(q.options||[]).length?`<ol type="A">${q.options.map(o=>`<li>${esc(o)}</li>`).join("")}</ol>`:""}<p class="muted"><b>参考答案：</b>${esc(q.answer||"")}</p><p class="muted"><b>解析：</b>${esc(q.analysis||"")}</p></div>`).join("");
  return `<div class="detail"><div class="template"><h4>${esc(exam.title||"考试")}</h4><p><b>对象：</b>${esc(exam.target||"")}　<b>主题：</b>${esc(exam.topic||"")}　<b>难度：</b>${esc(exam.difficulty||"")}</p><p><b>时长：</b>${esc(exam.duration||60)} 分钟　<b>总分：</b>${esc(exam.total_score||100)}　<b>通过分：</b>${esc(exam.pass_score||60)}</p><p>${esc(exam.description||"")}</p><div class="chips">${sections}</div></div>${qs}<div class="template"><h4>评分规则</h4><ol>${arr(exam.scoring_rules).map(x=>`<li>${esc(x)}</li>`).join("")}</ol></div></div>`;
}
function renderExamResults(exam){
  const rows=(trainingData.results||[]).filter(r=>r.exam_id===exam.id);
  return `<div class="detail"><div class="template"><h4>录入考试结果</h4><div class="model-form-grid"><div class="field"><label>考生姓名</label><input id="resultStudent" placeholder="张三"></div><div class="field"><label>部门</label><input id="resultDept" placeholder="技术部"></div><div class="field"><label>得分</label><input id="resultScore" type="number" min="0" max="${esc(exam.total_score||100)}"></div><div class="field"><label>状态</label><select id="resultStatus"><option>自动判断</option><option>通过</option><option>未通过</option><option>缺考</option><option>待补考</option></select></div></div><div class="field" style="margin-top:10px"><label>答题详情/备注</label><textarea id="resultComment" class="textarea-short" placeholder="可粘贴考生答案、答题详情、阅卷意见或补考安排"></textarea></div><div class="actions" style="margin-top:10px"><button class="btn blue" onclick="gradeTrainingResult()">AI/规则评分并保存</button><button class="btn primary" onclick="saveTrainingResult()">手动保存成绩</button></div><div id="gradeResult" class="muted" style="margin-top:10px">标准答案模板 + 关键场景人工复核；复杂语义题会提示需要复核。</div></div><table><thead><tr><th>考生</th><th>部门</th><th>得分</th><th>状态</th><th>提交时间</th><th>评分方式</th><th>备注</th></tr></thead><tbody>${rows.map(r=>`<tr><td><b>${esc(r.student_name||"")}</b></td><td>${esc(r.department||"")}</td><td>${esc(r.score||0)} / ${esc(r.total_score||exam.total_score||100)}</td><td>${pill(r.status)}</td><td>${esc((r.submitted_at||"").slice(0,16).replace("T"," "))}</td><td>${esc(r.grading?.method||"手动录入")}</td><td>${esc(r.comment||"")}</td></tr>`).join("")||"<tr><td colspan='7' class='muted'>暂无考试记录。</td></tr>"}</tbody></table></div>`;
}
function pill(v){let cls=v==="淘汰"||v==="未通过"?"bad":(v==="通过"?"ok":"warn");return `<span class="pill ${cls}">${v||"未同步"}</span>`}
function scorePill(v){if(v===undefined||v==="")return '<span class="muted">未评分</span>';let cls=v>=80?"ok":(v>=60?"warn":"bad");return `<span class="pill ${cls}">${v}</span>`}
function selectCandidate(id){selectedCandidate=candidates.find(c=>c.id===id);renderCandidateDetail(selectedCandidate)}
function renderCandidateDetail(c){
  const skills=(c.skills||[]).map(s=>`<span class="chip">${s}</span>`).join("");
  const m=c.match||{};
  $("#candidateDetail").innerHTML=`<div class="row"><b>姓名</b><span>${c.name||""}</span></div><div class="row"><b>岗位</b><span>${c.position||""}</span></div><div class="row"><b>联系方式</b><span>${c.phone||""} ${c.email||""}</span></div><div class="row"><b>解析状态</b><span>${parseStatusPill(c)}</span></div><div class="row"><b>阶段</b><span><select id="stageSelect" onchange="syncStageSelects('stageSelect')">${PIPELINE.map(s=>`<option ${s===c.stage?"selected":""}>${s}</option>`).join("")}</select></span></div><div class="row"><b>技能</b><div class="chips">${skills||"<span class='muted'>未识别</span>"}</div></div><div class="row"><b>匹配分</b><span>${scorePill(m.score)} ${m.suggestion||""}</span></div>${renderScoreDimensions(m)}${renderCandidateProfile(c)}<div class="row"><b>优势</b><span>${(m.strengths||[]).join("<br>")||"未评分"}</span></div><div class="row"><b>风险点</b><span>${(m.risks||[]).join("、")||"暂无"}</span></div><div class="row"><b>追问</b><span>${(m.questions||[]).join("<br>")||"暂无"}</span></div>${renderWorkflowState(c)}${renderManualReview(c)}${renderInterviewerPanel(c)}<div class="agent-card"><h3>招聘智能体</h3><div class="agent-actions"><button class="btn small primary" onclick="analyzeResume('${c.id}')">AI解析简历</button><button class="btn small blue" onclick="agentReview('${c.id}')">AI综合评审</button><button class="btn small ghost" onclick="loadNextAction('${c.id}')">推荐下一步</button><button class="btn small ghost" onclick="applyNextAction('${c.id}')">应用推荐下一步</button></div><div id="agentResult" style="margin-top:10px">${renderAgentState(c)}</div></div><div class="agent-card"><h3>群面/作业智能评分</h3><textarea id="groupText" placeholder="粘贴群面记录、面试官评价或群面作业内容"></textarea><div class="actions"><button class="btn small primary" onclick="groupScore('${c.id}')">AI群面评分并自动同步</button></div><div id="groupResult">${renderGroupState(c)}</div></div><div><b>企业微信话术</b><div id="templates" class="detail muted">正在生成话术...</div></div>${renderCommunicationPanel(c)}<div><b>HR备注</b><textarea id="notes">${c.notes||""}</textarea></div><div class="actions"><button class="btn primary" onclick="scoreOne('${c.id}')">规则重新评分</button><button class="btn blue" onclick="agentReview('${c.id}')">AI综合评审</button><button class="btn" onclick="saveDetail('${c.id}')">保存画像、沟通、状态和评分</button></div>`;
  loadTemplates(c.id);
}
function parseStatusPill(c){if(c.parse_status==="filename_only")return `<span class="pill warn" title="${esc(c.parse_error||"")}">仅文件名导入，待解析</span>`;return `<span class="pill ok">已解析</span>`}
function listText(v){return Array.isArray(v)?v.join("、"):(v||"")}
function renderCandidateProfile(c){return `<div class="agent-card"><h3>候选人画像与人才池</h3><div class="model-form-grid"><div class="field"><label>求职意向</label><input id="jobIntention" value="${esc(c.job_intention||"")}" placeholder="岗位方向、行业偏好"></div><div class="field"><label>薪资期望</label><input id="salaryExpectation" value="${esc(c.salary_expectation||c.expected_salary||"")}" placeholder="例如 15K / 200元天"></div><div class="field"><label>可到岗时间</label><input id="availability" value="${esc(c.availability||c.arrival_time||"")}" placeholder="例如 1周内 / 2026-07-15"></div><div class="field"><label>地点接受度</label><input id="locationPreference" value="${esc(c.location_preference||"")}" placeholder="城市、远程、通勤要求"></div><div class="field"><label>稳定性风险</label><input id="stabilityRisk" value="${esc(c.stability_risk||"")}" placeholder="空窗、频繁跳槽、到岗不确定等"></div><div class="field"><label>负责人</label><input id="candidateOwner" value="${esc(c.owner||currentUser?.display_name||currentUser?.username||"")}" placeholder="负责 HR"></div><div class="field"><label>下次跟进</label><input id="nextFollowTime" value="${esc(c.next_follow_time||"")}" placeholder="YYYY-MM-DD HH:mm"></div><div class="field"><label>跟进优先级</label><select id="followPriority"><option ${sel(c.follow_up_priority,"普通")}>普通</option><option ${sel(c.follow_up_priority,"高")}>高</option><option ${sel(c.follow_up_priority,"低")}>低</option></select></div></div><label class="hint" style="margin-top:8px"><input id="talentPool" type="checkbox" ${c.talent_pool?"checked":""}> 加入人才池，可后续复联</label><div class="field" style="margin-top:8px"><label>人才池标签</label><input id="talentTags" value="${esc(listText(c.talent_tags))}" placeholder="算法、前端、校招、可复联，用顿号或逗号分隔"></div></div>`}
function renderCommunicationPanel(c){const logs=c.communication_logs||[];return `<div class="agent-card"><h3>沟通记录与跟进</h3><div class="model-form-grid"><div class="field"><label>沟通方式</label><select id="commType"><option>电话</option><option>微信</option><option>邮件</option><option>面谈</option><option>系统备注</option></select></div><div class="field"><label>沟通结果</label><input id="commResult" placeholder="已约面 / 待回复 / 不合适 / 需复联"></div></div><textarea id="commContent" class="textarea-short" placeholder="记录候选人反馈、关键顾虑、薪资/地点/到岗信息"></textarea><div class="actions" style="margin-top:8px"><button class="btn small primary" type="button" onclick="addCommunicationLog()">添加沟通记录</button></div><div id="commTimeline" class="detail" style="margin-top:10px">${logs.length?logs.map(l=>`<div class="mini"><b>${esc(l.time||"")} · ${esc(l.type||"沟通")}</b><span>${esc(l.result||"")}<br>${esc(l.content||"")}</span></div>`).join(""):"<div class='muted'>暂无沟通记录。</div>"}</div></div>`}
function renderInterviewerPanel(c){const items=c.interviewer_reviews||[];return `<div class="agent-card"><h3>面试官协同评价</h3><div class="model-form-grid"><div class="field"><label>面试官</label><input id="reviewerName" placeholder="面试官姓名"></div><div class="field"><label>评分</label><input id="reviewerScore" type="number" min="0" max="100" placeholder="0-100"></div></div><textarea id="reviewerComment" class="textarea-short" placeholder="面试评价、通过建议、关注风险"></textarea><div class="actions" style="margin-top:8px"><button class="btn small primary" type="button" onclick="addInterviewerReview()">添加面试官评价</button></div><div id="reviewerList" class="detail" style="margin-top:10px">${items.length?items.map(x=>`<div class="mini"><b>${esc(x.reviewer||"面试官")} · ${esc(x.score??"未评分")}</b><span>${esc(x.comment||"")}<br>${esc((x.time||"").slice(0,16).replace("T"," "))}</span></div>`).join(""):"<div class='muted'>暂无面试官评价。</div>"}</div></div>`}
function renderScoreDimensions(m){const dims=m.dimensions||[];if(!dims.length)return "";return `<div class="agent-card"><h3>多维度评分依据</h3><div class="metrics">${dims.map(d=>`<div class="metric"><span class="muted">${esc(d.name)}</span><b>${esc(d.score??0)}</b><small>${arr(d.evidence).map(esc).join("、")||"暂无证据"}</small></div>`).join("")}</div><div class="hint">评分由岗位标准权重、候选人简历信号和风险关键词共同生成；启用 AI 评审后可叠加多指标复核。</div></div>`}
function renderManualReview(c){
  const g=c.group_interview||{}, a=c.assignment||{}, h=c.hr_interview||{};
  return `<div class="agent-card"><h3>HR人工评分与状态流转</h3><div class="review-form"><label class="hint"><input id="autoWorkflow" type="checkbox" checked> 保存后按评分和环节状态自动同步流程</label><div class="review-line"><b>当前阶段</b><select id="stageQuickSelect" onchange="syncStageSelects('stageQuickSelect')">${PIPELINE.map(s=>`<option ${s===c.stage?"selected":""}>${s}</option>`).join("")}</select></div><div class="review-line"><b>最终结果</b><select id="finalResult"><option ${sel(c.final_result,"待定")}>待定</option><option ${sel(c.final_result,"通过")}>通过</option><option ${sel(c.final_result,"淘汰")}>淘汰</option><option ${sel(c.final_result,"候补")}>候补</option></select></div>${manualReviewRow("group","群面",g)}${manualReviewRow("assignment","作业",a)}${manualReviewRow("hr","HR面",h)}<div class="review-summary"><div class="metric"><span class="muted">简历匹配</span><b>${mScore(c.match?.score)}</b></div><div class="metric"><span class="muted">流程均分</span><b id="manualAvg">${manualAverage(c)}</b></div><div class="metric"><span class="muted">当前结论</span><b>${c.final_result||"待定"}</b></div></div><div class="hint">人工分数建议填 0-100；点击保存后会写入候选人档案、数据看板和 CSV 导出。</div></div></div>`;
}
function renderWorkflowState(c){
  const w=c.workflow||{}, n=w.next_action||{}, rules=w.rules||appConfig?.workflow_rules||{};
  const ruleText=[["简历线",rules.resume_pass_score],["群面线",rules.group_pass_score],["作业线",rules.assignment_pass_score],["HR线",rules.hr_pass_score]].filter(x=>x[1]!==undefined).map(x=>`${x[0]} ${x[1]}`).join(" / ");
  return `<div class="agent-card"><h3>自动化安排</h3><div class="template"><p><b>下一步：</b>${n.action||"等待流程动作"}</p><p><b>原因：</b>${w.last_reason||n.reason||"尚未触发自动化规则"}</p><p><b>建议截止：</b>${n.due_at?(n.due_at||"").slice(0,16).replace("T"," "):"暂无"}</p><p><b>最近触发：</b>${w.last_trigger||"暂无"}</p><p><b>当前分数线：</b>${ruleText||"使用默认规则"}</p></div></div>`;
}
function syncStageSelects(sourceId){const v=$("#"+sourceId)?.value;["stageSelect","stageQuickSelect"].forEach(id=>{if(id!==sourceId&&$("#"+id))$("#"+id).value=v})}
function manualReviewRow(prefix,label,data={}){
  return `<div class="review-line"><b>${label}</b><div class="review-grid"><select id="${prefix}Status"><option ${sel(data.status,"待定")}>待定</option><option ${sel(data.status,"待安排")}>待安排</option><option ${sel(data.status,"已安排")}>已安排</option><option ${sel(data.status,"已完成")}>已完成</option><option ${sel(data.status,"已评分")}>已评分</option><option ${sel(data.status,"通过")}>通过</option><option ${sel(data.status,"淘汰")}>淘汰</option><option ${sel(data.status,"缺席")}>缺席</option></select><input id="${prefix}Score" type="number" min="0" max="100" step="1" value="${esc(data.score??"")}" placeholder="分数"></div></div>`;
}
function sel(value,option){return String(value||"")===option?"selected":""}
function mScore(v){return v===undefined||v===""?"未评分":v}
function manualAverage(c){const vals=[c.group_interview?.score,c.assignment?.score,c.hr_interview?.score].map(Number).filter(v=>Number.isFinite(v));return vals.length?Math.round(vals.reduce((s,v)=>s+v,0)/vals.length):"未评分"}
function splitTags(v){return String(v||"").split(/[、,，\n]+/).map(x=>x.trim()).filter(Boolean)}
function addCommunicationLog(){if(!selectedCandidate)return;const content=$("#commContent").value.trim(), result=$("#commResult").value.trim();if(!content&&!result){toast("请先填写沟通内容");return}selectedCandidate.communication_logs=[...(selectedCandidate.communication_logs||[]),{time:new Date().toISOString(),type:$("#commType").value,result,content,actor:currentUser?.display_name||currentUser?.username||""}];renderCandidateDetail(selectedCandidate);toast("沟通记录已加入，记得保存")}
function addInterviewerReview(){if(!selectedCandidate)return;const reviewer=$("#reviewerName").value.trim(), comment=$("#reviewerComment").value.trim(), score=scoreValue("#reviewerScore");if(!reviewer&&!comment){toast("请填写面试官或评价内容");return}selectedCandidate.interviewer_reviews=[...(selectedCandidate.interviewer_reviews||[]),{time:new Date().toISOString(),reviewer:reviewer||"面试官",score,comment}];renderCandidateDetail(selectedCandidate);toast("面试官评价已加入，记得保存")}
function renderAgentState(c){const parts=[];if(c.ai_parse)parts.push(renderParseAnalysis(c.ai_parse));if(c.agent_review)parts.push(renderReviewAnalysis(c.agent_review));return parts.join("")||"<span class='muted'>点击上方按钮，让智能体解析简历、评审候选人并推荐下一步。</span>"}
function collapsibleSection(title,body){return `<div class="collapsible"><button class="collapse-head" type="button" onclick="toggleCollapse(this)" aria-expanded="true"><span>${title}</span><span class="collapse-arrow">⌄</span></button><div class="collapse-body">${body}</div></div>`}
function toggleCollapse(btn){const box=btn.closest(".collapsible");const collapsed=box.classList.toggle("collapsed");btn.setAttribute("aria-expanded",String(!collapsed))}
function renderParseAnalysis(p){return collapsibleSection("LLM结构化解析",`<div class="template"><h4>${p.method||"AI解析"}</h4><p>${p.analysis_summary||p.summary||p.conclusion||"暂无分析摘要"}</p><p><b>教育分析：</b>${p.education_analysis||"无"}</p><p><b>技能分析：</b>${p.skill_analysis||"无"}</p><p><b>项目分析：</b>${p.project_analysis||"无"}</p><p><b>经历分析：</b>${p.experience_analysis||"无"}</p><p><b>风险点：</b>${arr(p.risk_points).join("、")||"无"}</p><p><b>建议追问：</b>${arr(p.follow_up_questions||p.questions).join("；")||"无"}</p><p><b>结论：</b>${p.conclusion||"无"}</p></div>`)}
function renderReviewAnalysis(r){const score=r.suitability_score??r.score??0;return collapsibleSection("岗位适配度评审",`<div class="template"><h4>${r.method||"AI评审"} · ${r.suitability_level||""}</h4><p><b>适配度：</b>${score} / 100</p><p><b>适配结论：</b>${r.fit_summary||r.reason||"无"}</p><p><b>匹配证据：</b>${arr(r.matched_evidence||r.strengths).join("；")||"无"}</p><p><b>差距/风险：</b>${arr(r.gap_risks||r.risks).join("；")||"无"}</p><p><b>建议追问：</b>${arr(r.questions).join("；")||"无"}</p><p><b>决策建议：</b>${r.decision||"无"}</p></div>`)}
function arr(v){return Array.isArray(v)?v:(v?[v]:[])}
function renderGroupState(c){return c.group_ai_review?`<pre>${pretty(c.group_ai_review)}</pre>`:"<span class='muted'>输入群面记录后可自动给出维度分、总分和推进建议。</span>"}
function pretty(v){return JSON.stringify(v,null,2)}
async function loadTemplates(id){const t=await api(`/api/candidates/${id}/templates`);$("#templates").innerHTML=Object.entries(t).map(([k,v])=>`<div class="template"><button class="copy" onclick="copyText(this)">复制</button><h4>${k}</h4><pre>${v}</pre></div>`).join("")}
function copyText(btn){const text=btn.parentElement.querySelector("pre").textContent;navigator.clipboard?.writeText(text);toast("话术已复制")}
async function saveDetail(id){
  const stage=$("#stageQuickSelect")?.value||$("#stageSelect").value;
  const payload={
    stage,
    notes:$("#notes").value,
    final_result:$("#finalResult")?.value||"待定",
    job_intention:$("#jobIntention")?.value.trim()||"",
    salary_expectation:$("#salaryExpectation")?.value.trim()||"",
    availability:$("#availability")?.value.trim()||"",
    location_preference:$("#locationPreference")?.value.trim()||"",
    stability_risk:$("#stabilityRisk")?.value.trim()||"",
    owner:$("#candidateOwner")?.value.trim()||"",
    next_follow_time:$("#nextFollowTime")?.value.trim()||"",
    follow_up_priority:$("#followPriority")?.value||"普通",
    talent_pool:$("#talentPool")?$("#talentPool").checked:false,
    talent_tags:splitTags($("#talentTags")?.value||""),
    communication_logs:selectedCandidate?.communication_logs||[],
    interviewer_reviews:selectedCandidate?.interviewer_reviews||[],
    group_interview:{status:$("#groupStatus")?.value||"待定",score:scoreValue("#groupScore")},
    assignment:{status:$("#assignmentStatus")?.value||"待定",score:scoreValue("#assignmentScore")},
    hr_interview:{status:$("#hrStatus")?.value||"待定",score:scoreValue("#hrScore")},
    auto_workflow:$("#autoWorkflow")?$("#autoWorkflow").checked:false
  };
  await api(`/api/candidates/${id}`,{method:"POST",body:JSON.stringify(payload)});
  toast("已保存状态、HR评分和备注");
  await refresh();
  selectedCandidate=candidates.find(c=>c.id===id);
  renderCandidateDetail(selectedCandidate);
}
function scoreValue(selector){const raw=$(selector)?.value;if(raw===""||raw===undefined)return "";const n=Number(raw);return Number.isFinite(n)?Math.max(0,Math.min(100,Math.round(n))):""}
async function scoreOne(id){await api(`/api/candidates/${id}/score`,{method:"POST",body:JSON.stringify({profile_id:selectedProfile?.id,auto_workflow:true})});toast("评分已更新，流程已自动检查");await refresh();selectedCandidate=candidates.find(c=>c.id===id);renderCandidateDetail(selectedCandidate)}
async function analyzeResume(id){$("#agentResult").innerHTML="<span class='muted'>AI 正在解析简历...</span>";await api(`/api/candidates/${id}/analyze-resume`,{method:"POST",body:JSON.stringify({profile_id:selectedProfile?.id})});toast("AI简历解析完成");await refresh();selectedCandidate=candidates.find(c=>c.id===id);renderCandidateDetail(selectedCandidate)}
async function agentReview(id){$("#agentResult").innerHTML="<span class='muted'>AI 正在综合评审...</span>";await api(`/api/candidates/${id}/agent-review`,{method:"POST",body:JSON.stringify({profile_id:selectedProfile?.id,auto_workflow:true})});toast("AI综合评审完成，流程已自动检查");await refresh();selectedCandidate=candidates.find(c=>c.id===id);renderCandidateDetail(selectedCandidate)}
async function generateInterviewPlan(){const c=currentInterviewCandidate();if(!c){toast("请先选择候选人");return}$("#interviewPanel").innerHTML="<span class='muted'>AI 正在生成面试方案...</span>";const updated=await api(`/api/candidates/${c.id}/interview-plan`,{method:"POST",body:JSON.stringify({profile_id:selectedProfile?.id})});toast("面试方案已生成");await refresh();selectedInterviewId=updated.id;interviewTab="plan";renderInterview()}
function examPayload(){return {title:$("#examTitle").value.trim(),target:$("#examTarget").value.trim(),topic:$("#examTopic").value.trim(),difficulty:$("#examDifficulty").value,duration:Number($("#examDuration").value||60),total_score:Number($("#examTotalScore").value||100),pass_score:Number($("#examPassScore").value||60),skills:$("#examSkills").value.trim(),description:$("#examDescription").value.trim()}}
function clearExamForm(){["examTitle","examTarget","examTopic","examSkills","examDescription"].forEach(id=>$("#"+id).value="");$("#examDifficulty").value="中等";$("#examDuration").value=60;$("#examTotalScore").value=100;$("#examPassScore").value=60}
async function generateTrainingExam(){const payload=examPayload();if(!payload.title&&!payload.topic){toast("请先填写考试名称或考试主题");return}$("#trainingPanel").innerHTML="<span class='muted'>AI 正在生成试卷...</span>";const exam=await api("/api/training/exams/generate",{method:"POST",body:JSON.stringify(payload)});selectedTrainingExamId=exam.id;trainingTab="preview";toast("试卷已生成");await refresh()}
async function saveTrainingResult(){const exam=currentTrainingExam();if(!exam){toast("请先选择试卷");return}const score=Number($("#resultScore").value||0);const status=$("#resultStatus").value==="自动判断"?"":$("#resultStatus").value;const payload={exam_id:exam.id,exam_title:exam.title,student_name:$("#resultStudent").value.trim(),department:$("#resultDept").value.trim(),score,total_score:exam.total_score||100,pass_score:exam.pass_score||60,status,comment:$("#resultComment").value.trim()};if(!payload.student_name){toast("请填写考生姓名");return}await api("/api/training/results",{method:"POST",body:JSON.stringify(payload)});toast("考试结果已保存");trainingTab="results";await refresh()}
async function gradeTrainingResult(){const exam=currentTrainingExam();if(!exam){toast("请先选择试卷");return}const student=$("#resultStudent").value.trim();const answers=$("#resultComment").value.trim();if(!student){toast("请填写考生姓名");return}if(!answers){toast("请粘贴考生答案或答题详情");return}$("#gradeResult").innerHTML="<span class='muted'>正在根据标准答案和考生答案评分...</span>";const r=await api("/api/training/results/grade",{method:"POST",body:JSON.stringify({exam_id:exam.id,exam_title:exam.title,student_name:student,department:$("#resultDept").value.trim(),answers,comment:answers})});const g=r.grading||{};$("#gradeResult").innerHTML=`<div class="template"><b>评分完成：</b>${esc(r.score)} / ${esc(r.total_score)}，${pill(r.status)}<br><span class="muted">方式：${esc(g.method||"智能评分")}；${g.review_required?"建议人工复核关键题":"无需强制复核"}</span></div>`;toast("考试评分已保存");trainingTab="results";await refresh()}
async function loadNextAction(id){const r=await api(`/api/candidates/${id}/next-action`);$("#agentResult").innerHTML=`<div><b>下一步推荐</b><pre>${pretty(r)}</pre></div>`}
async function applyNextAction(id){await api(`/api/candidates/${id}/apply-next-action`,{method:"POST",body:"{}"});toast("已应用推荐下一步");await refresh();selectedCandidate=candidates.find(c=>c.id===id);renderCandidateDetail(selectedCandidate)}
async function groupScore(id){const text=$("#groupText").value.trim();if(!text){toast("请先粘贴群面记录或作业内容");return}$("#groupResult").innerHTML="<span class='muted'>AI 正在评分...</span>";await api(`/api/candidates/${id}/group-score`,{method:"POST",body:JSON.stringify({text,auto_workflow:true})});toast("群面评分完成，流程已自动同步");await refresh();selectedCandidate=candidates.find(c=>c.id===id);renderCandidateDetail(selectedCandidate)}
function renderKanban(){
  $("#kanban").innerHTML=PIPELINE.map(s=>`<div class="lane"><h3>${s}</h3>${candidates.filter(c=>c.stage===s).map(c=>`<div class="mini" onclick="document.querySelector('[data-view=candidates]').click();selectCandidate('${c.id}')"><b>${c.name}</b><span>${c.position||""} · ${c.match?.score??"未评分"}分</span></div>`).join("")}</div>`).join("");
}
function renderProfiles(){
  if(!selectedProfile) selectedProfile=profiles[0];
  $("#profileList").innerHTML=profiles.map(p=>`<button class="${selectedProfile?.id===p.id?"active":""}" onclick="selectedProfile=profiles.find(x=>x.id==='${p.id}');jdDraft=null;renderProfiles()"><b>${p.name}</b><div class="muted">通过线 ${p.pass_score||70} 分</div></button>`).join("");
  const p=selectedProfile||{id:"",name:"",description:"",required_skills:[],abilities:[],risk_keywords:[],pass_score:70};
  const skillText=(p.required_skills||[]).map(x=>`${x.name},${x.weight}`).join("\n");
  const abilityText=(p.abilities||[]).map(x=>`${x.name},${(x.keywords||[]).join("|")},${x.weight}`).join("\n");
  const totalWeight=[...(p.required_skills||[]),...(p.abilities||[])].reduce((s,x)=>s+Number(x.weight||0),0);
  $("#profileForm").innerHTML=`<div class="profile-panel card">
    <div class="profile-head">
      <div><h2>评分标准配置</h2><p>把岗位 JD 拆成技能、能力、风险和通过线，形成标准化多维评分规则，降低 HR 主观判断波动。</p></div>
      <div class="profile-stats" id="profileStats">
        <div class="stat"><b>${(p.required_skills||[]).length}</b><span>技能项</span></div>
        <div class="stat"><b>${(p.abilities||[]).length}</b><span>能力项</span></div>
        <div class="stat"><b>${totalWeight}</b><span>总权重</span></div>
      </div>
    </div>
    <div class="profile-form">
      <div class="form-grid">
        <div class="field"><label>岗位名称</label><input id="pName" value="${p.name||""}" placeholder="例如：AI算法实习生"></div>
        <div class="field"><label>通过分</label><input id="pPass" type="number" min="0" max="100" value="${p.pass_score||70}"></div>
      </div>
      <div class="field"><label>岗位说明 <span class="hint">用于 HR 快速理解岗位画像</span></label><textarea id="pDesc" class="textarea-short">${p.description||""}</textarea></div>
      <div class="config-box">
        <h3>智能 JD 生成</h3>
        <div class="hint">针对“语言模板化、岗位细节难体现”，可补充行业知识库/RAG上下文，系统会结合岗位字段生成更贴近业务的 JD。</div>
        <div class="jd-builder" style="margin-top:10px">
          <div class="jd-form">
            <div class="input-row">
              <div class="field"><label>职位名称</label><input id="jdTitle" value="${esc(p.name||"")}" placeholder="例如：软件开发"></div>
              <div class="field"><label>工作地点</label><input id="jdCity" placeholder="例如：上海"></div>
            </div>
            <div class="input-row">
              <div class="field"><label>部门</label><input id="jdDept" placeholder="例如：技术部"></div>
              <div class="field"><label>工作类型</label><select id="jdWorkType"><option>全职</option><option>实习</option><option>兼职</option><option>校招</option></select></div>
            </div>
            <div class="input-row">
              <div class="field"><label>薪资范围</label><input id="jdSalary" placeholder="例如：15K / 10-25K"></div>
              <div class="field"><label>学历要求</label><select id="jdEducation"><option>本科</option><option>硕士</option><option>大专</option><option>不限</option></select></div>
            </div>
            <div class="input-row">
              <div class="field"><label>经验要求</label><input id="jdExperience" placeholder="例如：5-10年软件开发经验"></div>
              <div class="field"><label>招聘人数</label><input id="jdHeadcount" type="number" min="1" value="1"></div>
            </div>
            <div class="field"><label>技能标签</label><input id="jdSkills" value="${esc((p.required_skills||[]).map(x=>x.name).join("、"))}" placeholder="Python、Angular、Vue.js、TypeScript"></div>
            <div class="field"><label>业务/项目背景</label><textarea id="jdBackground" class="textarea-short" placeholder="例如：负责保险经纪系统、后台管理平台、数据可视化等项目建设。"></textarea></div>
            <div class="field"><label>行业知识库/RAG上下文</label><textarea id="jdKnowledge" class="textarea-short" placeholder="粘贴行业知识、团队技术栈、业务术语、岗位真实项目、优秀JD片段，生成时会作为岗位细节依据。"></textarea></div>
            <div class="field"><label>公司介绍</label><textarea id="jdCompany" class="textarea-short" placeholder="例如：我们是一家快速成长的科技公司，重视工程质量和人才培养。"></textarea></div>
            <div class="field"><label>福利待遇/工作亮点</label><input id="jdHighlights" placeholder="例如：五险一金、弹性办公、技术成长、扁平协作"></div>
            <div class="jd-actions">
              <button class="btn blue" type="button" onclick="generateJD()">生成 JD</button>
              <button class="btn" type="button" onclick="applyJDToProfile()">应用到岗位标准</button>
            </div>
          </div>
          <div class="template jd-preview" id="jdPreview">生成后的 JD 会显示在这里。</div>
        </div>
      </div>
      <div class="score-blocks">
        <div class="config-box">
          <h3>核心技能权重</h3>
          <div class="hint">每行格式：技能,权重。用于硬技能匹配，权重越高，对匹配分影响越大。</div>
          <textarea id="pSkills" class="textarea-tall">${skillText}</textarea>
        </div>
        <div class="config-box">
          <h3>能力关键词</h3>
          <div class="hint">每行格式：能力,关键词1|关键词2,权重。</div>
          <textarea id="pAbilities" class="textarea-tall">${abilityText}</textarea>
        </div>
      </div>
      <div class="risk-row">
        <div class="field"><label>风险关键词 <span class="hint">命中后会进入风险点</span></label><input id="pRisks" value="${(p.risk_keywords||[]).join(",")}" placeholder="例如：无法到岗,不接受实习"></div>
        <div class="config-box"><h3>多指标评分提示</h3><div class="hint">建议总权重控制在 80-120；技能项用于硬匹配，能力项用于项目经历和协作信号，风险关键词会进入风险控制维度，候选人详情会展示各维度评分依据。</div></div>
      </div>
      <div class="save-bar">
        <div class="save-note">保存后立即用于候选人重新评分，并写入 data/config/job_profiles.json。</div>
        <button class="btn primary" onclick="saveProfile()">保存岗位标准</button>
      </div>
    </div>
  </div>`;
  ["pSkills","pAbilities"].forEach(id=>$("#"+id).addEventListener("input",updateProfileStats));
}
function updateProfileStats(){const skills=parseSkills($("#pSkills").value), abilities=parseAbilities($("#pAbilities").value);const total=[...skills,...abilities].reduce((s,x)=>s+Number(x.weight||0),0);$("#profileStats").innerHTML=`<div class="stat"><b>${skills.length}</b><span>技能项</span></div><div class="stat"><b>${abilities.length}</b><span>能力项</span></div><div class="stat"><b>${total}</b><span>总权重</span></div>`}
function parseSkills(v){return v.split(/\n+/).map(l=>l.trim()).filter(Boolean).map(l=>{const [name,weight]=l.split(",");return {name:name.trim(),weight:Number(weight||5)}})}
function parseAbilities(v){return v.split(/\n+/).map(l=>l.trim()).filter(Boolean).map(l=>{const parts=l.split(",");return {name:(parts[0]||"").trim(),keywords:(parts[1]||"").split("|").map(x=>x.trim()).filter(Boolean),weight:Number(parts[2]||5)}})}
function jdPayload(){return {title:$("#jdTitle").value.trim()||$("#pName").value.trim(),city:$("#jdCity").value.trim(),department:$("#jdDept").value.trim(),work_type:$("#jdWorkType").value, salary:$("#jdSalary").value.trim(),education:$("#jdEducation").value,experience:$("#jdExperience").value.trim(),headcount:Number($("#jdHeadcount").value||1),skills:$("#jdSkills").value.trim(),background:$("#jdBackground").value.trim(),knowledge:$("#jdKnowledge")?.value.trim()||"",company:$("#jdCompany").value.trim(),highlights:$("#jdHighlights").value.trim()}}
async function generateJD(){try{$("#jdPreview").textContent="正在生成 JD...";const r=await api("/api/jd-generate",{method:"POST",body:JSON.stringify(jdPayload())});jdDraft=r;$("#jdPreview").textContent=r.jd_text||"生成完成，但没有返回正文。";toast((r.method||"JD")+" 已生成")}catch(e){$("#jdPreview").textContent="生成失败："+e.message;toast("JD 生成失败："+e.message)}}
function applyJDToProfile(){if(!jdDraft){toast("请先生成 JD");return}const profile=jdDraft.profile||{};if(profile.name)$("#pName").value=profile.name;if(jdDraft.jd_text)$("#pDesc").value=jdDraft.jd_text;if(profile.pass_score)$("#pPass").value=profile.pass_score;if(profile.required_skills)$("#pSkills").value=profile.required_skills.map(x=>`${x.name},${x.weight||5}`).join("\n");if(profile.abilities)$("#pAbilities").value=profile.abilities.map(x=>`${x.name},${(x.keywords||[x.name]).join("|")},${x.weight||5}`).join("\n");if(profile.risk_keywords)$("#pRisks").value=profile.risk_keywords.join("，");updateProfileStats();toast("已应用到岗位标准，确认后点击保存")}
async function saveProfile(){const p={...(selectedProfile||{}),id:selectedProfile?.id||Date.now().toString(36),name:$("#pName").value,description:$("#pDesc").value,pass_score:Number($("#pPass").value||70),required_skills:parseSkills($("#pSkills").value),abilities:parseAbilities($("#pAbilities").value),risk_keywords:$("#pRisks").value.split(/[,，]/).map(x=>x.trim()).filter(Boolean)};await api("/api/job-profiles",{method:"POST",body:JSON.stringify(p)});toast("岗位标准已保存");selectedProfile=p;await refresh()}
function esc(v){return String(v??"").replace(/[&<>"']/g,m=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[m]))}
const T={online:"\u5728\u7ebf",failed:"\u5931\u8d25",untested:"\u672a\u6d4b\u8bd5",custom:"\u81ea\u5b9a\u4e49",env:"\u73af\u5883\u914d\u7f6e",current:"\u5f53\u524d"};
function renderModelList(){if(!appConfig)return;const q=(document.querySelector("#modelSearch")?.value||"").toLowerCase(), source=document.querySelector("#modelSourceFilter")?.value||"", status=document.querySelector("#modelStatusFilter")?.value||"";let models=(appConfig.model_configs||[]).filter(m=>(!source||m.source===source)&&(!status||m.status===status)&&(!q||JSON.stringify(m).toLowerCase().includes(q)));document.querySelector("#modelList").innerHTML=models.map(m=>{const available=(m.available_models||[]).slice(0,8);const result=m.last_test_message?`<div class="model-test-result ${m.last_test_ok===false?"bad":""}"><b>${esc(m.last_test_message)}</b><br>${available.length?`\u53ef\u7528\u6a21\u578b\uff1a${available.map(esc).join("\u3001")}${(m.available_models||[]).length>8?"...":""}`:(m.models_error?`\u6a21\u578b\u5217\u8868\u8bfb\u53d6\u5931\u8d25\uff1a${esc(m.models_error)}`:"\u5df2\u5b8c\u6210\u8fde\u901a\u6d4b\u8bd5")}</div>`:"";return `<div class="model-card ${m.id===appConfig.active_model_id?"active":""}"><div class="model-title"><h3>${esc(m.name||"\u672a\u547d\u540d\u6a21\u578b")}</h3><div><span class="badge ${m.status===T.online?"ok":""}">${esc(m.status||T.untested)}</span> ${m.id===appConfig.active_model_id?'<span class="badge active">\u5f53\u524d</span>':""}</div></div><div class="model-meta"><span class="model-label">\u6a21\u578b</span><span>${esc(m.model||"\u65e0")}</span><span class="model-label">API</span><span>${esc(m.base_url||"\u65e0")}</span><span class="model-label">Key</span><span>${m.has_api_key?"\u5df2\u4fdd\u5b58":esc(m.api_key_env||"\u65e0")}</span><span class="model-label">Temp</span><span>${esc(m.temperature??0.2)}</span><span class="model-label">Timeout</span><span>${esc(m.timeout||120)}s</span><span class="model-label">Thinking</span><span>${esc(m.thinking||"OpenAI")}</span></div>${result}<div class="model-actions"><button class="btn small blue" onclick="useModel('${esc(m.id)}')">\u8bbe\u4e3a\u5f53\u524d</button><button class="btn small" onclick="testModel('${esc(m.id)}')">\u6d4b\u8bd5\u8fde\u901a</button><button class="btn small ghost" onclick="editModel('${esc(m.id)}')">\u7f16\u8f91</button><button class="btn small" onclick="deleteModel('${esc(m.id)}')">\u5220\u9664</button></div></div>`}).join("")||"<div class='muted'>\u6682\u65e0\u6a21\u578b\u914d\u7f6e\u3002</div>"}
function modelModalHtml(){return `<div id="modelModal" class="modal-mask" onclick="if(event.target===this)closeModelModal()"><div class="modal-panel"><div class="modal-head"><h3 id="modelFormTitle">\u65b0\u589e\u6a21\u578b</h3><button class="icon-btn" onclick="closeModelModal()">X</button></div><div class="modal-body"><div class="model-form"><div class="model-form-grid"><div class="field"><label>\u663e\u793a\u540d\u79f0</label><input id="modelName" placeholder="\u4f8b\u5982\uff1aqwen3.6-27b(A100)"></div><div class="field"><label>\u6a21\u578b\u6807\u8bc6</label><input id="modelIdName" placeholder="\u4f8b\u5982\uff1aqwen36-27b"></div></div><div class="field"><label>API \u5730\u5740</label><input id="modelBaseUrl" placeholder="\u53ef\u586b http://host:port/v1"></div><div class="model-form-grid"><div class="field"><label>API Key</label><input id="modelApiKey" type="password" placeholder="\u672c\u5730\u6a21\u578b\u53ef\u7559\u7a7a\uff0c\u7f16\u8f91\u65f6\u7559\u7a7a\u8868\u793a\u4e0d\u4fee\u6539"></div><div class="field"><label>\u73af\u5883\u53d8\u91cf\u540d</label><input id="modelKeyEnv" value="OPENAI_API_KEY"></div></div><div class="model-form-grid"><div class="field"><label>Temperature</label><input id="modelTemp" type="number" step="0.1" value="0.2"></div><div class="field"><label>Timeout \u79d2</label><input id="modelTimeout" type="number" value="120"></div></div><div class="model-form-grid"><div class="field"><label>Thinking / Provider</label><input id="modelThinking" placeholder="OpenAI / DashScope / \u4e0d\u5173\u95ed"></div><div class="field"><label>\u6765\u6e90</label><select id="modelSource"><option>${T.custom}</option><option>${T.env}</option></select></div></div><div class="actions"><button class="btn primary" onclick="saveModel()">\u4fdd\u5b58\u6a21\u578b</button><button class="btn" onclick="closeModelModal()">\u53d6\u6d88</button></div></div></div></div></div>`}
function openModelModal(){document.querySelector("#modelModal")?.classList.add("open")}
function closeModelModal(){document.querySelector("#modelModal")?.classList.remove("open")}
function newModel(){editingModelId=null;openModelModal();["modelName","modelIdName","modelBaseUrl","modelApiKey"].forEach(id=>document.querySelector("#"+id).value="");document.querySelector("#modelKeyEnv").value="OPENAI_API_KEY";document.querySelector("#modelTemp").value="0.2";document.querySelector("#modelTimeout").value="120";document.querySelector("#modelThinking").value="OpenAI";document.querySelector("#modelSource").value=T.custom;document.querySelector("#modelFormTitle").textContent="\u65b0\u589e\u6a21\u578b"}
function fillModelForm(m){if(!m)return;openModelModal();document.querySelector("#modelFormTitle").textContent="\u7f16\u8f91\u6a21\u578b";document.querySelector("#modelName").value=m.name||"";document.querySelector("#modelIdName").value=m.model||"";document.querySelector("#modelBaseUrl").value=m.base_url||"";document.querySelector("#modelApiKey").value="";document.querySelector("#modelKeyEnv").value=m.api_key_env||"OPENAI_API_KEY";document.querySelector("#modelTemp").value=m.temperature??0.2;document.querySelector("#modelTimeout").value=m.timeout||120;document.querySelector("#modelThinking").value=m.thinking||"";document.querySelector("#modelSource").value=m.source||T.custom}
function editModel(id){editingModelId=id;fillModelForm((appConfig.model_configs||[]).find(m=>m.id===id))}
async function persistConfig(extra={}){const payload={llm_enabled:document.querySelector("#cfgEnabled")?document.querySelector("#cfgEnabled").value==="true":appConfig.llm_enabled,model_configs:appConfig.model_configs,active_model_id:appConfig.active_model_id,reminder_hours:Number(document.querySelector("#cfgReminder")?.value||appConfig.reminder_hours||24),manager_register_code:appConfig.manager_register_code,...currentPromptConfig(),workflow_rules:collectWorkflowRules(),mask_contact_for_hr:$("#cfgMaskContact")?$("#cfgMaskContact").value==="true":!!appConfig.mask_contact_for_hr,data_retention_days:Number($("#cfgRetentionDays")?.value||appConfig.data_retention_days||365),ui_theme:appConfig.ui_theme||"light",ui_primary_color:appConfig.ui_primary_color||"#126b61",ui_font_size:appConfig.ui_font_size||14,ui_font_family:appConfig.ui_font_family||"Microsoft YaHei",...extra};appConfig=await api("/api/app-config",{method:"POST",body:JSON.stringify(payload)});applyAppearance();return appConfig}
async function saveModel(){const id=editingModelId||Date.now().toString(36);const old=(appConfig.model_configs||[]).find(m=>m.id===id)||{};const model={id,name:document.querySelector("#modelName").value.trim()||document.querySelector("#modelIdName").value.trim()||"\u672a\u547d\u540d\u6a21\u578b",model:document.querySelector("#modelIdName").value.trim(),base_url:document.querySelector("#modelBaseUrl").value.trim(),api_key:document.querySelector("#modelApiKey").value.trim(),api_key_env:document.querySelector("#modelKeyEnv").value.trim()||"OPENAI_API_KEY",temperature:Number(document.querySelector("#modelTemp").value||0.2),timeout:Number(document.querySelector("#modelTimeout").value||120),thinking:document.querySelector("#modelThinking").value.trim()||"OpenAI",source:document.querySelector("#modelSource").value,status:old.status||T.untested,last_test:old.last_test||"",last_test_ok:old.last_test_ok,last_test_message:old.last_test_message||"",available_models:old.available_models||[],models_error:old.models_error||"",model_found:old.model_found,has_api_key:old.has_api_key};appConfig.model_configs=[...(appConfig.model_configs||[]).filter(m=>m.id!==id),model];if(!appConfig.active_model_id)appConfig.active_model_id=id;await persistConfig();toast("\u6a21\u578b\u5df2\u4fdd\u5b58");editingModelId=null;renderModelSettings()}
async function useModel(id){appConfig.active_model_id=id;await persistConfig();toast("\u5df2\u8bbe\u4e3a\u5f53\u524d\u6a21\u578b");renderModelSettings()}
async function deleteModel(id){if((appConfig.model_configs||[]).length<=1){toast("\u81f3\u5c11\u4fdd\u7559\u4e00\u4e2a\u6a21\u578b");return}appConfig.model_configs=(appConfig.model_configs||[]).filter(m=>m.id!==id);if(appConfig.active_model_id===id)appConfig.active_model_id=appConfig.model_configs[0]?.id;await persistConfig();toast("\u6a21\u578b\u5df2\u5220\u9664");editingModelId=null;renderModelSettings()}
async function testModel(id){toast("\u6b63\u5728\u6d4b\u8bd5\u6a21\u578b\u8fde\u901a");const r=await api("/api/model-test",{method:"POST",body:JSON.stringify({id})});toast(r.ok?"\u6a21\u578b\u8fde\u901a\u6210\u529f":"\u6a21\u578b\u8fde\u901a\u5931\u8d25\uff1a"+r.message);appConfig=await api("/api/app-config");renderModelSettings()}
async function saveSettings(){await persistConfig();toast("\u7cfb\u7edf\u8bbe\u7f6e\u5df2\u4fdd\u5b58");renderSettings()}
async function rerunWorkflowRules(){await persistConfig();const r=await api("/api/workflow-rerun",{method:"POST",body:"{}"});toast(`已重新检查 ${r.count} 位候选人，更新 ${r.changed} 位`);await refresh()}
function renderModelSettings(){if(!currentUser||currentUser.role!=="manager"||!appConfig)return;const c=appConfig||{};const models=c.model_configs||[];const activeId=c.active_model_id;const online=models.filter(m=>m.status===T.online).length;document.querySelector("#modelSettingsPanel").innerHTML=`<div class="profile-form"><div class="config-box"><h3>\u6a21\u578b\u603b\u89c8</h3><div class="model-stats"><div class="metric"><span class="muted">\u603b\u6a21\u578b\u6570</span><b>${models.length}</b></div><div class="metric"><span class="muted">\u5f53\u524d\u542f\u7528</span><b>${esc(models.find(m=>m.id===activeId)?.name||"\u65e0")}</b></div><div class="metric"><span class="muted">\u5728\u7ebf</span><b>${online}</b></div><div class="metric"><span class="muted">\u81ea\u5b9a\u4e49</span><b>${models.filter(m=>m.source===T.custom).length}</b></div></div><div class="model-tools"><input id="modelSearch" placeholder="\u641c\u7d22\u6a21\u578b\u540d\u79f0\u3001\u6a21\u578b\u6807\u8bc6\u3001API \u5730\u5740..." oninput="renderModelList()"><select id="modelSourceFilter" onchange="renderModelList()"><option value="">\u5168\u90e8\u6765\u6e90</option><option>${T.env}</option><option>${T.custom}</option></select><select id="modelStatusFilter" onchange="renderModelList()"><option value="">\u5168\u90e8\u72b6\u6001</option><option>${T.online}</option><option>${T.failed}</option><option>${T.untested}</option></select><button class="btn primary" onclick="newModel()">+ \u6dfb\u52a0\u6a21\u578b</button></div><div id="modelList" class="model-list"></div></div><div class="config-box"><h3>\u8c03\u7528\u89c4\u5219</h3><div class="form-grid"><div class="field"><label>\u542f\u7528\u5927\u6a21\u578b</label><select id="cfgEnabled"><option value="false">\u5173\u95ed\uff0c\u4f7f\u7528\u672c\u5730\u515c\u5e95</option><option value="true" ${c.llm_enabled?"selected":""}>\u5f00\u542f\uff0c\u4f7f\u7528\u5f53\u524d\u6a21\u578b</option></select></div><div class="field"><label>\u63d0\u9192\u5c0f\u65f6\u6570</label><input id="cfgReminder" type="number" min="1" value="${c.reminder_hours||24}"></div></div>${workflowRuleSettings(c.workflow_rules||{})}${complianceSettings(c)}<div class="save-bar"><div class="save-note">\u8c03\u7528\u89c4\u5219\u4fdd\u5b58\u540e\u4f1a\u5f71\u54cd AI \u89e3\u6790\u3001\u8bc4\u5ba1\u3001\u81ea\u52a8\u63a8\u8fdb\u3001\u6570\u636e\u5408\u89c4\u548c\u63d0\u9192\u3002</div><button class="btn primary" onclick="persistConfig().then(()=>toast('\\u8c03\\u7528\\u89c4\\u5219\\u5df2\\u4fdd\\u5b58'))">\u4fdd\u5b58\u8c03\u7528\u89c4\u5219</button></div></div></div>${modelModalHtml()}`;renderModelList()}
function workflowRuleSettings(r={}){const v=(k,d)=>r[k]??d;return `<div class="config-box" style="margin-top:12px"><h3>流程分数线</h3><div class="hint">这些分数线控制自动流转：简历、群面、作业和 HR 面达到对应分数后自动进入下一环节。</div><div class="model-form-grid" style="margin-top:10px"><div class="field"><label>简历推进群面线</label><input id="ruleResumePass" type="number" min="0" max="100" value="${v("resume_pass_score",70)}"></div><div class="field"><label>简历待观察线</label><input id="ruleResumeWatch" type="number" min="0" max="100" value="${v("resume_watch_score",55)}"></div><div class="field"><label>群面推进作业线</label><input id="ruleGroupPass" type="number" min="0" max="100" value="${v("group_pass_score",70)}"></div><div class="field"><label>作业推进终面线</label><input id="ruleAssignmentPass" type="number" min="0" max="100" value="${v("assignment_pass_score",70)}"></div><div class="field"><label>HR面通过线</label><input id="ruleHrPass" type="number" min="0" max="100" value="${v("hr_pass_score",70)}"></div></div><div class="actions" style="margin-top:10px"><button class="btn small ghost" onclick="rerunWorkflowRules()">重新套用到现有候选人</button></div></div>`}
function complianceSettings(c={}){return `<div class="config-box" style="margin-top:12px"><h3>数据合规与权限</h3><div class="hint">控制普通 HR 的候选人数据可见范围和简历数据保留周期，适合多人共同使用但账号权限独立的场景。</div><div class="model-form-grid" style="margin-top:10px"><div class="field"><label>普通 HR 联系方式脱敏</label><select id="cfgMaskContact"><option value="false" ${!c.mask_contact_for_hr?"selected":""}>关闭，完整展示</option><option value="true" ${c.mask_contact_for_hr?"selected":""}>开启，仅管理员完整可见</option></select></div><div class="field"><label>数据保留天数</label><input id="cfgRetentionDays" type="number" min="30" value="${c.data_retention_days||365}"></div></div><div class="hint">开启脱敏后，非管理员通过候选人接口看到的手机号和邮箱会被遮盖；管理员仍可查看完整数据并导出。</div></div>`}
function promptFields(){return [
  ["resume_parse_prompt","cfgParsePrompt","AI解析简历","要求模型结构化解析简历，不补充简历没有的信息。"],
  ["candidate_review_prompt","cfgReviewPrompt","AI综合评审","用于岗位适配度、匹配证据、风险点、追问问题和推进建议。"],
  ["jd_generate_prompt","cfgJdPrompt","JD生成","用于智能 JD 生成和岗位评分标准抽取，可约束行业知识库/RAG上下文的使用方式。"],
  ["candidate_rescore_prompt","cfgRescorePrompt","简历AI复核评分","用于规则评分后的模型复核，要求输出 score、strengths、risks、questions、suggestion。"],
  ["next_action_prompt","cfgNextActionPrompt","流程下一步推荐","用于根据候选人当前阶段生成下一步动作、原因、话术和建议阶段。"],
  ["group_score_prompt","cfgGroupScorePrompt","群面/作业评分","用于根据群面记录、面试官评价或作业内容生成维度分、总分和决策建议。"],
  ["interview_plan_prompt","cfgInterviewPlanPrompt","智能面试方案","用于根据不同简历生成定制化面试问题、项目深挖、评分 Rubric 和风险验证。"],
  ["daily_report_prompt","cfgDailyPrompt","AI招聘日报","用于生成今日概览、风险、优先跟进、建议动作和流程卡点。"],
  ["training_exam_prompt","cfgTrainingExamPrompt","试卷生成","用于根据培训对象、主题、难度和技能点生成试卷、题目、答案和评分规则。"],
  ["exam_grading_prompt","cfgExamGradingPrompt","考试阅卷评分","用于根据试卷标准答案和考生答案评分，并标记是否需要人工复核。"],
]}
function currentPromptConfig(){const out={};promptFields().forEach(([key])=>out[key]=appConfig?.[key]||"");return out}
function renderPromptSettings(){if(!currentUser||currentUser.role!=="manager"||!appConfig)return;const c=appConfig||{};$("#promptSettingsPanel").innerHTML=`<div class="profile-form"><div class="config-box"><h3>提示词管理</h3><div class="hint">以下功能都会调用模型。修改后保存到 data/config/app_config.json，下一次调用对应智能体立即生效。请尽量保留要求输出 JSON 的字段约束。</div></div>${promptFields().map(([key,id,title,hint])=>`<div class="config-box"><h3>${esc(title)}</h3><div class="field"><label>${esc(title)}提示词</label><textarea id="${id}" class="textarea-tall">${esc(c[key]||"")}</textarea></div><div class="hint">${esc(hint)}</div></div>`).join("")}<div class="save-bar"><div class="save-note">保存后会影响 JD、简历解析、候选评审、流程推荐、群面评分、面试方案、日报、试卷生成和考试阅卷。</div><button class="btn primary" onclick="savePromptSettings()">保存全部提示词</button></div></div>`}
function renderSettings(){if(!currentUser)return;const c=appConfig||{};const globalSettings=currentUser.role==="manager"?`<div class="config-box"><h3>全局外观</h3><div class="form-grid"><div class="field"><label>主题模式</label><select id="cfgTheme"><option value="light" ${c.ui_theme!=="dark"?"selected":""}>浅色模式</option><option value="dark" ${c.ui_theme==="dark"?"selected":""}>深色模式</option></select></div><div class="field"><label>主色</label><input id="cfgPrimaryColor" type="color" value="${c.ui_primary_color||"#126b61"}"></div></div><div class="form-grid"><div class="field"><label>字体</label><select id="cfgFontFamily"><option ${c.ui_font_family==="Microsoft YaHei"?"selected":""}>Microsoft YaHei</option><option ${c.ui_font_family==="SimHei"?"selected":""}>SimHei</option><option ${c.ui_font_family==="Arial"?"selected":""}>Arial</option></select></div><div class="field"><label>字号</label><input id="cfgFontSize" type="number" min="12" max="18" value="${c.ui_font_size||14}"></div></div><div class="save-bar"><div class="save-note">全局外观配置会保存到 data/config/app_config.json。</div><button class="btn primary" onclick="saveBasicSettings()">保存外观设置</button></div></div>`:"";$("#settingsPanel").innerHTML=`<div class="profile-form"><div class="config-box"><div class="settings-head">${avatarHtml(currentUser,"large")}<div><h3>个人账户</h3><div class="hint">维护头像、账号、联系方式和安全密码。</div></div></div><div class="form-grid"><div class="field"><label>用户名</label><input id="acctUsername" value="${esc(currentUser.username||"")}"></div><div class="field"><label>显示名称</label><input id="acctDisplay" value="${esc(currentUser.display_name||"")}"></div></div><div class="form-grid"><div class="field"><label>部门</label><input id="acctDepartment" value="${esc(currentUser.department||"")}" placeholder="例如：招聘组"></div><div class="field"><label>职位/身份</label><input id="acctTitle" value="${esc(currentUser.title||"")}" placeholder="例如：高级 HR"></div></div><div class="form-grid"><div class="field"><label>手机号</label><input id="acctPhone" value="${esc(currentUser.phone||"")}"></div><div class="field"><label>邮箱</label><input id="acctEmail" value="${esc(currentUser.email||"")}"></div></div><div class="form-grid"><div class="field"><label>头像图片地址</label><input id="acctAvatarUrl" value="${esc(currentUser.avatar_url||"")}" placeholder="可填图片 URL，留空使用首字头像"></div><div class="field"><label>头像底色</label><input id="acctAvatarColor" type="color" value="${currentUser.avatar_color||"#126b61"}"></div></div><div class="field"><label>个人简介</label><textarea id="acctBio" class="textarea-short" placeholder="例如：负责 AI/算法实习生招聘、群面组织和候选人跟进">${esc(currentUser.bio||"")}</textarea></div><div class="save-bar"><div class="save-note">资料会保存到 data/config/users.json。</div><button class="btn primary" onclick="saveAccountProfile()">保存个人资料</button></div></div><div class="config-box"><h3>修改密码</h3><div class="form-grid"><div class="field"><label>原密码</label><input id="acctOldPassword" type="password" placeholder="修改密码时必填"></div><div class="field"><label>新密码</label><input id="acctNewPassword" type="password" placeholder="至少 6 位"></div></div><div class="save-bar"><div class="save-note">保存后下次登录使用新密码。</div><button class="btn primary" onclick="saveAccountPassword()">更新密码</button></div></div>${globalSettings}</div>`}
function collectWorkflowRules(){const current=appConfig?.workflow_rules||{};const read=(id,fallback)=>{const el=$("#"+id);const n=Number(el?.value);return Number.isFinite(n)?Math.max(0,Math.min(100,Math.round(n))):fallback};return {resume_pass_score:read("ruleResumePass",current.resume_pass_score??70),resume_watch_score:read("ruleResumeWatch",current.resume_watch_score??55),group_pass_score:read("ruleGroupPass",current.group_pass_score??70),assignment_pass_score:read("ruleAssignmentPass",current.assignment_pass_score??70),hr_pass_score:read("ruleHrPass",current.hr_pass_score??70)}}
async function persistConfig(extra={}){const payload={llm_enabled:$("#cfgEnabled")?$("#cfgEnabled").value==="true":appConfig.llm_enabled,model_configs:appConfig.model_configs,active_model_id:appConfig.active_model_id,reminder_hours:Number($("#cfgReminder")?.value||appConfig.reminder_hours||24),manager_register_code:appConfig.manager_register_code,...currentPromptConfig(),workflow_rules:collectWorkflowRules(),mask_contact_for_hr:$("#cfgMaskContact")?$("#cfgMaskContact").value==="true":!!appConfig.mask_contact_for_hr,data_retention_days:Number($("#cfgRetentionDays")?.value||appConfig.data_retention_days||365),ui_theme:appConfig.ui_theme||"light",ui_primary_color:appConfig.ui_primary_color||"#126b61",ui_font_size:appConfig.ui_font_size||14,ui_font_family:appConfig.ui_font_family||"Microsoft YaHei",...extra};appConfig=await api("/api/app-config",{method:"POST",body:JSON.stringify(payload)});applyAppearance();return appConfig}
function collectPromptSettings(){const out={};promptFields().forEach(([key,id])=>{const el=$("#"+id);out[key]=el?el.value:(appConfig?.[key]||"")});return out}
async function savePromptSettings(){await persistConfig(collectPromptSettings());toast("提示词已保存");renderPromptSettings()}
async function saveAccountProfile(){const r=await api("/api/account",{method:"POST",body:JSON.stringify({username:$("#acctUsername").value.trim(),display_name:$("#acctDisplay").value.trim(),department:$("#acctDepartment").value.trim(),title:$("#acctTitle").value.trim(),phone:$("#acctPhone").value.trim(),email:$("#acctEmail").value.trim(),avatar_url:$("#acctAvatarUrl").value.trim(),avatar_color:$("#acctAvatarColor").value,bio:$("#acctBio").value.trim()})});currentUser=r.user;renderUserbar();renderSettings();toast("个人资料已保存")}
async function saveAccountPassword(){const oldp=$("#acctOldPassword").value,newp=$("#acctNewPassword").value;if(!newp){toast("请输入新密码");return}const r=await api("/api/account",{method:"POST",body:JSON.stringify({old_password:oldp,new_password:newp})});currentUser=r.user;$("#acctOldPassword").value="";$("#acctNewPassword").value="";toast("密码已更新")}
async function saveBasicSettings(){await persistConfig({ui_theme:$("#cfgTheme").value,ui_primary_color:$("#cfgPrimaryColor").value,ui_font_size:Number($("#cfgFontSize").value||14),ui_font_family:$("#cfgFontFamily").value});toast("外观设置已保存");renderSettings()}
async function saveSettings(){await saveBasicSettings()}
function showCollectGuide(){const box=$("#collectGuide");box.style.display=box.style.display==="none"?"block":"none"}
async function collectSelectedSource(){const source=$("#collectSource").value;const cfg=(collectionSources||[]).find(s=>s.id===source);if(!cfg||!cfg.api_url||source==="local"){openResumePicker(source);return}$("#collectResult").textContent="正在通过 API 采集...";try{const r=await api("/api/resume-api-collect",{method:"POST",body:JSON.stringify({source})});$("#collectResult").textContent=`已从${r.source||r.api_source||cfg.name}采集 ${r.saved_count||0} 份，候选人 ${r.scan?.total||0} 人`;toast("API 采集完成");await refresh()}catch(e){$("#collectResult").textContent="API 采集失败，可改用上传/粘贴："+e.message;toast("API 采集失败")}}
function collectFromPlatform(source){$("#collectSource").value=source;collectSelectedSource()}
function openResumePicker(source){if(source)$("#collectSource").value=source;$("#resumeFiles").click()}
function fileToBase64(file){return new Promise((resolve,reject)=>{const r=new FileReader();r.onload=()=>resolve(String(r.result).split(",")[1]||"");r.onerror=reject;r.readAsDataURL(file)})}
async function uploadResumeFiles(files){files=[...(files||[])];if(!files.length)return;$("#collectResult").textContent="正在导入文件...";const payload={source:$("#collectSource").value,files:[]};for(const file of files){payload.files.push({name:file.name,content:await fileToBase64(file)})}const r=await api("/api/resume-collect",{method:"POST",body:JSON.stringify(payload)});$("#resumeFiles").value="";$("#collectResult").textContent=`已从${r.source}导入 ${r.saved_count} 份，候选人 ${r.scan?.total||0} 人`;toast(`简历采集完成：${r.saved_count} 份`);await refresh()}
async function importPlatformText(){const text=$("#platformResumeText").value.trim();if(!text){toast("请先粘贴候选人资料");return}$("#collectResult").textContent="正在导入文本...";const name=($("#collectSource").selectedOptions[0]?.textContent||"平台")+"候选人.txt";const r=await api("/api/resume-collect",{method:"POST",body:JSON.stringify({source:$("#collectSource").value,texts:[{name,text}]})});$("#collectResult").textContent=`已从${r.source}导入 ${r.saved_count} 条文本，候选人 ${r.scan?.total||0} 人`;$("#platformResumeText").value="";toast("平台候选人已导入");await refresh()}
document.querySelectorAll("nav button").forEach(b=>b.onclick=()=>{document.querySelectorAll("nav button,.view").forEach(x=>x.classList.remove("active"));b.classList.add("active");$("#"+b.dataset.view).classList.add("active");$("#title").textContent=b.textContent});
$("#scanBtn").onclick=async()=>{toast("正在扫描并同步简历文件夹");const r=await api("/api/scan",{method:"POST",body:JSON.stringify({remove_missing:true})});const fallback=(r.failed||[]).filter(x=>x.imported_from_filename).length;toast(`扫描完成：${r.scanned} 个文件，候选人 ${r.total} 人${r.removed_count?`，移除 ${r.removed_count} 条已删文件记录`:""}${fallback?`，${fallback} 份先按文件名导入`:""}`);await refresh()};
$("#scoreAllBtn").onclick=async()=>{await api("/api/score-all",{method:"POST",body:JSON.stringify({profile_id:selectedProfile?.id})});toast("批量评分完成");await refresh()};
$("#exportBtn").onclick=()=>location.href="/api/export/candidates.csv";
$("#search").oninput=renderCandidates; $("#stageFilter").onchange=renderCandidates;
$("#newProfileBtn").onclick=()=>{jdDraft=null;selectedProfile={id:"",name:"新岗位",description:"",required_skills:[],abilities:[],risk_keywords:[],pass_score:70};renderProfiles()};
$("#loginTab").onclick=()=>setAuthMode("login");$("#registerTab").onclick=()=>setAuthMode("register");
function setAuthMode(mode){authMode=mode;$("#loginTab").classList.toggle("active",mode==="login");$("#registerTab").classList.toggle("active",mode==="register");$("#authSubmit").textContent=mode==="login"?"登录":"注册";$("#authDisplay").style.display=mode==="register"?"block":"none";$("#roleRow").style.display=mode==="register"?"grid":"none"}
$("#authSubmit").onclick=async()=>{try{const payload={username:$("#authUsername").value.trim(),password:$("#authPassword").value,display_name:$("#authDisplay").value.trim(),role:$("#authRole").value,manager_code:$("#managerCode").value.trim()};if(authMode==="register"){await api("/api/register",{method:"POST",body:JSON.stringify(payload)});toast("注册成功，请登录");setAuthMode("login");return}const r=await api("/api/login",{method:"POST",body:JSON.stringify(payload)});currentUser=r.user;showAuth(false);renderUserbar();await refresh()}catch(e){toast(e.message)}};
document.addEventListener("click",()=>$("#userbar")?.classList.remove("open"));
async function switchUser(){await logout(true)}
async function logout(clearForm=false){await api("/api/logout",{method:"POST",body:"{}"});currentUser=null;adminOverview=null;appConfig=null;$("#userbar").style.display="none";["adminNav","modelSettingsNav","promptSettingsNav","settingsNav"].forEach(id=>$("#"+id).style.display="none");if(clearForm){setAuthMode("login");$("#authUsername").value="";$("#authPassword").value="";$("#authDisplay").value="";$("#managerCode").value=""}showAuth(true)}
init().catch(e=>toast(e.message));
</script>
</body></html>"""


class Handler(SimpleHTTPRequestHandler):
    def log_message(self, fmt, *args):
        if sys.stderr:
            sys.stderr.write("%s - %s\n" % (self.address_string(), fmt % args))

    def current_user(self):
        cookie = self.headers.get("Cookie", "")
        token = ""
        for part in cookie.split(";"):
            if part.strip().startswith("hr_session="):
                token = part.strip().split("=", 1)[1]
        return get_user_by_session(token)

    def require_user(self):
        user = self.current_user()
        if not user:
            self.send_json({"error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
            return None
        return user

    def require_manager(self):
        user = self.require_user()
        if not user:
            return None
        if user.get("role") != "manager":
            self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
            return None
        return user

    def send_json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_body(self):
        length = int(self.headers.get("Content-Length", "0") or 0)
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/":
                body = HTML.encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            elif parsed.path == "/api/me":
                user = self.current_user()
                self.send_json({"user": public_user(user) if user else None})
            elif parsed.path == "/api/admin/overview":
                if not self.require_manager():
                    return
                self.send_json(manager_overview())
            elif parsed.path == "/api/app-config":
                if not self.require_manager():
                    return
                self.send_json(public_app_config())
            elif parsed.path == "/api/collection-sources":
                user = self.require_user()
                if not user:
                    return
                self.send_json([public_collection_source(s) for s in normalize_user_collection_sources(user) if s.get("enabled", True)])
            elif parsed.path == "/api/training":
                user = self.require_user()
                if not user:
                    return
                self.send_json({"exams": load_training_exams(), "results": load_training_results(), "summary": training_summary()})
            elif parsed.path == "/api/candidates":
                user = self.require_user()
                if not user:
                    return
                self.send_json(candidates_for_user(load_candidates(), user))
            elif parsed.path == "/api/summary":
                if not self.require_user():
                    return
                self.send_json(build_summary(load_candidates()))
            elif parsed.path == "/api/daily-report":
                if not self.require_user():
                    return
                candidates = load_candidates()
                self.send_json(RecruitingAgent().daily_report(candidates))
            elif parsed.path == "/api/job-profiles":
                if not self.require_user():
                    return
                self.send_json(load_job_profiles())
            elif re.match(r"^/api/candidates/[^/]+/templates$", parsed.path):
                if not self.require_user():
                    return
                cid = parsed.path.split("/")[3]
                candidate = find_candidate(load_candidates(), cid)
                if not candidate:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json(build_templates(candidate))
            elif re.match(r"^/api/candidates/[^/]+/next-action$", parsed.path):
                if not self.require_user():
                    return
                cid = parsed.path.split("/")[3]
                candidate = find_candidate(load_candidates(), cid)
                if not candidate:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                self.send_json(RecruitingAgent().recommend_next_action(candidate))
            elif parsed.path == "/api/export/candidates.csv":
                if not self.require_user():
                    return
                candidates = load_candidates()
                export_candidates_csv(candidates)
                path = EXPORT_DIR / "candidates.csv"
                body = path.read_bytes()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "text/csv; charset=utf-8")
                self.send_header("Content-Disposition", "attachment; filename=candidates.csv")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            else:
                self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_json({"error": str(exc), "trace": traceback.format_exc()}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            body = self.read_body()
            if parsed.path == "/api/register":
                try:
                    user = create_user(
                        body.get("username", ""),
                        body.get("password", ""),
                        body.get("role", "hr"),
                        body.get("display_name", ""),
                        body.get("manager_code", ""),
                    )
                    self.send_json({"user": user})
                except ValueError as exc:
                    self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
                return
            if parsed.path == "/api/login":
                user = verify_user(body.get("username", ""), body.get("password", ""))
                if not user:
                    self.send_json({"error": "用户名或密码错误"}, HTTPStatus.UNAUTHORIZED)
                    return
                token = create_session(user)
                payload = json.dumps({"user": public_user(user)}, ensure_ascii=False).encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Set-Cookie", f"hr_session={token}; Path=/; HttpOnly; SameSite=Lax")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
                return
            if parsed.path == "/api/logout":
                cookie = self.headers.get("Cookie", "")
                token = ""
                for part in cookie.split(";"):
                    if part.strip().startswith("hr_session="):
                        token = part.strip().split("=", 1)[1]
                clear_session(token)
                payload = b'{"ok":true}'
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Set-Cookie", "hr_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
                return

            user = self.require_user()
            if not user:
                return
            actor = user.get("username")

            if parsed.path == "/api/account":
                try:
                    updated = update_user_account(user.get("id"), body)
                    self.send_json({"user": updated})
                except ValueError as exc:
                    self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
                return

            if parsed.path == "/api/workflow-rerun":
                if user.get("role") != "manager":
                    self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
                    return
                candidates = load_candidates()
                changed = 0
                rules = load_workflow_rules()
                for c in candidates:
                    if apply_workflow_automation(c, "rules_rerun", rules):
                        changed += 1
                    c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("workflow_rules_rerun", None, {"count": len(candidates), "changed": changed, "rules": rules}, actor=actor)
                self.send_json({"ok": True, "count": len(candidates), "changed": changed})
                return

            if parsed.path == "/api/scan":
                result = scan_resumes(actor=actor, remove_missing=body.get("remove_missing", True))
                append_event("scan_resumes", None, result, actor=actor)
                self.send_json(result)
            elif parsed.path == "/api/resume-collect":
                result = save_collected_resumes(body, actor=actor, user=user)
                self.send_json(result)
            elif parsed.path == "/api/resume-api-collect":
                result = collect_resumes_from_api(body.get("source"), actor=actor, user=user)
                status = HTTPStatus.OK if result.get("ok", True) else HTTPStatus.BAD_REQUEST
                self.send_json(result, status)
            elif parsed.path == "/api/jd-generate":
                result = generate_job_description(body)
                append_event("jd_generated", None, {"title": result.get("title"), "method": result.get("method")}, actor=actor)
                self.send_json(result)
            elif parsed.path == "/api/training/exams/generate":
                result = generate_training_exam(body)
                append_event("training_exam_generated", None, {"title": result.get("title"), "method": result.get("method")}, actor=actor)
                self.send_json(result)
            elif parsed.path == "/api/training/results":
                result = upsert_training_result(body, actor=actor)
                append_event("training_result_saved", None, {"exam": result.get("exam_title"), "student": result.get("student_name"), "score": result.get("score")}, actor=actor)
                self.send_json(result)
            elif parsed.path == "/api/training/results/grade":
                try:
                    result = grade_training_answers(body, actor=actor)
                    append_event("training_result_graded", None, {"exam": result.get("exam_title"), "student": result.get("student_name"), "score": result.get("score")}, actor=actor)
                    self.send_json(result)
                except ValueError as exc:
                    self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            elif parsed.path == "/api/score-all":
                profile = pick_profile(body.get("profile_id"))
                candidates = load_candidates()
                for c in candidates:
                    c["match"] = score_candidate(c, profile)
                    c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("score_all", None, {"profile": profile.get("name")}, actor=actor)
                self.send_json({"ok": True, "count": len(candidates)})
            elif re.match(r"^/api/candidates/[^/]+/score$", parsed.path):
                cid = parsed.path.split("/")[3]
                profile = pick_profile(body.get("profile_id"))
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                result = score_candidate(c, profile)
                llm = try_llm_rescore(c, profile)
                if llm and not llm.get("error"):
                    result["llm_review"] = llm
                    result["method"] = "规则匹配 + 大模型复评"
                elif llm and llm.get("error"):
                    result["llm_error"] = llm["error"]
                c["match"] = result
                if body.get("auto_workflow", True):
                    apply_workflow_automation(c, "resume_score")
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("candidate_scored", cid, {"profile": profile.get("name"), "score": result.get("score")}, actor=actor)
                self.send_json(c)
            elif re.match(r"^/api/candidates/[^/]+/agent-review$", parsed.path):
                cid = parsed.path.split("/")[3]
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                review = RecruitingAgent().review_candidate(c, body.get("profile_id"))
                c["agent_review"] = review
                c["match"] = {
                    **c.get("match", {}),
                    "score": review.get("score", c.get("match", {}).get("score", 0)),
                    "strengths": review.get("strengths", c.get("match", {}).get("strengths", [])),
                    "risks": review.get("risks", c.get("match", {}).get("risks", [])),
                    "questions": review.get("questions", c.get("match", {}).get("questions", [])),
                    "suggestion": review.get("decision", c.get("match", {}).get("suggestion", "")),
                    "method": review.get("method", "Agent评审"),
                    "scored_at": now_iso(),
                }
                if body.get("auto_workflow", True):
                    apply_workflow_automation(c, "agent_review")
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("agent_review", cid, {"decision": review.get("decision"), "score": review.get("score")}, actor=actor)
                self.send_json(c)
            elif re.match(r"^/api/candidates/[^/]+/analyze-resume$", parsed.path):
                cid = parsed.path.split("/")[3]
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                parsed_resume = RecruitingAgent().analyze_resume(c, body.get("profile_id"))
                c["ai_parse"] = parsed_resume
                if parsed_resume.get("target_position"):
                    c["position"] = parsed_resume.get("target_position") or c.get("position")
                if parsed_resume.get("skills"):
                    c["skills"] = parsed_resume.get("skills")
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("resume_analyzed", cid, {"method": parsed_resume.get("method")}, actor=actor)
                self.send_json(c)
            elif re.match(r"^/api/candidates/[^/]+/group-score$", parsed.path):
                cid = parsed.path.split("/")[3]
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                result = RecruitingAgent().score_group_interview(c, body.get("text", ""))
                c["group_ai_review"] = result
                c["group_interview"] = {
                    **c.get("group_interview", {}),
                    "status": "已评分",
                    "score": result.get("group_score", ""),
                }
                if body.get("auto_workflow", True):
                    apply_workflow_automation(c, "group_score")
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("group_scored", cid, {"score": result.get("group_score"), "decision": result.get("decision")}, actor=actor)
                self.send_json(c)
            elif re.match(r"^/api/candidates/[^/]+/interview-plan$", parsed.path):
                cid = parsed.path.split("/")[3]
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                result = RecruitingAgent().interview_plan(c, body.get("profile_id"))
                c["interview_plan"] = result
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("interview_plan_generated", cid, {"method": result.get("method")}, actor=actor)
                self.send_json(c)
            elif re.match(r"^/api/candidates/[^/]+/apply-next-action$", parsed.path):
                cid = parsed.path.split("/")[3]
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                recommendation = RecruitingAgent().recommend_next_action(c)
                if recommendation.get("suggested_stage"):
                    c["stage"] = normalize_stage(recommendation.get("suggested_stage"))
                c["workflow"] = {
                    **c.get("workflow", {}),
                    "auto_enabled": True,
                    "last_trigger": "apply_next_action",
                    "last_reason": recommendation.get("reason") or recommendation.get("action") or "应用下一步推荐",
                    "next_action": workflow_next_action(c.get("stage")),
                    "updated_at": now_iso(),
                }
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("workflow_next_action_applied", cid, {"stage": c.get("stage"), "recommendation": recommendation}, actor=actor)
                self.send_json(c)
            elif re.match(r"^/api/candidates/[^/]+$", parsed.path):
                cid = parsed.path.split("/")[3]
                candidates = load_candidates()
                c = find_candidate(candidates, cid)
                if not c:
                    self.send_json({"error": "candidate not found"}, HTTPStatus.NOT_FOUND)
                    return
                for key in [
                    "stage", "notes", "final_result", "expected_salary", "arrival_time",
                    "job_intention", "location_preference", "salary_expectation",
                    "availability", "stability_risk", "owner", "talent_pool", "talent_tags",
                    "next_follow_time", "follow_up_priority",
                ]:
                    if key in body:
                        c[key] = body[key]
                if isinstance(body.get("communication_logs"), list):
                    c["communication_logs"] = body["communication_logs"]
                if isinstance(body.get("interviewer_reviews"), list):
                    c["interviewer_reviews"] = body["interviewer_reviews"]
                for key in ["group_interview", "assignment", "hr_interview"]:
                    if key in body and isinstance(body[key], dict):
                        c[key] = body[key]
                if body.get("auto_workflow"):
                    apply_workflow_automation(c, "manual_save")
                c["updated_at"] = now_iso()
                save_candidates(candidates)
                append_event("candidate_updated", cid, body, actor=actor)
                self.send_json(c)
            elif parsed.path == "/api/job-profiles":
                profiles = load_job_profiles()
                profile = body
                if not profile.get("id"):
                    profile["id"] = hashlib.sha1((profile.get("name", "") + now_iso()).encode("utf-8")).hexdigest()[:10]
                profiles = [p for p in profiles if p.get("id") != profile["id"]]
                profiles.append(profile)
                write_json(JOB_PROFILES_FILE, profiles)
                append_event("job_profile_saved", None, {"profile": profile.get("name")}, actor=actor)
                self.send_json(profile)
            elif parsed.path == "/api/app-config":
                if user.get("role") != "manager":
                    self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
                    return
                config = save_app_config_from_payload(body)
                append_event("app_config_saved", None, {"llm_enabled": config.get("llm_enabled"), "llm_model": config.get("llm_model")}, actor=actor)
                self.send_json(config)
            elif parsed.path == "/api/model-test":
                if user.get("role") != "manager":
                    self.send_json({"error": "forbidden"}, HTTPStatus.FORBIDDEN)
                    return
                config = load_app_config()
                ensure_model_configs(config)
                model = find_model_config(config.get("model_configs", []), body.get("id"))
                if not model:
                    self.send_json({"error": "model not found"}, HTTPStatus.NOT_FOUND)
                    return
                result = test_model_config(model)
                model["status"] = result.get("status", "失败")
                model["last_test"] = result.get("time", now_iso())
                model["last_test_ok"] = result.get("ok", False)
                model["last_test_message"] = result.get("message", "")
                model["available_models"] = result.get("available_models", [])
                model["model_found"] = result.get("model_found", False)
                model["models_error"] = result.get("models_error", "")
                write_json(APP_CONFIG_FILE, config)
                append_event("model_tested", None, {"model": model.get("name"), "status": model.get("status")}, actor=actor)
                self.send_json(result)
            else:
                self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_json({"error": str(exc), "trace": traceback.format_exc()}, HTTPStatus.INTERNAL_SERVER_ERROR)


def find_candidate(candidates, cid):
    return next((c for c in candidates if c.get("id") == cid), None)


def pick_profile(profile_id=None):
    profiles = load_job_profiles()
    if profile_id:
        for profile in profiles:
            if profile.get("id") == profile_id:
                return profile
    return profiles[0]


def bootstrap():
    ensure_dirs()
    load_job_profiles()
    load_app_config()
    if not CANDIDATES_FILE.exists():
        scan_resumes()
    else:
        load_candidates()


def main():
    bootstrap()
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "11011"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"{APP_NAME} 已启动: http://{host}:{port}")
    print(f"数据目录: {DATA_DIR}")
    print("按 Ctrl+C 停止服务")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务已停止")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
